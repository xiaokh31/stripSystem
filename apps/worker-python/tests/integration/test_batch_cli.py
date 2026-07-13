from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from worker_python.batch import run_batch
from worker_python.cli import app


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
TEMPLATE_PATH = REPO_ROOT / "samples" / "templates" / "卸柜报告-En.xlsx"
REAL_FIXTURE = FIXTURE_DIR / "CAAU8011090 UNLOADING PLAN.xlsx"


def test_batch_runner_generates_phase0_outputs_and_reports_failed_files(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    shutil.copy2(REAL_FIXTURE, input_dir / REAL_FIXTURE.name)
    failed_file = input_dir / "corrupt-upload.xlsx"
    failed_file.write_bytes(b"not an xlsx workbook")

    result = run_batch(
        input_dir=input_dir,
        template_path=TEMPLATE_PATH,
        output_dir=tmp_path / "storage",
    )

    assert result.processedCount == 2
    assert result.successCount == 1
    assert result.failedCount == 1
    assert len(list(result.parsedJsonDir.glob("*.json"))) == 2
    assert list(result.reportDir.glob("*.xlsx"))
    assert list(result.labelDir.glob("*.pdf"))
    assert result.taskReport.htmlPath.is_file()
    assert result.taskReport.correctionsPath.is_file()

    import_manifest = json.loads(
        (tmp_path / "storage" / "original_files" / "import_manifest.json").read_text(encoding="utf-8")
    )
    assert len(import_manifest["records"]) == 2
    assert all(record["sha256"] for record in import_manifest["records"])

    failed_json = next(result.parsedJsonDir.glob("corrupt-upload-*.json"))
    failed_payload = json.loads(failed_json.read_text(encoding="utf-8"))
    assert failed_payload["sha256"]
    assert failed_payload["task_status"] == "ERROR"
    assert failed_payload["errors"]

    html = result.taskReport.htmlPath.read_text(encoding="utf-8")
    assert REAL_FIXTURE.name in html
    assert failed_file.name in html
    assert "Unable to read Excel workbook" in html


def test_batch_runner_preserves_detailed_pallet_rule_outputs(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    workbook_path = input_dir / "TSTU1234567 detailed pallet rules.xlsx"
    _write_detailed_rule_workbook(workbook_path)

    result = run_batch(
        input_dir=input_dir,
        template_path=TEMPLATE_PATH,
        output_dir=tmp_path / "storage",
        generated_at=datetime(2026, 6, 25, 9, 30),
    )

    assert result.processedCount == 1
    assert result.successCount == 1
    assert result.warningFileCount == 1

    payload = json.loads(next(result.parsedJsonDir.glob("*.json")).read_text(encoding="utf-8"))
    assert payload["task_status"] == "WARNING"
    assert payload["parsed_result"]["containerNo"] == "TSTU1234567"

    plans = {
        (plan["destinationCode"], plan["packageType"]): plan
        for plan in payload["pallet_result"]["plans"]
    }
    expected = {
        ("YYC4", "CARTON"): (
            "FOOTPRINT_HEIGHT_VOLUME_LOW_1_7",
            2.04,
            "CEIL",
            2,
        ),
        ("YYC6", "CARTON"): (
            "FOOTPRINT_HEIGHT_VOLUME_LOW_1_7",
            2.04,
            "CEIL",
            1,
        ),
        ("YEG2", "CARTON"): (
            "FOOTPRINT_HEIGHT_VOLUME_LOW_1_7",
            2.04,
            "CEIL",
            7,
        ),
        ("YVR2", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            2,
        ),
        ("YVR3", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            2,
        ),
        ("YVR4", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            1,
        ),
        ("YEG1", "CARTON"): (
            "YEG1_FOOTPRINT_HEIGHT_PLUS_4",
            2.04,
            "CEIL",
            6,
        ),
        ("UPS", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            3,
        ),
        ("Private Address / ADDR-CARTON", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            2,
        ),
        ("Private Address / ADDR-UNKNOWN", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            2,
        ),
        ("Commercial Address / ADDR-WOOD", "WOODEN_CRATE"): (
            "WOODEN_CRATE_PIECE_COUNT",
            None,
            "PIECE_COUNT",
            7,
        ),
        ("Other Oversize", "CARTON"): (
            "OVERSIZE_PIECE_COUNT",
            None,
            "PIECE_COUNT",
            2,
        ),
        ("Mixed Cargo", "CARTON"): (
            "OTHER_DESTINATION_FOOTPRINT_HEIGHT_2_2",
            2.64,
            "CEIL",
            1,
        ),
        ("Mixed Cargo", "WOODEN_CRATE"): (
            "WOODEN_CRATE_PIECE_COUNT",
            None,
            "PIECE_COUNT",
            3,
        ),
    }

    for key, (rule_code, basis, rounding_mode, final_pallets) in expected.items():
        plan = plans[key]
        assert plan["ruleCode"] == rule_code
        if basis is None:
            assert plan["calculationBasisCbm"] is None
        else:
            assert Decimal(plan["calculationBasisCbm"]) == Decimal(str(basis))
        assert plan["roundingMode"] == rounding_mode
        assert plan["finalPallets"] == final_pallets
        assert len(plan["palletIds"]) == final_pallets

    warning_codes = {warning["code"] for warning in payload["warnings"]}
    assert "PACKAGE_TYPE_CONFIRMATION_REQUIRED" not in warning_codes
    assert "DESTINATION_RANGE_EXCEEDED" not in warning_codes

    label_result = payload["label_result"]
    pallet_result = payload["pallet_result"]
    assert label_result["labelCount"] == pallet_result["totalFinalPallets"]
    assert len(label_result["palletIds"]) == len(set(label_result["palletIds"]))
    assert len(label_result["qrPayloads"]) == len(set(label_result["qrPayloads"]))
    assert all(qr.split("|")[-1] in label_result["palletIds"] for qr in label_result["qrPayloads"])
    assert Path(label_result["outputPath"]).read_bytes().startswith(b"%PDF")

    pdf_text = Path(label_result["outputPath"]).read_bytes().decode("latin1", errors="ignore")
    width, height = _first_media_box_size(pdf_text)
    assert width == pytest.approx(_mm_points(150), abs=0.01)
    assert height == pytest.approx(_mm_points(100), abs=0.01)

    workbook = load_workbook(payload["report_result"]["outputPath"], data_only=False)
    try:
        report_rows = [
            (sheet, row)
            for sheet in workbook.worksheets
            for row in (4, 6, 8, 10, 12, 14, 16, 18)
        ]
        for (worksheet, row), plan in zip(
            report_rows, payload["pallet_result"]["plans"]
        ):
            assert worksheet[f"N{row}"].value == plan["destinationCode"]
            assert worksheet[f"O{row}"].value == plan["finalPallets"]
            assert worksheet[f"P{row}"].value == plan["totalCartons"]
        report_destinations = {
            worksheet[f"N{row}"].value: worksheet[f"O{row}"].value
            for worksheet, row in report_rows
            if worksheet[f"N{row}"].value
        }
        assert report_destinations["UPS"] == 3
        assert sum(
            int(worksheet[f"O{row}"].value or 0)
            for worksheet, row in report_rows
        ) == pallet_result["totalFinalPallets"]
    finally:
        workbook.close()

    task_report_html = next((tmp_path / "storage" / "task_reports").glob("task-report-*.html")).read_text(
        encoding="utf-8"
    )
    assert "rule YEG1 footprint volume plus four pallets" in task_report_html
    assert "rule Wooden crate piece count" in task_report_html
    assert "rounding Piece count" in task_report_html
    assert "Private or commercial address package type was not recognized" not in task_report_html
    assert "package Carton" in task_report_html


def test_unloading_worker_batch_cli_prints_summary(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    shutil.copy2(REAL_FIXTURE, input_dir / REAL_FIXTURE.name)
    output_dir = tmp_path / "storage"

    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "batch",
            "--input-dir",
            str(input_dir),
            "--template",
            str(TEMPLATE_PATH),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0
    assert "Batch completed" in result.output
    assert "Processed: 1" in result.output
    assert "Task report:" in result.output
    assert list((output_dir / "parsed_json").glob("*.json"))
    assert list((output_dir / "reports").glob("*.xlsx"))
    assert list((output_dir / "labels").glob("*.pdf"))


def test_unloading_worker_parse_file_cli_outputs_parser_only_json() -> None:
    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "parse-file",
            "--input-file",
            str(REAL_FIXTURE),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["parse_scope"] == "parser-only"
    assert payload["source_file"] == str(REAL_FIXTURE.resolve())
    assert payload["task_status"] in {"SUCCESS", "WARNING"}
    assert payload["parsed_result"]["containerNo"] == "CAAU8011090"
    assert payload["parsed_result"]["parserVersion"] == "unloading-plan-cn-v1"
    assert payload["parsed_result"]["lines"]
    assert payload["pallet_result"]["plans"]
    assert payload["report_result"] is None
    assert payload["label_result"] is None


def test_unloading_worker_parse_file_cli_retains_api_pallet_policy() -> None:
    runner = CliRunner()
    policy = {
        "policyVersion": "pallet-footprint-v1",
        "settingsRevision": "settings-revision-test",
        "palletLengthM": "1.0",
        "palletWidthM": "1.1",
        "lowHeightM": "1.7",
        "otherHeightM": "2.2",
        "yeg1ExtraPallets": 4,
    }

    result = runner.invoke(
        app,
        [
            "parse-file",
            "--input-file",
            str(REAL_FIXTURE),
            "--pallet-policy-json",
            json.dumps(policy),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["pallet_policy"] == policy
    assert payload["pallet_result"]["plans"]
    assert all(
        isinstance(plan["palletCapacityCbm"], str)
        for plan in payload["pallet_result"]["plans"]
    )
    assert all(
        plan["policySnapshot"]["palletWidthM"] == "1.1"
        for plan in payload["pallet_result"]["plans"]
    )


def test_unloading_worker_write_report_cli_generates_excel_only(tmp_path: Path) -> None:
    runner = CliRunner()
    parsed = runner.invoke(
        app,
        [
            "parse-file",
            "--input-file",
            str(REAL_FIXTURE),
        ],
    )
    assert parsed.exit_code == 0
    parsed_payload = json.loads(parsed.output)
    report_payload_path = tmp_path / "report-payload.json"
    report_payload_path.write_text(
        json.dumps(
            {
                "company": "Bestar",
                "parsed_result": parsed_payload["parsed_result"],
                "pallet_result": parsed_payload["pallet_result"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "reports"

    result = runner.invoke(
        app,
        [
            "write-report",
            "--payload",
            str(report_payload_path),
            "--template",
            str(TEMPLATE_PATH),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["task_status"] in {"SUCCESS", "WARNING"}
    assert payload["report_result"]["outputPath"].endswith(".xlsx")
    assert Path(payload["report_result"]["outputPath"]).is_file()
    assert not list(tmp_path.glob("*.pdf"))


def test_unloading_worker_write_labels_cli_generates_pdf_only(tmp_path: Path) -> None:
    runner = CliRunner()
    parsed = runner.invoke(
        app,
        [
            "parse-file",
            "--input-file",
            str(REAL_FIXTURE),
        ],
    )
    assert parsed.exit_code == 0
    parsed_payload = json.loads(parsed.output)
    label_payload_path = tmp_path / "label-payload.json"
    label_payload_path.write_text(
        json.dumps(
            {
                "parsed_result": parsed_payload["parsed_result"],
                "pallet_result": parsed_payload["pallet_result"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "labels"

    result = runner.invoke(
        app,
        [
            "write-labels",
            "--payload",
            str(label_payload_path),
            "--output-dir",
            str(output_dir),
            "--label-date",
            "2026-06-26",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["task_status"] in {"SUCCESS", "WARNING"}
    assert payload["label_result"]["outputPath"].endswith(".pdf")
    assert Path(payload["label_result"]["outputPath"]).is_file()
    assert payload["label_result"]["labelCount"] > 0
    assert payload["label_result"]["palletIds"]
    assert payload["label_result"]["qrPayloads"][0].startswith("SSP1|PALLET|")
    assert not list(tmp_path.glob("*.xlsx"))


def test_unloading_worker_write_print_calibration_cli_generates_pdf(
    tmp_path: Path,
) -> None:
    runner = CliRunner()
    output_dir = tmp_path / "labels"

    result = runner.invoke(
        app,
        [
            "write-print-calibration",
            "--output-dir",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    calibration = payload["print_calibration_result"]
    assert payload["task_status"] == "SUCCESS"
    assert calibration["outputPath"].endswith("print-calibration.pdf")
    assert calibration["pageWidthMm"] == 150
    assert calibration["pageHeightMm"] == 100
    assert calibration["qrBoxMm"] == 25
    assert Path(calibration["outputPath"]).is_file()


def _write_detailed_rule_workbook(path: Path) -> None:
    workbook = load_workbook(REAL_FIXTURE)
    worksheet = workbook["Sheet1"]
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "CAAU8011090" in cell.value:
                cell.value = cell.value.replace("CAAU8011090", "TSTU1234567")
    if worksheet.max_row > 6:
        worksheet.delete_rows(7, worksheet.max_row - 6)
    rows = (
        ("WB-YYC4", "", "", 10, 100, 3.41, "YYC4", "LTL", ""),
        ("WB-YYC6", "", "", 10, 100, 1.70, "YYC6", "LTL", ""),
        ("WB-YEG2", "", "", 10, 100, 13.236, "YEG2", "LTL", ""),
        ("WB-YVR2", "", "", 10, 100, 4.39, "YVR2", "LTL", ""),
        ("WB-YVR3", "", "", 10, 100, 4.41, "YVR3", "LTL", ""),
        ("WB-YVR4", "", "", 10, 100, 0.50, "YVR4", "LTL", ""),
        ("WB-YEG1", "", "", 10, 100, 3.40, "YEG1", "LTL", ""),
        ("WB-UPS-57", "", "", 57, 100, 5.40, "UPS", "快递派送", ""),
        ("ADDR-CARTON", "", "", 12, 100, 3.61, "Private Address", "LTL", "carton"),
        ("ADDR-UNKNOWN", "", "", 10, 100, 3.61, "Private Address", "LTL", ""),
        ("ADDR-WOOD", "", "", 7, 100, 9.00, "Commercial Address", "LTL", "wooden crate"),
        ("OVERSIZE-TWO", "", "", 2, 100, 5.60, "Other Oversize", "LTL", ""),
        ("MIXED-CARTON", "", "", 10, 100, 1.00, "Mixed Cargo", "LTL", "carton"),
        ("MIXED-WOOD", "", "", 3, 100, 9.00, "Mixed Cargo", "LTL", "wooden crate"),
    )
    for row in rows:
        worksheet.append(row)

    workbook.save(path)


def _first_media_box_size(pdf_text: str) -> tuple[float, float]:
    match = re.search(r"/MediaBox\s*\[\s*0\s+0\s+([0-9.]+)\s+([0-9.]+)\s*\]", pdf_text)
    assert match is not None
    return float(match.group(1)), float(match.group(2))


def _mm_points(mm: int) -> float:
    return mm * (72 / 25.4)
