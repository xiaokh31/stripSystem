from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from worker_python.cli import app


REPO_ROOT = Path(__file__).resolve().parents[4]
MANIFEST_PATH = REPO_ROOT / "docs" / "fixtures.md"
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
TEMPLATE_PATH = REPO_ROOT / "samples" / "templates" / "卸柜报告-En.xlsx"


def test_phase0_e2e_regression_uses_all_real_fixtures(tmp_path: Path) -> None:
    fixture_paths = _registered_unloading_fixture_paths()
    fixture_names = {fixture.name for fixture in fixture_paths}
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "storage"

    input_dir.mkdir()
    for fixture_path in fixture_paths:
        shutil.copy2(fixture_path, input_dir / fixture_path.name)

    result = CliRunner().invoke(
        app,
        [
            "batch",
            "--input-dir",
            str(input_dir),
            "--template",
            str(TEMPLATE_PATH),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    assert f"Processed: {len(fixture_paths)}" in result.output

    parsed_payloads = _parsed_payloads(output_dir)
    assert len(parsed_payloads) == len(fixture_paths)
    assert {payload["original_filename"] for payload in parsed_payloads} == fixture_names

    for payload in parsed_payloads:
        assert payload["schema_version"] == 1
        assert payload["batch_version"] == "phase0-batch-v1"
        assert payload["sha256"]
        assert payload["detection"]["format_type"] in {
            "UNLOADING_PLAN_CN",
            "BESTAR_RECEIVING",
            "UNKNOWN",
        }
        assert payload["task_status"] in {"SUCCESS", "WARNING", "ERROR"}
        assert isinstance(payload["warnings"], list)
        assert isinstance(payload["errors"], list)

    supported_payloads = [
        payload
        for payload in parsed_payloads
        if payload["detection"]["format_type"] in {"UNLOADING_PLAN_CN", "BESTAR_RECEIVING"}
    ]
    pallet_ids = [
        pallet_id
        for payload in supported_payloads
        for pallet_id in (payload["label_result"] or {}).get("palletIds", [])
    ]
    qr_payloads = [
        qr_payload
        for payload in supported_payloads
        for qr_payload in (payload["label_result"] or {}).get("qrPayloads", [])
    ]
    assert supported_payloads
    assert {payload["detection"]["format_type"] for payload in supported_payloads} >= {
        "UNLOADING_PLAN_CN",
        "BESTAR_RECEIVING",
    }
    assert any(payload["parsed_result"]["lines"] for payload in supported_payloads)
    assert any(payload["pallet_result"]["plans"] for payload in supported_payloads)
    assert any(payload["report_result"] and payload["report_result"]["outputPath"] for payload in supported_payloads)
    assert any(payload["label_result"] and payload["label_result"]["outputPath"] for payload in supported_payloads)
    assert len(pallet_ids) == len(set(pallet_ids))
    assert len(qr_payloads) == len(set(qr_payloads))
    assert all(qr_payload.split("|")[-1] in pallet_ids for qr_payload in qr_payloads)

    report_files = sorted((output_dir / "reports").glob("*.xlsx"))
    label_files = sorted((output_dir / "labels").glob("*.pdf"))
    report_manifest = json.loads((output_dir / "reports" / "report_manifest.json").read_text(encoding="utf-8"))
    label_manifest = json.loads((output_dir / "labels" / "label_manifest.json").read_text(encoding="utf-8"))
    assert report_files
    assert label_files
    assert len(report_files) == len(report_manifest["records"])
    assert len(label_files) == len(label_manifest["records"])
    assert label_files[0].read_bytes().startswith(b"%PDF")

    workbook = load_workbook(report_files[0], data_only=False)
    try:
        assert "Sheet1" in workbook.sheetnames
    finally:
        workbook.close()

    task_report_path = next((output_dir / "task_reports").glob("task-report-*.html"))
    task_report_html = task_report_path.read_text(encoding="utf-8")
    for fixture_name in fixture_names:
        assert fixture_name in task_report_html

    corrections_path = next((output_dir / "corrections").glob("corrections-*.json"))
    corrections = json.loads(corrections_path.read_text(encoding="utf-8"))
    assert len(corrections["corrections"]) == len(fixture_paths)

    import_manifest = json.loads(
        (output_dir / "original_files" / "import_manifest.json").read_text(encoding="utf-8")
    )
    assert len(import_manifest["records"]) == len(fixture_paths)
    assert len({record["sha256"] for record in import_manifest["records"]}) == len(fixture_paths)


def _registered_unloading_fixture_paths() -> list[Path]:
    paths = []
    for line in MANIFEST_PATH.read_text(encoding="utf-8").splitlines():
        if not line.startswith("| samples/unloading-plans/"):
            continue
        fixture_path = REPO_ROOT / line.strip("|").split("|")[0].strip()
        paths.append(fixture_path)
    return sorted(paths)


def _parsed_payloads(output_dir: Path) -> list[dict]:
    return [
        json.loads(path.read_text(encoding="utf-8"))
        for path in sorted((output_dir / "parsed_json").glob("*.json"))
    ]
