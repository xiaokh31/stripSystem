from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from worker_python.parser_profiles import inspection as inspection_module
from worker_python.imports import compute_sha256
from worker_python.parser import parse_bestar_receiving, parse_unloading_plan_cn
from worker_python.parser_profiles import (
    ALLOWED_MAPPING_OPERATION_CODES,
    PROFILE_CONTRACT_CODES,
    InspectionLimits,
    MappingDefinition,
    ProfileDefinitionError,
    WorkbookInspectionError,
    execute_mapping,
    fingerprint_definition_json_schema,
    inspect_workbook,
    mapping_definition_json_schema,
    profile_parse_result_json_schema,
    suggest_mappings,
    workbook_inspection_json_schema,
)


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
UNLOADING_FIXTURE = FIXTURE_DIR / "CAAU8011090 UNLOADING PLAN.xlsx"
STANDARD_FIXTURE = FIXTURE_DIR / "Unloading Plan SMCU1012780.xlsx"
BESTAR_FIXTURE = FIXTURE_DIR / "137675 JXJU3246131  PO#3404  BESTAR.xlsx"


def _unloading_definition() -> dict:
    return {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "fixture-unloading-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Sheet1"},
        "header": {"row": 6, "rowCount": 1},
        "dataRange": {"startRow": 7, "maxRows": 200},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "B3"}],
            "transforms": [
                {"op": "trim"},
                {"op": "regex_extract", "pattern": "([A-Z]{4}\\d{7})", "group": 1},
            ],
        },
        "fields": {
            "waybillNo": {"sources": [{"kind": "column", "header": "运单号"}]},
            "fbaNo": {"sources": [{"kind": "column", "header": "FBA NO."}]},
            "poNumber": {"sources": [{"kind": "column", "header": "PO#"}]},
            "cartons": {
                "sources": [{"kind": "column", "header": "箱数/件数"}],
                "transforms": [{"op": "parse_integer", "groupSeparator": ","}],
            },
            "weight": {
                "sources": [{"kind": "column", "header": "重量"}],
                "transforms": [{"op": "parse_decimal", "groupSeparator": ","}],
            },
            "volumeCbm": {
                "sources": [{"kind": "column", "header": "体积"}],
                "transforms": [{"op": "parse_decimal", "groupSeparator": ","}],
            },
            "destinationCode": {
                "sources": [{"kind": "column", "header": "派送目的地"}],
                "transforms": [{"op": "trim"}],
            },
            "deliveryMethod": {"sources": [{"kind": "column", "header": "派送方式"}]},
            "note": {"sources": [{"kind": "column", "header": "特殊指令/备注"}]},
        },
        "rowPredicates": [
            {"op": "skip_blank", "headers": ["运单号", "箱数/件数", "体积"]},
            {
                "op": "exclude",
                "source": {"kind": "column", "header": "运单号"},
                "operator": "in",
                "values": ["合计", "总计", "TOTAL"],
            },
        ],
        "groupBy": ["destinationCode", "packageType"],
    }


def _bestar_definition() -> dict:
    return {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "fixture-bestar-v1",
        "formatType": "BESTAR_RECEIVING",
        "sheet": {"index": 0},
        "header": {"row": 11, "rowCount": 1},
        "dataRange": {"startRow": 12, "maxRows": 50},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "D3"}],
            "transforms": [{"op": "trim"}],
        },
        "metadataFields": {
            "company": {
                "scope": "workbook",
                "sources": [{"kind": "cell", "cell": "D2"}],
                "transforms": [{"op": "trim"}],
            },
            "poNumber": {
                "scope": "workbook",
                "sources": [{"kind": "cell", "cell": "H3"}],
            },
            "customer": {
                "scope": "workbook",
                "sources": [{"kind": "cell", "cell": "D7"}],
                "transforms": [{"op": "trim"}],
            },
            "clearOrderNo": {
                "scope": "workbook",
                "sources": [{"kind": "cell", "cell": "D8"}],
            },
        },
        "fields": {
            "itemNo": {
                "sources": [{"kind": "column", "header": "ITEM#"}],
                "transforms": [{"op": "trim"}],
            },
            "description": {"sources": [{"kind": "column", "header": "DESCRIPTION"}]},
            "cartons": {
                "sources": [{"kind": "column", "header": "TOTAL # OF CARTONS"}],
                "transforms": [{"op": "parse_integer"}],
            },
            "totalSkidCount": {
                "sources": [{"kind": "column", "header": "TOTAL SKID COUNT"}],
                "transforms": [{"op": "parse_integer"}],
            },
            "destinationCode": {"sources": [{"kind": "constant", "value": None}]},
        },
        "rowPredicates": [
            {"op": "skip_blank", "headers": ["ITEM#", "TOTAL # OF CARTONS"]},
            {
                "op": "exclude",
                "source": {"kind": "column", "header": "ITEM#"},
                "operator": "in",
                "values": ["TOTAL", "SUMMARY"],
            },
        ],
        "groupBy": ["destinationCode"],
    }


def test_inspection_is_bounded_and_describes_real_merged_headers() -> None:
    result = inspect_workbook(BESTAR_FIXTURE)

    assert result.contractVersion == "workbook-inspection-v1"
    assert result.workbookType == "OOXML_XLSX"
    assert result.inputSha256 == compute_sha256(BESTAR_FIXTURE)
    assert result.sheets[0].name
    assert result.sheets[0].visibility == "visible"
    assert result.sheets[0].boundedDimensions.maxRow >= 12
    assert result.sheets[0].mergedRanges
    assert any(area.row == 11 for area in result.sheets[0].candidateHeaderAreas)
    assert result.limits.maxRowsPerSheet == 500
    assert all(issue.code != "WORKBOOK_READ_FAILED" for issue in result.issues)


def test_inspection_reports_stable_limit_and_missing_formula_cache_codes(
    tmp_path: Path,
) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Header"
    sheet["A2"] = "=1+1"
    sheet["B10"] = "outside"
    workbook.save(path)
    workbook.close()

    result = inspect_workbook(
        path,
        limits=InspectionLimits(maxRowsPerSheet=5, maxColumnsPerSheet=1, maxCells=5),
    )

    assert {issue.code for issue in result.issues} >= {
        "INSPECTION_ROW_LIMIT_EXCEEDED",
        "INSPECTION_COLUMN_LIMIT_EXCEEDED",
        "FORMULA_CACHED_VALUE_MISSING",
    }
    formula = next(cell for cell in result.sheets[0].sampleCells if cell.cell == "A2")
    assert formula.isFormula is True
    assert formula.hasCachedValue is False
    assert formula.cachedValue is None
    assert formula.cachedValueType is None


def test_inspection_rejects_legacy_xls_with_structured_code(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not-an-ooxml-workbook")

    with pytest.raises(WorkbookInspectionError) as exc_info:
        inspect_workbook(path)

    assert [(issue.code, issue.path) for issue in exc_info.value.issues] == [
        ("WORKBOOK_TYPE_UNSUPPORTED", "workbook")
    ]


def test_inspection_structures_corrupt_ooxml_and_reports_merged_range_limit(
    tmp_path: Path,
) -> None:
    corrupt = tmp_path / "corrupt.xlsx"
    with zipfile.ZipFile(corrupt, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types />")

    with pytest.raises(WorkbookInspectionError) as exc_info:
        inspect_workbook(corrupt)
    assert [issue.code for issue in exc_info.value.issues] == ["WORKBOOK_READ_FAILED"]

    merged = tmp_path / "merged.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    sheet.merge_cells("A2:B2")
    sheet.merge_cells("A3:B3")
    workbook.save(merged)
    workbook.close()

    result = inspect_workbook(
        merged, limits=InspectionLimits(maxMergedRangesPerSheet=2)
    )
    assert len(result.sheets[0].mergedRanges) == 2
    assert any(
        issue.code == "INSPECTION_MERGED_RANGE_LIMIT_EXCEEDED"
        for issue in result.issues
    )


def test_archive_budgets_reject_before_worksheet_xml_is_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    oversized = tmp_path / "oversized.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "x" * 5_000
    workbook.save(oversized)
    workbook.close()

    def forbidden_merge_read(*args: object, **kwargs: object) -> object:
        raise AssertionError("worksheet XML read before archive budget rejection")

    monkeypatch.setattr(
        inspection_module, "_merged_ranges_by_sheet", forbidden_merge_read
    )
    with pytest.raises(WorkbookInspectionError) as exc_info:
        inspect_workbook(
            oversized,
            limits=InspectionLimits(maxArchiveEntryBytes=1_024),
        )
    assert [issue.code for issue in exc_info.value.issues] == [
        "WORKBOOK_ARCHIVE_ENTRY_LIMIT_EXCEEDED"
    ]

    too_many = tmp_path / "too-many.xlsx"
    with zipfile.ZipFile(too_many, "w") as archive:
        for index in range(11):
            archive.writestr(f"entry-{index}", "x")
    with pytest.raises(WorkbookInspectionError) as exc_info:
        inspect_workbook(
            too_many,
            limits=InspectionLimits(maxArchiveEntries=10),
        )
    assert [issue.code for issue in exc_info.value.issues] == [
        "WORKBOOK_ARCHIVE_ENTRY_COUNT_LIMIT_EXCEEDED"
    ]

    too_large = tmp_path / "too-large.xlsx"
    with zipfile.ZipFile(too_large, "w") as archive:
        archive.writestr("large-entry", "x" * 11_000)
    with pytest.raises(WorkbookInspectionError) as exc_info:
        inspect_workbook(
            too_large,
            limits=InspectionLimits(
                maxArchiveEntryBytes=20_000,
                maxArchiveTotalBytes=10_240,
            ),
        )
    assert [issue.code for issue in exc_info.value.issues] == [
        "WORKBOOK_ARCHIVE_TOTAL_SIZE_LIMIT_EXCEEDED"
    ]


def test_mapping_schema_is_strict_versioned_and_rejects_dangerous_definitions() -> None:
    schema = mapping_definition_json_schema()
    assert schema["title"] == "MappingDefinition"
    assert "parser-profile-mapping-v1" in json.dumps(schema, sort_keys=True)
    assert "workbook-inspection-v1" in json.dumps(
        workbook_inspection_json_schema(), sort_keys=True
    )
    assert "workbook-fingerprint-v1" in json.dumps(
        fingerprint_definition_json_schema(), sort_keys=True
    )
    assert "parser-profile-engine-v1" in json.dumps(
        profile_parse_result_json_schema(), sort_keys=True
    )

    for mutation, code, path in [
        (
            {
                "fields": {
                    "cartons": {
                        "sources": [{"kind": "column", "header": "件数"}],
                        "transforms": [{"op": "python", "code": "open('/tmp/x')"}],
                    }
                }
            },
            "MAPPING_OPERATION_UNKNOWN",
            "fields.cartons.transforms.0.op",
        ),
        (
            {"networkTarget": "https://example.test"},
            "MAPPING_DEFINITION_UNKNOWN_FIELD",
            "networkTarget",
        ),
        (
            {
                "fields": {
                    "cartons": {"sources": [{"kind": "cell", "cell": "../../secret"}]}
                }
            },
            "MAPPING_SOURCE_CELL_INVALID",
            "fields.cartons.sources.0.cell",
        ),
        (
            {
                "fields": {
                    "cartons": {
                        "sources": [{"kind": "column", "header": "件数"}],
                        "transforms": [{"op": "regex_extract", "pattern": "(a+)+$"}],
                    }
                }
            },
            "MAPPING_REGEX_UNSAFE",
            "fields.cartons.transforms.0.regex_extract.pattern",
        ),
        (
            {
                "fields": {
                    "cartons": {
                        "sources": [{"kind": "column", "header": "件数"}],
                        "transforms": [{"op": "regex_extract", "pattern": "a*a*b"}],
                    }
                }
            },
            "MAPPING_REGEX_UNSAFE",
            "fields.cartons.transforms.0.regex_extract.pattern",
        ),
        (
            {
                "container": {
                    "scope": "workbook",
                    "sources": [{"kind": "constant", "value": "INVENTED000"}],
                }
            },
            "MAPPING_CONTAINER_CONSTANT_FORBIDDEN",
            None,
        ),
        (
            {
                "fields": {
                    "cartons": {
                        "sources": [{"kind": "column", "header": "件数"}],
                        "transforms": [
                            {
                                "op": "regex_extract",
                                "pattern": "(?:a{1,3}){1,100000}",
                            }
                        ],
                    }
                }
            },
            "MAPPING_REGEX_UNSAFE",
            "fields.cartons.transforms.0.regex_extract.pattern",
        ),
        (
            {
                "fields": {
                    "cartons": {
                        "sources": [{"kind": "column", "header": "件数"}],
                        "transforms": [
                            {
                                "op": "parse_integer",
                                "groupSeparator": ".",
                                "decimalSeparator": ".",
                            }
                        ],
                    }
                }
            },
            "MAPPING_DEFINITION_INVALID",
            "fields.cartons.transforms.0.parse_integer",
        ),
    ]:
        payload = _unloading_definition()
        payload.update(mutation)
        with pytest.raises(ProfileDefinitionError) as exc_info:
            MappingDefinition.validate_definition(payload)
        assert any(
            issue.code == code and issue.path == path for issue in exc_info.value.issues
        )


@pytest.mark.parametrize(
    ("transforms", "expected"),
    [
        ([{"op": "trim"}, {"op": "case", "mode": "upper"}], "ABC"),
        ([{"op": "blank", "values": ["N/A"]}], None),
        (
            [{"op": "parse_decimal", "groupSeparator": ",", "decimalSeparator": "."}],
            "1234.5",
        ),
        ([{"op": "parse_integer", "groupSeparator": ","}], "1234"),
        ([{"op": "lookup", "dictionary": {"Warehouse A": "YEG1"}}], "YEG1"),
        ([{"op": "concatenate", "separator": " / "}], "A  / B"),
        ([{"op": "regex_extract", "pattern": "([A-Z]+)-(\\d+)", "group": 2}], "42"),
        ([{"op": "multiply", "factor": 2}, {"op": "divide", "divisor": 4}], "5.0"),
        (
            [{"op": "unit_conversion", "fromUnit": "CUBIC_FEET", "toUnit": "CBM"}],
            "0.283168",
        ),
    ],
)
def test_each_allowlisted_transform_executes_through_public_mapping_interface(
    tmp_path: Path, transforms: list[dict], expected: object
) -> None:
    path = tmp_path / "transforms.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["value", "fallback"])
    sheet["D1"] = "TEST0000000"
    input_key = next(
        operation["op"] for operation in transforms if operation["op"] != "trim"
    )
    raw = {
        "case": " abc ",
        "blank": "N/A",
        "parse_decimal": "1,234.5",
        "parse_integer": "1,234",
        "lookup": "Warehouse A",
        "concatenate": " A ",
        "regex_extract": "ABC-42",
        "multiply": 10,
        "unit_conversion": 10,
    }[input_key]
    sheet.append([raw, "B"])
    workbook.save(path)
    workbook.close()

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "transform-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 2},
        "container": {"scope": "workbook", "sources": [{"kind": "cell", "cell": "D1"}]},
        "fields": {
            "destinationCode": {
                "sources": [
                    {"kind": "column", "header": "value"},
                    {"kind": "column", "header": "fallback"},
                ],
                "transforms": transforms,
            },
            "cartons": {"sources": [{"kind": "constant", "value": 1}]},
            "volumeCbm": {"sources": [{"kind": "constant", "value": 1}]},
        },
        "groupBy": ["destinationCode"],
    }
    result = execute_mapping(path, definition, replay_input_hash="a" * 64)

    assert result.errors == ()
    assert result.lines[0].destinationCode == expected


def test_coalesce_filters_provenance_unknown_columns_and_stable_warnings(
    tmp_path: Path,
) -> None:
    path = tmp_path / "profile.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Dest", "Alt Dest", "Cartons", "Volume", "Unknown Customer Column"])
    sheet["G1"] = "TEST0000000"
    sheet.append([None, " yeg1 ", "2", "0", "keep-me"])
    sheet.append(["TOTAL", None, "2", "0", "summary"])
    sheet.append([None, None, None, None, None])
    workbook.save(path)
    workbook.close()

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "coalesce-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 10},
        "container": {"scope": "workbook", "sources": [{"kind": "cell", "cell": "G1"}]},
        "fields": {
            "destinationCode": {
                "sources": [
                    {"kind": "column", "header": "Dest"},
                    {"kind": "column", "header": "Alt Dest"},
                ],
                "transforms": [
                    {"op": "coalesce"},
                    {"op": "trim"},
                    {"op": "case", "mode": "upper"},
                ],
            },
            "cartons": {
                "sources": [{"kind": "column", "header": "Cartons"}],
                "transforms": [{"op": "parse_integer"}],
            },
            "volumeCbm": {
                "sources": [{"kind": "column", "header": "Volume"}],
                "transforms": [{"op": "parse_decimal"}],
            },
        },
        "rowPredicates": [
            {"op": "skip_blank", "headers": ["Dest", "Alt Dest", "Cartons"]},
            {
                "op": "stop",
                "source": {"kind": "column", "header": "Dest"},
                "operator": "equals",
                "value": "STOP",
            },
            {
                "op": "exclude",
                "source": {"kind": "column", "header": "Dest"},
                "operator": "equals",
                "value": "TOTAL",
            },
        ],
        "groupBy": ["destinationCode"],
    }

    result = execute_mapping(path, definition, replay_input_hash="b" * 64)

    assert len(result.lines) == 1
    line = result.lines[0]
    assert line.destinationCode == "YEG1"
    assert line.raw_json["Unknown Customer Column"] == "keep-me"
    provenance = line.provenance["destinationCode"]
    assert provenance.sourceRefs[1].cell == "B2"
    assert provenance.transformChain == ("coalesce", "trim", "case")
    assert {issue.code for issue in result.warnings} >= {"ZERO_VOLUME_WITH_CARTONS"}
    assert result.rawMetadata["mappingSchemaVersion"] == "parser-profile-mapping-v1"
    assert result.rawMetadata["fingerprintVersion"] == "workbook-fingerprint-v1"
    assert result.rawMetadata["replayInputHash"] == "b" * 64


def test_multi_row_header_include_filter_and_formula_cache_error(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multi-row.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination", "Quantity", "Calculated"])
    sheet["E1"] = "TEST0000000"
    sheet.append(["Code", "Cartons", "Volume"])
    sheet.append(["YEG1", 2, "=1+1"])
    sheet.append(["YYC4", 3, 1.5])
    workbook.save(path)
    workbook.close()

    inspection = inspect_workbook(path)
    assert any(
        area.row == 1 and area.rowCount == 2
        for area in inspection.sheets[0].candidateHeaderAreas
    )

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "multi-header-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1, "rowCount": 2},
        "dataRange": {"startRow": 3, "maxRows": 10},
        "container": {"scope": "workbook", "sources": [{"kind": "cell", "cell": "E1"}]},
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination / Code"}]
            },
            "cartons": {
                "sources": [{"kind": "column", "header": "Quantity / Cartons"}]
            },
            "volumeCbm": {
                "sources": [{"kind": "column", "header": "Calculated / Volume"}]
            },
        },
        "rowPredicates": [
            {
                "op": "include",
                "source": {"kind": "column", "header": "Destination / Code"},
                "operator": "equals",
                "value": "YEG1",
            }
        ],
        "groupBy": ["destinationCode"],
    }
    result = execute_mapping(path, definition, replay_input_hash="c" * 64)

    assert len(result.lines) == 1
    assert result.lines[0].destinationCode == "YEG1"
    assert any(issue.code == "MAPPING_FORMULA_CACHE_MISSING" for issue in result.errors)


def test_formula_only_required_row_cannot_be_silently_skipped(tmp_path: Path) -> None:
    path = tmp_path / "formula-only.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination", "Cartons", "Volume", "Container"])
    sheet.append([None, None, "=1+1", None])
    sheet["D1"] = "TEST0000000"
    workbook.save(path)
    workbook.close()

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "formula-only-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 5},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "D1"}],
        },
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}]
            },
            "cartons": {"sources": [{"kind": "column", "header": "Cartons"}]},
            "volumeCbm": {"sources": [{"kind": "column", "header": "Volume"}]},
        },
        "rowPredicates": [
            {
                "op": "skip_blank",
                "headers": ["Destination", "Cartons", "Volume"],
            }
        ],
    }

    result = execute_mapping(path, definition, replay_input_hash="e" * 64)
    assert any(
        issue.code == "MAPPING_FORMULA_CACHE_MISSING" and issue.path == "C2"
        for issue in result.errors
    )


def test_mapping_rejects_wide_and_over_cell_budget_workbooks(tmp_path: Path) -> None:
    wide = tmp_path / "wide.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Destination"
    sheet["B1"] = "Cartons"
    sheet["C1"] = "Volume"
    sheet.cell(row=1, column=101, value="Unknown")
    sheet.append(["YEG1", 1, 1])
    workbook.save(wide)
    workbook.close()

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "budget-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 10},
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}]
            },
            "cartons": {"sources": [{"kind": "column", "header": "Cartons"}]},
            "volumeCbm": {"sources": [{"kind": "column", "header": "Volume"}]},
        },
    }
    result = execute_mapping(wide, definition, replay_input_hash="f" * 64)
    assert [issue.code for issue in result.errors] == ["MAPPING_COLUMN_LIMIT_EXCEEDED"]

    bounded = tmp_path / "bounded.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination", "Cartons", "Volume"])
    sheet.append(["YEG1", 1, 1])
    sheet.append(["YYC4", 2, 2])
    workbook.save(bounded)
    workbook.close()
    result = execute_mapping(
        bounded,
        definition,
        replay_input_hash="f" * 64,
        limits=InspectionLimits(
            maxColumnsPerSheet=3,
            maxCells=5,
        ),
    )
    assert [issue.code for issue in result.errors] == ["MAPPING_CELL_LIMIT_EXCEEDED"]


def test_mapping_enforces_row_budget_even_for_a_narrow_sheet(tmp_path: Path) -> None:
    path = tmp_path / "too-many-rows.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination"])
    for _ in range(501):
        sheet.append(["YEG1"])
    workbook.save(path)
    workbook.close()

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "row-budget-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 1_000},
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}]
            }
        },
    }
    result = execute_mapping(path, definition, replay_input_hash="1" * 64)
    assert [issue.code for issue in result.errors] == ["MAPPING_ROW_BUDGET_EXCEEDED"]


def test_regex_execution_has_timeout_and_returns_stable_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "regex-timeout.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination"])
    sheet["B1"] = "TEST0000000"
    sheet.append(["YEG1"])
    workbook.save(path)
    workbook.close()

    observed_timeouts: list[float] = []

    def timeout_search(pattern: str, text: str, *, timeout: float) -> object:
        observed_timeouts.append(timeout)
        raise TimeoutError

    monkeypatch.setattr(
        "worker_python.parser_profiles.mapping.bounded_regex.search",
        timeout_search,
    )
    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "regex-timeout-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 2},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "B1"}],
        },
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}],
                "transforms": [{"op": "regex_extract", "pattern": "[A-Z]+"}],
            }
        },
    }
    result = execute_mapping(path, definition, replay_input_hash="2" * 64)
    assert len(observed_timeouts) == 1
    assert 0 < observed_timeouts[0] <= 0.05
    assert [issue.code for issue in result.errors] == ["MAPPING_REGEX_TIMEOUT"]


def test_regex_request_operation_budget_is_shared_across_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "regex-budget.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination", "Note"])
    sheet["C1"] = "TEST0000000"
    sheet.append(["YEG1", "KEEP"])
    workbook.save(path)
    workbook.close()
    monkeypatch.setattr("worker_python.parser_profiles.mapping.MAX_REGEX_OPERATIONS", 1)
    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "regex-budget-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 2},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "C1"}],
        },
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}],
                "transforms": [{"op": "regex_extract", "pattern": "[A-Z]+"}],
            },
            "note": {
                "sources": [{"kind": "column", "header": "Note"}],
                "transforms": [{"op": "regex_extract", "pattern": "[A-Z]+"}],
            },
        },
    }
    result = execute_mapping(path, definition, replay_input_hash="3" * 64)
    assert any(
        issue.code == "MAPPING_REGEX_BUDGET_EXCEEDED" and issue.path == "fields.note"
        for issue in result.errors
    )


def test_regex_predicate_timeout_is_a_structured_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "predicate-timeout.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination"])
    sheet["B1"] = "TEST0000000"
    sheet.append(["YEG1"])
    workbook.save(path)
    workbook.close()

    def timeout_search(pattern: str, text: str, *, timeout: float) -> object:
        raise TimeoutError

    monkeypatch.setattr(
        "worker_python.parser_profiles.mapping.bounded_regex.search",
        timeout_search,
    )
    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "predicate-timeout-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 2},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "B1"}],
        },
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}]
            }
        },
        "rowPredicates": [
            {
                "op": "include",
                "source": {"kind": "column", "header": "Destination"},
                "operator": "regex",
                "pattern": "[A-Z]+",
            }
        ],
    }
    result = execute_mapping(path, definition, replay_input_hash="4" * 64)
    assert result.lines == ()
    assert [issue.code for issue in result.errors] == ["MAPPING_REGEX_TIMEOUT"]


def test_regex_predicate_input_limit_is_a_structured_error(tmp_path: Path) -> None:
    path = tmp_path / "predicate-input-limit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Destination"])
    sheet["B1"] = "TEST0000000"
    sheet.append(["A" * 10_001])
    workbook.save(path)
    workbook.close()
    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "predicate-input-limit-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 2},
        "container": {
            "scope": "workbook",
            "sources": [{"kind": "cell", "cell": "B1"}],
        },
        "fields": {
            "destinationCode": {
                "sources": [{"kind": "column", "header": "Destination"}]
            }
        },
        "rowPredicates": [
            {
                "op": "include",
                "source": {"kind": "column", "header": "Destination"},
                "operator": "regex",
                "pattern": "[A-Z]+",
            }
        ],
    }
    result = execute_mapping(path, definition, replay_input_hash="5" * 64)
    assert result.lines == ()
    assert [issue.code for issue in result.errors] == [
        "MAPPING_REGEX_INPUT_LIMIT_EXCEEDED"
    ]


def test_mapping_row_budget_is_explicit_and_ambiguous_suggestions_are_all_unapproved(
    tmp_path: Path,
) -> None:
    path = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["仓库代码", "派送目的地", "件数", "体积"])
    sheet["F1"] = "TEST0000000"
    sheet.append(["YEG1", "YYC4", 1, 1])
    sheet.append(["YEG2", "YYC6", 2, 2])
    workbook.save(path)
    workbook.close()

    suggestions = [
        suggestion
        for suggestion in suggest_mappings(inspect_workbook(path))
        if suggestion.canonicalField == "destinationCode"
    ]
    assert [suggestion.source.header for suggestion in suggestions] == [
        "仓库代码",
        "派送目的地",
    ]
    assert all(suggestion.approved is False for suggestion in suggestions)

    definition = {
        "schemaVersion": "parser-profile-mapping-v1",
        "profileVersion": "row-limit-v1",
        "formatType": "UNLOADING_PLAN_CN",
        "sheet": {"name": "Data"},
        "header": {"row": 1},
        "dataRange": {"startRow": 2, "maxRows": 1},
        "container": {"scope": "workbook", "sources": [{"kind": "cell", "cell": "F1"}]},
        "fields": {
            "destinationCode": {"sources": [{"kind": "column", "header": "仓库代码"}]},
            "cartons": {"sources": [{"kind": "column", "header": "件数"}]},
            "volumeCbm": {"sources": [{"kind": "column", "header": "体积"}]},
        },
    }
    result = execute_mapping(path, definition, replay_input_hash="d" * 64)
    assert len(result.lines) == 1
    assert any(issue.code == "MAPPING_ROW_LIMIT_EXCEEDED" for issue in result.warnings)


def test_real_unloading_fixture_profile_matches_builtin_key_fields_and_provenance() -> (
    None
):
    built_in = parse_unloading_plan_cn(UNLOADING_FIXTURE)
    result = execute_mapping(
        UNLOADING_FIXTURE,
        _unloading_definition(),
        replay_input_hash=compute_sha256(UNLOADING_FIXTURE),
    )

    assert result.containerNo == built_in.containerNo
    assert len(result.lines) == len(built_in.lines)
    assert [line.cartons for line in result.lines] == [
        line.cartons for line in built_in.lines
    ]
    assert [line.volumeCbm for line in result.lines] == pytest.approx(
        [line.volumeCbm for line in built_in.lines]
    )
    assert [line.destinationCode for line in result.lines] == [
        line.destinationCode for line in built_in.lines
    ]
    first = result.lines[0]
    assert first.provenance["destinationCode"].sourceRefs[0].cell == "G7"
    assert first.provenance["cartons"].sourceRefs[0].cell == "D7"
    assert first.provenance["volumeCbm"].sourceRefs[0].cell == "F7"


def test_real_bestar_fixture_profile_matches_builtin_cartons_and_raw_columns() -> None:
    built_in = parse_bestar_receiving(BESTAR_FIXTURE)
    result = execute_mapping(
        BESTAR_FIXTURE,
        _bestar_definition(),
        replay_input_hash=compute_sha256(BESTAR_FIXTURE),
    )

    assert result.containerNo == built_in.containerNo
    assert result.company == "BESTAR"
    assert result.poNumber == built_in.poNumber
    assert result.customer == built_in.customer
    assert result.clearOrderNo == built_in.clearOrderNo
    assert [line.cartons for line in result.lines] == [
        line.totalCartons for line in built_in.lines
    ]
    assert [line.itemNo for line in result.lines] == [
        line.itemNo for line in built_in.lines
    ]
    assert result.lines[0].raw_json["ITEM#"] == built_in.lines[0].raw_json["ITEM#"]
    assert (
        result.destinationSummaries[0].status == built_in.destinationSummaries[0].status
    )
    assert any(issue.code == "NEED_MANUAL_DESTINATION" for issue in result.warnings)


def test_suggestions_return_all_ambiguous_aliases_as_unapproved() -> None:
    inspection = inspect_workbook(STANDARD_FIXTURE)
    suggestions = suggest_mappings(inspection)
    destination = [
        item for item in suggestions if item.canonicalField == "destinationCode"
    ]

    assert destination
    assert all(item.approved is False for item in destination)
    assert all(item.reasonCode == "HEADER_ALIAS_MATCH" for item in destination)
    assert all(item.evidence.normalizedHeader for item in destination)


def test_contract_contains_only_stable_codes_not_localized_messages() -> None:
    inspection_payload = inspect_workbook(BESTAR_FIXTURE).model_dump(mode="json")
    mapping_payload = execute_mapping(
        BESTAR_FIXTURE,
        _bestar_definition(),
        replay_input_hash=compute_sha256(BESTAR_FIXTURE),
    ).model_dump(mode="json")
    serialized = json.dumps(
        [inspection_payload, mapping_payload], ensure_ascii=False, sort_keys=True
    )

    assert '"message"' not in serialized
    assert "缺少" not in serialized
    assert "Please" not in serialized


def test_stable_code_and_operation_registries_cover_worker_contract() -> None:
    source_dir = (
        REPO_ROOT
        / "apps"
        / "worker-python"
        / "src"
        / "worker_python"
        / "parser_profiles"
    )
    source = "\n".join(
        path.read_text(encoding="utf-8") for path in source_dir.glob("*.py")
    )
    emitted_codes = set(re.findall(r'code=["\']([A-Z][A-Z0-9_]+)["\']', source))

    assert emitted_codes <= PROFILE_CONTRACT_CODES
    assert ALLOWED_MAPPING_OPERATION_CODES == {
        "trim",
        "case",
        "blank",
        "parse_decimal",
        "parse_integer",
        "coalesce",
        "lookup",
        "concatenate",
        "regex_extract",
        "multiply",
        "divide",
        "unit_conversion",
        "skip_blank",
        "skip_summary",
        "include",
        "exclude",
        "stop",
    }
