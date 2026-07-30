from __future__ import annotations

import hashlib
import json
import xml.etree.ElementTree as ET
from copy import copy
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText

import worker_python.reports.excel_report_writer as report_writer
from worker_python.imports import ImportRegistry
from worker_python.pallets import calculate_pallets, inputs_from_destination_summaries
from worker_python.parser import parse_bestar_receiving, parse_unloading_plan_cn
from worker_python.reports.cell_map import (
    ADDITIONAL_DESTINATION_ROWS,
    DESTINATION_ROWS,
    DestinationLayoutMode,
    EXPANDED_DESTINATION_ROWS,
    PRIMARY_DESTINATION_ROWS,
    layout_mode_for_page_count,
    rows_for_page_count,
)
from worker_python.reports.excel_report_writer import (
    DEFAULT_TEMPLATE_PATH,
    write_excel_report,
)


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
STANDARD_FIXTURE = FIXTURE_DIR / "CAAU8011090 UNLOADING PLAN.xlsx"
OVERFLOW_FIXTURE = FIXTURE_DIR / "ZCSU9025988B unloading plan.xlsx"
BESTAR_FIXTURE = FIXTURE_DIR / "137675 JXJU3246131  PO#3404  BESTAR.xlsx"
EXPECTED_TEMPLATE_SHA256 = (
    "31a613e86a76447bfcbb308f1a23f6072dd1a5381f1992fbc0757a2735c92027"
)
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML_NS = {"m": SPREADSHEET_NS}


def test_excel_report_writer_generates_openable_report_from_real_parsed_result(
    tmp_path: Path,
) -> None:
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)

    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert result.errors == ()
    assert result.outputPath.is_file()
    assert parsed.containerNo in result.outputPath.name
    assert result.totalCartons == 896
    assert result.writtenDestinationCount == 9
    assert result.totalDestinationCount == 9
    assert len(result.orderedDestinationDigest) == 64

    workbook = load_workbook(result.outputPath, data_only=False)
    worksheet = workbook["Sheet1"]
    assert worksheet["D1"].value == "2026-06-25"
    assert worksheet["H1"].value == "09:30"
    assert worksheet["K1"].value == "CAAU8011090"
    assert worksheet["D2"].value == "Bestar"
    assert tuple(
        (
            worksheet[f"N{row}"].value,
            worksheet[f"O{row}"].value,
            worksheet[f"P{row}"].value,
        )
        for row in range(4, 4 + len(pallet_result.plans))
    ) == tuple(
        (plan.destinationCode, plan.finalPallets, plan.totalCartons)
        for plan in pallet_result.plans
    )
    assert worksheet["P20"].value == 896
    assert _populated_sheet_names(workbook) == ["Sheet1"]
    assert (
        sum(
            int(sheet[row_cells.pallet_count_cell].value or 0)
            for sheet in workbook.worksheets
            for row_cells in DESTINATION_ROWS
        )
        == pallet_result.totalFinalPallets
    )
    workbook.close()


def test_excel_report_writer_does_not_modify_template_file(tmp_path: Path) -> None:
    before = _sha256(DEFAULT_TEMPLATE_PATH)
    assert before == EXPECTED_TEMPLATE_SHA256
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)

    write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert _sha256(DEFAULT_TEMPLATE_PATH) == before


def test_excel_report_writer_preserves_palletizing_standards_rich_text(
    tmp_path: Path,
) -> None:
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)
    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    template = load_workbook(DEFAULT_TEMPLATE_PATH, rich_text=True)
    generated = load_workbook(result.outputPath, rich_text=True)
    try:
        template_sheet = template["Sheet1"]
        generated_sheet = generated["Sheet1"]
        template_value = template_sheet["C21"].value
        generated_value = generated_sheet["C21"].value
        assert isinstance(template_value, CellRichText)
        assert isinstance(generated_value, CellRichText)
        template_runs = _standards_runs(DEFAULT_TEMPLATE_PATH)
        generated_runs = _standards_runs(result.outputPath)
        assert generated_runs == template_runs
        assert len(generated_runs) > 1
        assert "".join(run.text for run in generated_runs) == str(template_value)
        assert "".join(run.text for run in generated_runs).endswith("when stored.")
        assert "\n" in "".join(run.text for run in generated_runs)
        assert {run.font_size for run in generated_runs} == {"10", "11"}
        assert {run.font_name for run in generated_runs} == {"Arial", "宋体"}
        assert all(run.bold for run in generated_runs)

        generated_cell_xml = _standards_cell_xml(result.outputPath)
        assert generated_cell_xml.find(".//m:r", XML_NS) is not None
        assert generated_cell_xml.find(".//m:rPr", XML_NS) is not None

        assert template_sheet.calculate_dimension() == "B1:P25"
        assert (
            generated_sheet.calculate_dimension()
            == template_sheet.calculate_dimension()
        )
        assert {str(item) for item in generated_sheet.merged_cells.ranges} == {
            str(item) for item in template_sheet.merged_cells.ranges
        }
        assert "C21:I25" in {str(item) for item in generated_sheet.merged_cells.ranges}
        assert generated_sheet["C21"].alignment.vertical == "top"
        assert generated_sheet["C21"].alignment.wrap_text is True
        generated_standards_height = sum(
            generated_sheet.row_dimensions[row].height or 0 for row in range(21, 26)
        )
        template_standards_height = sum(
            template_sheet.row_dimensions[row].height or 0 for row in range(21, 26)
        )
        assert generated_standards_height > template_standards_height
        assert all(
            (generated_sheet.row_dimensions[row].height or 0)
            > (template_sheet.row_dimensions[row].height or 0)
            for row in range(21, 26)
        )
        assert {
            column: generated_sheet.column_dimensions[column].width
            for column in "CDEFGHI"
        } == {
            column: template_sheet.column_dimensions[column].width
            for column in "CDEFGHI"
        }
        page_layout = _page_layout(generated_sheet)
        assert page_layout["paperSize"] == 9
        assert page_layout["orientation"] == "landscape"
        assert page_layout["scale"] is None
        assert page_layout["fitToWidth"] == 1
        assert page_layout["fitToHeight"] == 1
        assert page_layout["fitToPage"] is True
        assert page_layout["autoPageBreaks"] is False
        assert page_layout["printArea"] == "'Sheet1'!$A$1:$P$25"
        assert generated_sheet.row_breaks.count == 0
        assert generated_sheet.col_breaks.count == 0

        # Business cells are still written while the untouched rich-text cell survives.
        assert generated_sheet["D1"].value == "2026-06-25"
        assert generated_sheet["H1"].value == "09:30"
        assert generated_sheet["K1"].value == "CAAU8011090"
        assert generated_sheet["N4"].value == "Private Address / QDCA2605058915"
        assert generated_sheet["P20"].value == 896
    finally:
        template.close()
        generated.close()


def test_excel_report_writer_preserves_editable_business_cell_contract(
    tmp_path: Path,
) -> None:
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)
    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    template = load_workbook(DEFAULT_TEMPLATE_PATH, rich_text=True)
    generated = load_workbook(result.outputPath, rich_text=True)
    try:
        template_sheet = template["Sheet1"]
        generated_sheet = generated["Sheet1"]
        assert generated_sheet.protection.sheet == template_sheet.protection.sheet
        assert generated_sheet.protection.sheet is False
        assert {str(item) for item in generated_sheet.merged_cells.ranges} == {
            str(item) for item in template_sheet.merged_cells.ranges
        }
        for column in "BCDEFGHIJKLMNOP":
            assert (
                generated_sheet.column_dimensions[column].hidden
                == template_sheet.column_dimensions[column].hidden
            )
            assert (
                generated_sheet.column_dimensions[column].width
                == template_sheet.column_dimensions[column].width
            )
        for row in range(1, 26):
            assert (
                generated_sheet.row_dimensions[row].hidden
                == template_sheet.row_dimensions[row].hidden
            )
        writer_owned_cells = {
            "D1",
            "H1",
            "K1",
            "D2",
            "P20",
            "C21",
            *(
                coordinate
                for row_cells in DESTINATION_ROWS
                for coordinate in (
                    row_cells.pallet_label_cell,
                    row_cells.destination_cell,
                    row_cells.pallet_count_cell,
                    row_cells.carton_count_cell,
                )
            ),
        }
        for row in range(1, 26):
            for column in range(2, 17):
                template_cell = template_sheet.cell(row=row, column=column)
                generated_cell = generated_sheet.cell(row=row, column=column)
                if (
                    template_cell.coordinate in writer_owned_cells
                    or isinstance(template_cell, MergedCell)
                    or isinstance(generated_cell, MergedCell)
                ):
                    continue
                assert generated_cell.style_id == template_cell.style_id
                assert (
                    generated_cell.protection.locked
                    == template_cell.protection.locked
                )
                assert (
                    generated_cell.protection.hidden
                    == template_cell.protection.hidden
                )
    finally:
        template.close()
        generated.close()


def test_excel_report_writer_records_generated_report(tmp_path: Path) -> None:
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)

    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert result.manifestPath.is_file()
    manifest_text = result.manifestPath.read_text(encoding="utf-8")
    assert "CAAU8011090" in manifest_text
    assert str(result.outputPath) in manifest_text
    manifest = json.loads(manifest_text)
    assert manifest["records"][0]["expected_destination_count"] == 9
    assert manifest["records"][0]["written_destination_count"] == 9
    assert (
        manifest["records"][0]["ordered_destination_digest"]
        == result.orderedDestinationDigest
    )
    assert manifest["records"][0]["layout_modes"] == ["EXPANDED"]
    assert manifest["records"][0]["page_evidence"] == [
        {
            "page": 1,
            "layout_mode": "EXPANDED",
            "expected_destination_count": 9,
            "written_destination_count": 9,
            "expected_physical_rows": list(range(4, 13)),
            "written_physical_rows": list(range(4, 13)),
        }
    ]


def test_excel_report_writer_overwrites_same_container_report(
    tmp_path: Path,
) -> None:
    parsed, pallet_result = _parsed_and_pallets(STANDARD_FIXTURE, tmp_path)

    first = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )
    second = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 31),
    )

    assert first.outputPath == second.outputPath
    assert first.outputPath.is_file()
    manifest = json.loads(second.manifestPath.read_text(encoding="utf-8"))
    assert len(manifest["records"]) == 1
    assert manifest["records"][0]["generated_at"] == "2026-06-25T09:31:00"


def test_excel_report_writer_uses_all_rows_in_printed_order_before_overflow_sheets(
    tmp_path: Path,
) -> None:
    parsed, pallet_result = _parsed_and_pallets(OVERFLOW_FIXTURE, tmp_path)

    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert result.totalDestinationCount == result.writtenDestinationCount
    assert result.writtenDestinationCount > 8
    assert not any(
        warning.code == "DESTINATION_RANGE_EXCEEDED" for warning in result.warnings
    )
    assert result.outputPath.is_file()
    workbook = load_workbook(result.outputPath, data_only=False, rich_text=True)
    try:
        assert _populated_sheet_names(workbook) == ["Sheet1"]
        assert tuple(
            workbook["Sheet1"][f"N{row}"].value
            for row in range(4, 4 + len(pallet_result.plans))
        ) == tuple(plan.destinationCode for plan in pallet_result.plans)
        assert workbook["Sheet1"]["N5"].value is not None
        written_pallets = sum(
            int(sheet[row_cells.pallet_count_cell].value or 0)
            for sheet in workbook.worksheets
            for row_cells in DESTINATION_ROWS
        )
        assert written_pallets == pallet_result.totalFinalPallets
    finally:
        workbook.close()


def test_excel_report_writer_adds_a_sheet_only_after_all_white_rows_are_used(
    tmp_path: Path,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"EDGE-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, len(DESTINATION_ROWS) + 2)
    )

    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="OVERFLOW17"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    workbook = load_workbook(result.outputPath, data_only=False, rich_text=True)
    try:
        assert _populated_sheet_names(workbook) == ["Sheet1", "Sheet2"]
        assert workbook["Sheet1"]["N19"].value == "EDGE-16"
        assert workbook["Sheet2"]["N4"].value == "EDGE-17"
        assert isinstance(workbook["Sheet2"]["C21"].value, CellRichText)
        assert "C21:I25" in {
            str(item) for item in workbook["Sheet2"].merged_cells.ranges
        }
    finally:
        workbook.close()


def test_excel_report_writer_keeps_sixteen_short_destinations_on_one_sheet(
    tmp_path: Path,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"SHORT-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, len(DESTINATION_ROWS) + 1)
    )

    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="SHORT16"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    workbook = load_workbook(result.outputPath, data_only=False, rich_text=True)
    try:
        assert _populated_sheet_names(workbook) == ["Sheet1"]
        assert workbook["Sheet1"]["N19"].value == "SHORT-16"
    finally:
        workbook.close()


def test_excel_report_writer_does_not_paginate_normal_rows_before_capacity(
    tmp_path: Path,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=(
                f"LONG-{index:02d} "
                "Industrial Receiving Boulevard Appointment Required Door A"
            ),
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, len(DESTINATION_ROWS) + 1)
    )

    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="HEIGHTOVERFLOW"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert result.errors == ()
    workbook = load_workbook(result.outputPath, data_only=False, rich_text=True)
    try:
        populated = [
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.calculate_dimension() not in {"A1", "A1:A1"}
        ]
        assert len(populated) == 1
        written = [
            str(worksheet[row_cells.destination_cell].value)
            for worksheet in populated
            for row_cells in DESTINATION_ROWS
            if worksheet[row_cells.destination_cell].value is not None
        ]
        assert written == [plan.destinationCode for plan in plans]
        assert all(
            worksheet.print_area == f"'{worksheet.title}'!$A$1:$P$25"
            for worksheet in populated
        )
        assert all(worksheet.row_breaks.count == 0 for worksheet in populated)
        assert all(worksheet.col_breaks.count == 0 for worksheet in populated)
    finally:
        workbook.close()


def test_excel_report_writer_requires_review_for_extreme_destination_layout(
    tmp_path: Path,
) -> None:
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="TOOTALL"),
        pallet_result=SimpleNamespace(
            plans=(
                SimpleNamespace(
                    destinationCode="X" * 1000,
                    finalPallets=1,
                    totalCartons=10,
                ),
            )
        ),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    assert result.errors
    assert result.errors[0].code == "REPORT_LAYOUT_REVIEW_REQUIRED"
    assert result.errors[0].message == "REPORT_LAYOUT_REVIEW_REQUIRED"
    assert result.errors[0].stage == "planning.layout-review"
    assert result.errors[0].row == 4
    assert result.errors[0].requiredHeightPoints is not None
    assert result.errors[0].availableHeightPoints is not None
    assert not result.outputPath.exists()


@pytest.mark.parametrize(
    ("count", "expected_mode", "expected_rows"),
    (
        (0, DestinationLayoutMode.PRIMARY_ONLY, ()),
        (1, DestinationLayoutMode.PRIMARY_ONLY, (4,)),
        (2, DestinationLayoutMode.PRIMARY_ONLY, (4, 6)),
        (8, DestinationLayoutMode.PRIMARY_ONLY, (4, 6, 8, 10, 12, 14, 16, 18)),
        (9, DestinationLayoutMode.EXPANDED, tuple(range(4, 13))),
        (10, DestinationLayoutMode.EXPANDED, tuple(range(4, 14))),
        (16, DestinationLayoutMode.EXPANDED, tuple(range(4, 20))),
    ),
)
def test_destination_cell_map_selects_layout_before_assigning_rows(
    count: int,
    expected_mode: DestinationLayoutMode,
    expected_rows: tuple[int, ...],
) -> None:
    assert layout_mode_for_page_count(count) is expected_mode
    assert tuple(row.row for row in rows_for_page_count(count)) == expected_rows


@pytest.mark.parametrize("count", (-1, 17))
def test_destination_cell_map_rejects_invalid_page_count(count: int) -> None:
    with pytest.raises(ValueError):
        rows_for_page_count(count)


@pytest.mark.parametrize(
    ("count", "expected_page_rows", "expected_layout_modes"),
    (
        (0, ((),), ("PRIMARY_ONLY",)),
        (1, ((4,),), ("PRIMARY_ONLY",)),
        (2, ((4, 6),), ("PRIMARY_ONLY",)),
        (8, ((4, 6, 8, 10, 12, 14, 16, 18),), ("PRIMARY_ONLY",)),
        (9, (tuple(range(4, 13)),), ("EXPANDED",)),
        (10, (tuple(range(4, 14)),), ("EXPANDED",)),
        (16, (tuple(range(4, 20)),), ("EXPANDED",)),
        (17, (tuple(range(4, 20)), (4,)), ("EXPANDED", "PRIMARY_ONLY")),
        (
            24,
            (tuple(range(4, 20)), (4, 6, 8, 10, 12, 14, 16, 18)),
            ("EXPANDED", "PRIMARY_ONLY"),
        ),
        (
            25,
            (tuple(range(4, 20)), tuple(range(4, 13))),
            ("EXPANDED", "EXPANDED"),
        ),
        (
            32,
            (tuple(range(4, 20)), tuple(range(4, 20))),
            ("EXPANDED", "EXPANDED"),
        ),
        (
            33,
            (tuple(range(4, 20)), tuple(range(4, 20)), (4,)),
            ("EXPANDED", "EXPANDED", "PRIMARY_ONLY"),
        ),
    ),
)
def test_excel_report_writer_uses_adaptive_rows_before_capacity_pagination(
    tmp_path: Path,
    count: int,
    expected_page_rows: tuple[tuple[int, ...], ...],
    expected_layout_modes: tuple[str, ...],
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"CAPACITY-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, count + 1)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo=f"CAPACITY{count}"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / f"reports-{count}",
        report_datetime=datetime(2026, 7, 28, 9, 30),
    )

    assert result.errors == ()
    assert result.writtenDestinationCount == count
    assert result.totalDestinationCount == count
    assert result.layoutModes == expected_layout_modes
    assert tuple(
        evidence.expectedPhysicalRows for evidence in result.pageEvidence
    ) == expected_page_rows
    assert tuple(
        evidence.writtenPhysicalRows for evidence in result.pageEvidence
    ) == expected_page_rows
    workbook = load_workbook(result.outputPath, data_only=False, rich_text=True)
    try:
        populated = workbook.worksheets[: len(expected_page_rows)]
        assert tuple(row.row for row in DESTINATION_ROWS) == tuple(range(4, 20))
        actual_page_rows = tuple(
            tuple(
                row.row
                for row in DESTINATION_ROWS
                if worksheet[row.destination_cell].value is not None
            )
            for worksheet in populated
        )
        assert actual_page_rows == expected_page_rows
        expected_offset = 0
        for worksheet, page_rows in zip(populated, expected_page_rows):
            assert tuple(
                worksheet[f"N{row}"].value for row in page_rows
            ) == tuple(
                plan.destinationCode
                for plan in plans[
                    expected_offset : expected_offset + len(page_rows)
                ]
            )
            expected_offset += len(page_rows)
            unused_rows = sorted(set(range(4, 20)) - set(page_rows))
            for row in unused_rows:
                assert all(
                    worksheet[f"{column}{row}"].value is None
                    for column in "CNOP"
                )
        for worksheet, page_rows in zip(
            populated[:-1], expected_page_rows[:-1]
        ):
            assert len(page_rows) == len(DESTINATION_ROWS)
            assert all(
                worksheet[row.destination_cell].value is not None
                for row in DESTINATION_ROWS
            )
        assert all(worksheet.protection.sheet is False for worksheet in populated)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("count", "expected_rows"),
    (
        (8, (4, 6, 8, 10, 12, 14, 16, 18)),
        (9, tuple(range(4, 13))),
    ),
)
def test_excel_report_writer_preserves_template_row_styles_for_each_layout_mode(
    tmp_path: Path,
    count: int,
    expected_rows: tuple[int, ...],
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"STYLE-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, count + 1)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo=f"STYLE{count}"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / f"style-{count}",
        report_datetime=datetime(2026, 7, 29, 9, 30),
    )

    template = load_workbook(DEFAULT_TEMPLATE_PATH, rich_text=True)
    generated = load_workbook(result.outputPath, rich_text=True)
    try:
        expected_sheet = template["Sheet1"]
        actual_sheet = generated["Sheet1"]
        assert actual_sheet.protection.sheet is False
        assert {
            str(item) for item in actual_sheet.merged_cells.ranges
        } == {str(item) for item in expected_sheet.merged_cells.ranges}
        for row in range(4, 20):
            for column in "CNOP":
                expected_cell = expected_sheet[f"{column}{row}"]
                actual_cell = actual_sheet[f"{column}{row}"]
                assert copy(actual_cell.fill) == copy(expected_cell.fill)
                assert copy(actual_cell.font) == copy(expected_cell.font)
                assert copy(actual_cell.border) == copy(expected_cell.border)
                assert actual_cell.number_format == expected_cell.number_format
                assert copy(actual_cell.protection) == copy(
                    expected_cell.protection
                )
            assert (
                actual_sheet.row_dimensions[row].hidden
                == expected_sheet.row_dimensions[row].hidden
            )
            if row not in expected_rows:
                assert (
                    actual_sheet.row_dimensions[row].height
                    == expected_sheet.row_dimensions[row].height
                )
                assert all(
                    actual_sheet[f"{column}{row}"].value is None
                    for column in "CNOP"
                )
        assert all(
            actual_sheet[f"N{row}"].value is not None for row in expected_rows
        )
        assert all(
            actual_sheet[row.destination_cell].value is None
            for row in ADDITIONAL_DESTINATION_ROWS
        ) is (count == 8)
    finally:
        template.close()
        generated.close()


@pytest.mark.parametrize("count", (8, 9))
@pytest.mark.parametrize(
    "destination",
    (
        "Private Address / Industrial Receiving Calgary Dock Door",
        "卡尔加里仓超长中文收货地址工业园区第八大道",
        "YYC4 Receiving\nDoor A Appointment",
        "LONG-" + ("X" * 72) + "-TOKEN-END",
        None,
    ),
)
def test_adaptive_layout_preserves_text_boundaries_in_both_modes(
    tmp_path: Path,
    count: int,
    destination: str | None,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=destination if index == count else f"BOUNDARY-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, count + 1)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo=f"BOUNDARY{count}"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / f"boundary-{count}-{destination is None}",
        report_datetime=datetime(2026, 7, 29, 9, 30),
    )

    assert result.errors == ()
    expected_value = destination or "NEED_MANUAL_DESTINATION"
    target_row = rows_for_page_count(count)[-1].row
    workbook = load_workbook(result.outputPath, rich_text=True)
    try:
        worksheet = workbook["Sheet1"]
        assert worksheet[f"C{target_row}"].value == expected_value
        assert worksheet[f"N{target_row}"].value == expected_value
        assert worksheet[f"O{target_row}"].value == count
        assert worksheet[f"P{target_row}"].value == count * 10
        assert worksheet[f"N{target_row}"].alignment.wrap_text is True
    finally:
        workbook.close()
    assert any(issue.code == "MISSING_DESTINATION" for issue in result.warnings) is (
        destination is None
    )


def test_excel_report_writer_preserves_duplicate_destination_occurrences(
    tmp_path: Path,
) -> None:
    plans = (
        SimpleNamespace(
            destinationCode="DUPLICATE-DEST",
            finalPallets=1,
            totalCartons=11,
        ),
        SimpleNamespace(
            destinationCode="DUPLICATE-DEST",
            finalPallets=2,
            totalCartons=22,
        ),
        SimpleNamespace(
            destinationCode="DUPLICATE-DEST",
            finalPallets=3,
            totalCartons=33,
        ),
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="DUPLICATE03"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 7, 28, 9, 30),
    )

    assert result.errors == ()
    workbook = load_workbook(result.outputPath, data_only=False)
    try:
        worksheet = workbook["Sheet1"]
        actual = [
            (
                worksheet[row.destination_cell].value,
                worksheet[row.pallet_count_cell].value,
                worksheet[row.carton_count_cell].value,
            )
            for row in PRIMARY_DESTINATION_ROWS[:3]
        ]
        assert actual == [
            ("DUPLICATE-DEST", 1, 11),
            ("DUPLICATE-DEST", 2, 22),
            ("DUPLICATE-DEST", 3, 33),
        ]
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("coordinate", "value", "expected_stage"),
    (
        ("N6", "WRONG-DESTINATION", "reopen.mirror"),
        ("O6", 999, "reopen.row"),
        ("P6", 999, "reopen.row"),
        ("C6", "WRONG-MIRROR", "reopen.mirror"),
    ),
)
def test_saved_report_validator_fails_closed_for_mutated_business_rows(
    tmp_path: Path,
    coordinate: str,
    value: object,
    expected_stage: str,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"VALIDATE-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, 4)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="VALIDATE03"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 7, 28, 9, 30),
    )
    workbook = load_workbook(result.outputPath)
    try:
        workbook["Sheet1"][coordinate] = value
        workbook.save(result.outputPath)
    finally:
        workbook.close()

    template = load_workbook(DEFAULT_TEMPLATE_PATH, rich_text=True)
    try:
        page_plans, planning_issue = report_writer._plan_report_pages(
            template["Sheet1"], plans
        )
    finally:
        template.close()
    assert planning_issue is None
    _, _, issue = report_writer._validate_saved_report(
        result.outputPath,
        expected_plans=report_writer._canonical_plan_identities(plans),
        page_plans=page_plans,
    )
    assert issue is not None
    assert issue.code == "REPORT_DESTINATION_CONSERVATION_FAILED"
    assert issue.stage == expected_stage


def test_saved_report_validator_rejects_residual_white_row_in_primary_mode(
    tmp_path: Path,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"PRIMARY-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, 9)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="RESIDUALWHITE"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 7, 29, 9, 30),
    )
    workbook = load_workbook(result.outputPath)
    try:
        for column, value in zip("CNOP", ("RESIDUAL", "RESIDUAL", 1, 1)):
            workbook["Sheet1"][f"{column}5"] = value
        workbook.save(result.outputPath)
    finally:
        workbook.close()

    page_plans = _page_plans(plans)
    _, _, issue = report_writer._validate_saved_report(
        result.outputPath,
        expected_plans=report_writer._canonical_plan_identities(plans),
        page_plans=page_plans,
    )
    assert issue is not None
    assert issue.stage == "reopen.unused-row"
    assert issue.row == 5


@pytest.mark.parametrize("mutation", ("mode", "physical-row"))
def test_saved_report_validator_rejects_corrupted_page_plan_contract(
    tmp_path: Path,
    mutation: str,
) -> None:
    plans = tuple(
        SimpleNamespace(
            destinationCode=f"PLAN-{index:02d}",
            finalPallets=index,
            totalCartons=index * 10,
        )
        for index in range(1, 3)
    )
    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="PLANCONTRACT"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / mutation,
        report_datetime=datetime(2026, 7, 29, 9, 30),
    )
    page_plan = _page_plans(plans)[0]
    if mutation == "mode":
        corrupted = replace(
            page_plan,
            layoutMode=DestinationLayoutMode.EXPANDED,
        )
    else:
        corrupted = replace(
            page_plan,
            assignments=(
                page_plan.assignments[0],
                replace(
                    page_plan.assignments[1],
                    rowCells=EXPANDED_DESTINATION_ROWS[1],
                ),
            ),
        )

    _, _, issue = report_writer._validate_saved_report(
        result.outputPath,
        expected_plans=report_writer._canonical_plan_identities(plans),
        page_plans=(corrupted,),
    )
    assert issue is not None
    assert issue.stage == "reopen.page-plan"


def test_conservation_failure_does_not_replace_previous_success(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    plans = (
        SimpleNamespace(
            destinationCode="SAFE-OLD",
            finalPallets=1,
            totalCartons=10,
        ),
    )
    kwargs = {
        "parsed_result": SimpleNamespace(containerNo="ATOMIC03"),
        "pallet_result": SimpleNamespace(plans=plans),
        "output_dir": tmp_path / "reports",
    }
    first = write_excel_report(
        **kwargs,
        report_datetime=datetime(2026, 7, 28, 9, 30),
    )
    old_bytes = first.outputPath.read_bytes()
    old_manifest = first.manifestPath.read_bytes()
    monkeypatch.setattr(
        report_writer,
        "_validate_saved_report",
        lambda *_args, **_kwargs: (
            0,
            (),
            report_writer._conservation_issue(
                stage="reopen.test-mutation",
                expected_count=1,
                actual_count=0,
            ),
        ),
    )

    failed = write_excel_report(
        **kwargs,
        report_datetime=datetime(2026, 7, 28, 9, 31),
    )

    assert failed.errors[0].code == "REPORT_DESTINATION_CONSERVATION_FAILED"
    assert first.outputPath.read_bytes() == old_bytes
    assert first.manifestPath.read_bytes() == old_manifest
    assert not list((tmp_path / "reports").glob(".*.xlsx"))


def test_excel_report_writer_marks_missing_bestar_destination_for_manual_entry(
    tmp_path: Path,
) -> None:
    registry = ImportRegistry(tmp_path / "original_files")
    imported = registry.import_file(BESTAR_FIXTURE)
    parsed = parse_bestar_receiving(imported.stored_path)
    pallet_result = calculate_pallets(
        inputs_from_destination_summaries(parsed.destinationSummaries),
        container_no=parsed.containerNo,
    )

    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    workbook = load_workbook(result.outputPath, data_only=False)
    worksheet = workbook["Sheet1"]
    assert worksheet["N4"].value == "NEED_MANUAL_DESTINATION"
    assert any(warning.code == "MISSING_DESTINATION" for warning in result.warnings)
    workbook.close()


def test_excel_report_writer_auto_expands_destination_row_height(
    tmp_path: Path,
) -> None:
    long_destination = "Private Address / SZCA2604054725 / Surrey"
    parsed = SimpleNamespace(containerNo="AUTOROW123")
    pallet_result = SimpleNamespace(
        plans=(
            SimpleNamespace(
                destinationCode=long_destination,
                finalPallets=1,
                totalCartons=12,
            ),
        )
    )

    result = write_excel_report(
        parsed_result=parsed,
        pallet_result=pallet_result,
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    template = load_workbook(DEFAULT_TEMPLATE_PATH, data_only=False)
    template_height = template["Sheet1"].row_dimensions[4].height
    template.close()

    workbook = load_workbook(result.outputPath, data_only=False)
    worksheet = workbook["Sheet1"]
    assert worksheet["N4"].value == long_destination
    assert worksheet["C4"].value == long_destination
    assert worksheet["N4"].alignment.wrap_text is True
    assert worksheet["C4"].alignment.wrap_text is True
    assert worksheet.row_dimensions[4].height > (template_height or 0) + 40
    workbook.close()


def test_excel_report_writer_wraps_long_destination_in_white_business_row(
    tmp_path: Path,
) -> None:
    long_destination = "Private Address / SZCA2604054725 / Surrey Receiving Door"
    plans = tuple(
        SimpleNamespace(
            destinationCode=(long_destination if index == 2 else f"DEST-{index}"),
            finalPallets=1,
            totalCartons=12,
        )
        for index in range(1, 10)
    )

    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="WHITEOVERFLOW9"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    workbook = load_workbook(result.outputPath, data_only=False)
    try:
        worksheet = workbook["Sheet1"]
        assert _populated_sheet_names(workbook) == ["Sheet1"]
        assert worksheet["N5"].value == long_destination
        assert worksheet["C5"].value == long_destination
        assert worksheet["N5"].alignment.wrap_text is True
        assert worksheet["C5"].alignment.wrap_text is True
        assert worksheet.row_dimensions[5].height > 16
        assert worksheet["C21"].value is not None
    finally:
        workbook.close()


def test_excel_report_writer_preserves_line_break_and_expands_white_business_row(
    tmp_path: Path,
) -> None:
    destination = "YYC4\nDoor A"
    plans = tuple(
        SimpleNamespace(
            destinationCode=(destination if index == 2 else f"DEST-{index}"),
            finalPallets=1,
            totalCartons=12,
        )
        for index in range(1, 10)
    )

    result = write_excel_report(
        parsed_result=SimpleNamespace(containerNo="LINEBREAK9"),
        pallet_result=SimpleNamespace(plans=plans),
        output_dir=tmp_path / "reports",
        report_datetime=datetime(2026, 6, 25, 9, 30),
    )

    workbook = load_workbook(result.outputPath, data_only=False)
    try:
        worksheet = workbook["Sheet1"]
        assert worksheet["N5"].value == destination
        assert worksheet["C5"].value == destination
        assert worksheet.row_dimensions[5].height >= 33
    finally:
        workbook.close()


def _parsed_and_pallets(fixture_path: Path, tmp_path: Path):
    registry = ImportRegistry(tmp_path / "original_files")
    imported = registry.import_file(fixture_path)
    parsed = parse_unloading_plan_cn(imported.stored_path)
    pallet_result = calculate_pallets(
        inputs_from_destination_summaries(parsed.destinationSummaries),
        container_no=parsed.containerNo,
    )
    return parsed, pallet_result


def _page_plans(plans: tuple[SimpleNamespace, ...]):
    template = load_workbook(DEFAULT_TEMPLATE_PATH, rich_text=True)
    try:
        page_plans, issue = report_writer._plan_report_pages(
            template["Sheet1"], plans
        )
    finally:
        template.close()
    assert issue is None
    return page_plans


def _populated_sheet_names(workbook) -> list[str]:
    return [
        worksheet.title
        for worksheet in workbook.worksheets
        if worksheet.calculate_dimension() not in {"A1", "A1:A1"}
    ]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class _NormalizedRun(SimpleNamespace):
    text: str
    font_name: str | None
    font_size: str | None
    bold: bool
    properties: tuple[tuple[str, tuple[tuple[str, str], ...], str], ...]

    def __eq__(self, other: object) -> bool:
        return isinstance(other, _NormalizedRun) and vars(self) == vars(other)


def _standards_runs(path: Path) -> tuple[_NormalizedRun, ...]:
    with ZipFile(path) as archive:
        cell = _standards_cell_xml_from_archive(archive)
        string_node = cell.find("m:is", XML_NS)
        if cell.attrib.get("t") == "s":
            value = cell.find("m:v", XML_NS)
            assert value is not None and value.text is not None
            shared_strings = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            string_node = shared_strings.findall("m:si", XML_NS)[int(value.text)]
        assert string_node is not None

        runs = string_node.findall("m:r", XML_NS)
        assert runs, "Standards cell must contain rich-text runs"
        normalized: list[_NormalizedRun] = []
        for run in runs:
            properties = run.find("m:rPr", XML_NS)
            assert properties is not None
            font_name = properties.find("m:rFont", XML_NS)
            font_size = properties.find("m:sz", XML_NS)
            bold = properties.find("m:b", XML_NS)
            normalized.append(
                _NormalizedRun(
                    text="".join(
                        node.text or "" for node in run.findall("m:t", XML_NS)
                    ),
                    font_name=font_name.attrib.get("val")
                    if font_name is not None
                    else None,
                    font_size=font_size.attrib.get("val")
                    if font_size is not None
                    else None,
                    bold=bold is not None
                    and bold.attrib.get("val", "1") not in {"0", "false"},
                    properties=tuple(
                        sorted(_normalized_run_property(child) for child in properties)
                    ),
                )
            )
        return tuple(normalized)


def _standards_cell_xml(path: Path) -> ET.Element:
    with ZipFile(path) as archive:
        return ET.fromstring(ET.tostring(_standards_cell_xml_from_archive(archive)))


def _standards_cell_xml_from_archive(archive: ZipFile) -> ET.Element:
    worksheet = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    cell = worksheet.find(".//m:c[@r='C21']", XML_NS)
    assert cell is not None
    return cell


def _normalized_run_property(
    child: ET.Element,
) -> tuple[str, tuple[tuple[str, str], ...], str]:
    name = child.tag.rsplit("}", 1)[-1]
    attributes = dict(child.attrib)
    if name == "b" and "val" not in attributes:
        attributes["val"] = "1"
    return name, tuple(sorted(attributes.items())), child.text or ""


def _page_layout(worksheet) -> dict[str, object]:
    setup = worksheet.page_setup
    margins = worksheet.page_margins
    setup_properties = worksheet.sheet_properties.pageSetUpPr
    return {
        "paperSize": setup.paperSize,
        "orientation": setup.orientation,
        "scale": setup.scale,
        "fitToWidth": setup.fitToWidth,
        "fitToHeight": setup.fitToHeight,
        "fitToPage": setup_properties.fitToPage if setup_properties else None,
        "autoPageBreaks": setup_properties.autoPageBreaks if setup_properties else None,
        "margins": (
            margins.left,
            margins.right,
            margins.top,
            margins.bottom,
            margins.header,
            margins.footer,
        ),
        "printArea": str(worksheet.print_area),
    }
