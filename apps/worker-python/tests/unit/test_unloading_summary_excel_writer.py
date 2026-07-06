from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from worker_python.unloading_summary import write_unloading_summary_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
WORKFORM_TEMPLATE = REPO_ROOT / "samples" / "workform" / "Bestar_work_form.xlsx"


def test_unloading_summary_writer_generates_grouped_openable_workbook(
    tmp_path: Path,
) -> None:
    result = write_unloading_summary_workbook(
        payload={
            "month": "2026-06",
            "rows": [
                {
                    "sequence": 1,
                    "containerId": "container-1",
                    "containerNo": "BEAU5946301",
                    "dateBusinessTag": "6.1海柜",
                    "destinationText": "YYC4",
                    "quantityText": "40件 / 8托",
                    "referenceText": "124115028975",
                    "appointmentText": "06/03/2026 19:00 MDT",
                    "splitOrVarianceText": "28",
                    "operationNote": "Office reviewed",
                },
                {
                    "sequence": 1,
                    "containerId": "container-1",
                    "containerNo": "BEAU5946301",
                    "dateBusinessTag": "6.1海柜",
                    "destinationText": "YYC6",
                    "quantityText": "32件",
                    "referenceText": "86971028976",
                    "appointmentText": "06/02/2026 09:00 MDT",
                },
                {
                    "sequence": 2,
                    "containerId": "container-2",
                    "containerNo": "CAAU4743359",
                    "dateBusinessTag": "6.1美转加",
                    "destinationText": "UPS",
                    "quantityText": "6件 / 1托",
                    "referenceText": "shipment-1",
                    "appointmentText": "06/04/2026 13:00 MDT",
                },
            ],
            "reviewItems": [],
        },
        output_dir=tmp_path,
        generated_at=datetime(2026, 7, 6, 10, 30),
    )

    assert result.errors == ()
    assert result.outputPath.is_file()
    assert result.rowCount == 3
    assert result.sourceContainerCount == 2

    workbook = load_workbook(result.outputPath)
    sheet = workbook["6月拆柜数据"]
    assert sheet["A1"].value == "1、BEAU5946301"
    assert sheet["B1"].value == "6.1海柜"
    assert sheet["C1"].value == "YYC4"
    assert sheet["D1"].value == "40件 / 8托"
    assert sheet["E1"].value == "124115028975"
    assert sheet["F1"].value == "06/03/2026 19:00 MDT"
    assert sheet["A2"].value is None
    assert sheet["C2"].value == "YYC6"
    assert sheet["A4"].value == "2、CAAU4743359"
    assert sheet.column_dimensions["A"].width == 35.082
    workbook.close()


def test_unloading_summary_writer_adds_review_sheet_for_missing_fields(
    tmp_path: Path,
) -> None:
    result = write_unloading_summary_workbook(
        payload={
            "month": "2026-06",
            "rows": [
                {
                    "sequence": 1,
                    "containerId": "container-1",
                    "containerNo": "BEAU5946301",
                    "dateBusinessTag": "6.1海柜",
                    "destinationText": "YYC4",
                    "quantityText": "40件 / 8托",
                },
            ],
            "reviewItems": [
                {
                    "code": "MISSING_UNLOADING_COMPLETED_AT",
                    "containerNo": "LOADED123",
                    "field": "completedAt",
                    "message": "Missing completed date.",
                }
            ],
        },
        output_dir=tmp_path,
    )

    assert {warning.code for warning in result.warnings} >= {
        "MISSING_REFERENCE_TEXT",
        "MISSING_APPOINTMENT_TEXT",
        "REVIEW_ITEMS_PRESENT",
    }
    workbook = load_workbook(result.outputPath)
    review = workbook["Review"]
    assert review["B2"].value == "MISSING_REFERENCE_TEXT"
    assert review["B4"].value == "REVIEW_ITEMS_PRESENT"
    assert review["B5"].value == "MISSING_UNLOADING_COMPLETED_AT"
    workbook.close()


def test_unloading_summary_writer_does_not_modify_workform_template(
    tmp_path: Path,
) -> None:
    before = _sha256(WORKFORM_TEMPLATE)

    write_unloading_summary_workbook(
        payload={
            "month": "2026-06",
            "rows": [],
            "reviewItems": [],
        },
        output_dir=tmp_path,
    )

    assert _sha256(WORKFORM_TEMPLATE) == before


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
