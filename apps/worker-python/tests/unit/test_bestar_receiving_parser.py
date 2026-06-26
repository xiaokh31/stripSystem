from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from worker_python.imports import ImportRegistry
from worker_python.parser import FormatType, parse_bestar_receiving


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
BESTAR_RECEIVING_FIXTURE = FIXTURE_DIR / "137675 JXJU3246131  PO#3404  BESTAR.xlsx"
UNLOADING_PLAN_CN_FIXTURE = FIXTURE_DIR / "CAAU8011090 UNLOADING PLAN.xlsx"


def test_parse_bestar_receiving_extracts_header_and_item_lines(tmp_path: Path) -> None:
    imported = ImportRegistry(tmp_path / "original_files").import_file(BESTAR_RECEIVING_FIXTURE)

    result = parse_bestar_receiving(imported.stored_path)

    assert result.formatType == FormatType.BESTAR_RECEIVING
    assert result.containerNo == "JXJU3246131"
    assert result.poNumber == "3404"
    assert result.customer == "GILLYBOO"
    assert result.clearOrderNo == "137675"
    assert result.errors == ()
    assert len(result.lines) == 3

    first = result.lines[0]
    assert first.rowNumber == 12
    assert first.itemNo == "FSHTSCIS"
    assert first.description is None
    assert first.totalCartons == 267
    assert first.totalSkidCount is None
    assert first.raw_json["ITEM#"] == "FSHTSCIS "
    assert "TOTAL # OF CARTONS" in first.raw_json


def test_parse_bestar_receiving_marks_destination_as_manual(tmp_path: Path) -> None:
    imported = ImportRegistry(tmp_path / "original_files").import_file(BESTAR_RECEIVING_FIXTURE)

    result = parse_bestar_receiving(imported.stored_path)

    assert any(warning.code == "NEED_MANUAL_DESTINATION" for warning in result.warnings)
    assert len(result.destinationSummaries) == 1
    summary = result.destinationSummaries[0]
    assert summary.destinationCode is None
    assert summary.status == "NEED_MANUAL_DESTINATION"
    assert summary.totalCartons == 1267
    assert summary.totalSkidCount is None
    assert summary.lineCount == 3


def test_parse_bestar_receiving_skips_total_row_without_double_counting(
    tmp_path: Path,
) -> None:
    imported = ImportRegistry(tmp_path / "original_files").import_file(BESTAR_RECEIVING_FIXTURE)

    result = parse_bestar_receiving(imported.stored_path)

    assert [line.totalCartons for line in result.lines] == [267, 600, 400]
    assert sum(line.totalCartons or 0 for line in result.lines) == 1267
    assert any(warning.code == "SUMMARY_ROW_SKIPPED" for warning in result.warnings)


def test_parse_bestar_receiving_reports_missing_container(tmp_path: Path) -> None:
    source_copy = tmp_path / "bestar-receiving-no-container.xlsx"
    source_copy.write_bytes(BESTAR_RECEIVING_FIXTURE.read_bytes())
    workbook = load_workbook(source_copy)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet["D3"] = None
    workbook.save(source_copy)
    workbook.close()
    imported = ImportRegistry(tmp_path / "original_files").import_file(source_copy)

    result = parse_bestar_receiving(imported.stored_path)

    assert result.formatType == FormatType.BESTAR_RECEIVING
    assert result.containerNo is None
    assert any(error.code == "MISSING_CONTAINER_NO" for error in result.errors)
    assert result.lines


def test_parse_bestar_receiving_rejects_chinese_unloading_plan(tmp_path: Path) -> None:
    imported = ImportRegistry(tmp_path / "original_files").import_file(UNLOADING_PLAN_CN_FIXTURE)

    result = parse_bestar_receiving(imported.stored_path)

    assert result.formatType == FormatType.UNLOADING_PLAN_CN
    assert result.lines == ()
    assert any(error.code == "UNSUPPORTED_FORMAT" for error in result.errors)
