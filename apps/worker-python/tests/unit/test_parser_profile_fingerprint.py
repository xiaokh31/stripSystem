from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from worker_python.parser_profiles import (
    FingerprintDefinition,
    build_structural_fingerprint,
    inspect_workbook,
    rank_profile_matches,
)


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE_DIR = REPO_ROOT / "samples" / "unloading-plans"
LAYOUT_A = FIXTURE_DIR / "CAAU8011090 UNLOADING PLAN.xlsx"
LAYOUT_A_OTHER_DATA = FIXTURE_DIR / "DRYU9800413 - Unloading Plan.xlsx"
LAYOUT_B = FIXTURE_DIR / "137675 JXJU3246131  PO#3404  BESTAR.xlsx"


def _definition(profile_id: str = "profile-a") -> dict:
    return {
        "profileId": profile_id,
        "algorithmVersion": "workbook-fingerprint-v1",
        "workbookType": "OOXML_XLSX",
        "sheet": {"name": "Sheet1"},
        "anchors": [
            {
                "value": "运单号",
                "required": True,
                "row": 6,
                "column": 1,
                "rowTolerance": 0,
                "columnTolerance": 0,
            },
            {
                "value": "箱数/件数",
                "required": True,
                "row": 6,
                "column": 4,
                "rowTolerance": 0,
                "columnTolerance": 0,
            },
            {
                "value": "体积",
                "required": True,
                "row": 6,
                "column": 6,
                "rowTolerance": 0,
                "columnTolerance": 0,
            },
        ],
        "requiredRelativeColumns": [
            {
                "anchor": "运单号",
                "header": "箱数/件数",
                "offset": 3,
                "expectedValueTypes": ["number"],
            },
            {
                "anchor": "运单号",
                "header": "体积",
                "offset": 5,
                "expectedValueTypes": ["number"],
            },
        ],
        "dataStart": {"rowOffsetFromHeader": 1},
    }


def test_fingerprint_is_stable_across_different_cargo_and_row_counts() -> None:
    definition = FingerprintDefinition.model_validate(_definition())
    first = build_structural_fingerprint(inspect_workbook(LAYOUT_A), definition)
    second = build_structural_fingerprint(
        inspect_workbook(LAYOUT_A_OTHER_DATA), definition
    )

    assert first.algorithmVersion == "workbook-fingerprint-v1"
    assert first.matched is True
    assert second.matched is True
    assert first.hash == second.hash
    assert first.hash.startswith("sha256:")
    assert all(
        reason.code == "FINGERPRINT_ANCHOR_MATCHED"
        for reason in first.reasons
        if reason.matched
    )


def test_fingerprint_rejects_different_layout_with_stable_drift_codes() -> None:
    result = build_structural_fingerprint(
        inspect_workbook(LAYOUT_B),
        FingerprintDefinition.model_validate(_definition()),
    )

    assert result.matched is False
    assert {reason.code for reason in result.reasons} >= {"FINGERPRINT_SHEET_MISSING"}


def test_fingerprint_rejects_incompatible_ooxml_type(tmp_path: Path) -> None:
    xlsm_path = tmp_path / "layout.xlsm"
    xlsm_path.write_bytes(LAYOUT_A.read_bytes())
    result = build_structural_fingerprint(
        inspect_workbook(xlsm_path),
        FingerprintDefinition.model_validate(_definition()),
    )

    assert result.matched is False
    assert any(
        reason.code == "FINGERPRINT_WORKBOOK_TYPE_MISMATCH" for reason in result.reasons
    )


def test_required_anchor_movement_is_drift_not_filename_or_data_match(
    tmp_path: Path,
) -> None:
    moved = tmp_path / "CAAU8011090 UNLOADING PLAN.xlsx"
    workbook = load_workbook(LAYOUT_A)
    sheet = workbook["Sheet1"]
    sheet["A7"] = sheet["A6"].value
    sheet["A6"] = None
    workbook.save(moved)
    workbook.close()

    result = build_structural_fingerprint(
        inspect_workbook(moved),
        FingerprintDefinition.model_validate(_definition()),
    )

    assert result.matched is False
    assert any(
        reason.code == "FINGERPRINT_REQUIRED_ANCHOR_MISSING"
        for reason in result.reasons
    )
    assert "CAAU8011090" not in result.hash


def test_profile_ranking_is_stable_and_collision_never_selects_a_winner() -> None:
    inspection = inspect_workbook(LAYOUT_A)
    candidates = rank_profile_matches(
        inspection,
        [
            FingerprintDefinition.model_validate(_definition("profile-z")),
            FingerprintDefinition.model_validate(_definition("profile-a")),
        ],
    )

    assert [candidate.profileId for candidate in candidates.candidates] == [
        "profile-a",
        "profile-z",
    ]
    assert candidates.selectedProfileId is None
    assert candidates.issueCode == "FINGERPRINT_PROFILE_COLLISION"


def test_collision_never_uses_anchor_count_to_choose_a_winner() -> None:
    inspection = inspect_workbook(LAYOUT_A)
    stronger = _definition("profile-stronger")
    weaker = _definition("profile-weaker")
    weaker["anchors"] = weaker["anchors"][:2]
    weaker["requiredRelativeColumns"] = weaker["requiredRelativeColumns"][:1]

    candidates = rank_profile_matches(
        inspection,
        [
            FingerprintDefinition.model_validate(stronger),
            FingerprintDefinition.model_validate(weaker),
        ],
    )

    assert all(candidate.matched for candidate in candidates.candidates)
    assert candidates.selectedProfileId is None
    assert candidates.issueCode == "FINGERPRINT_PROFILE_COLLISION"


def test_fingerprint_rejects_declared_type_and_formula_cache_drift(
    tmp_path: Path,
) -> None:
    type_drift = tmp_path / "type-drift.xlsx"
    workbook = load_workbook(LAYOUT_A)
    sheet = workbook["Sheet1"]
    sheet["D7"] = "not-a-number"
    workbook.save(type_drift)
    workbook.close()

    typed_result = build_structural_fingerprint(
        inspect_workbook(type_drift),
        FingerprintDefinition.model_validate(_definition()),
    )
    assert typed_result.matched is False
    assert any(
        reason.code == "FINGERPRINT_COLUMN_TYPE_MISMATCH"
        for reason in typed_result.reasons
    )

    formula_drift = tmp_path / "formula-drift.xlsx"
    workbook = load_workbook(LAYOUT_A)
    sheet = workbook["Sheet1"]
    sheet["D7"] = "=1+1"
    workbook.save(formula_drift)
    workbook.close()
    definition = _definition()
    definition["requiredRelativeColumns"][0]["requireCachedFormula"] = True

    formula_result = build_structural_fingerprint(
        inspect_workbook(formula_drift),
        FingerprintDefinition.model_validate(definition),
    )
    assert formula_result.matched is False
    assert any(
        reason.code == "FINGERPRINT_FORMULA_CACHE_MISSING"
        for reason in formula_result.reasons
    )


def test_fingerprint_matches_data_markers_and_reports_stop_marker_drift(
    tmp_path: Path,
) -> None:
    with_stop = tmp_path / "with-stop.xlsx"
    workbook = load_workbook(LAYOUT_A)
    sheet = workbook["Sheet1"]
    sheet["A100"] = "PROFILE STOP"
    workbook.save(with_stop)
    workbook.close()
    definition = _definition()
    definition["dataStop"] = {"header": "运单号", "value": "PROFILE STOP"}

    matched = build_structural_fingerprint(
        inspect_workbook(with_stop),
        FingerprintDefinition.model_validate(definition),
    )
    assert matched.matched is True
    assert matched.structuralEvidence["dataStart"]["matched"] is True
    assert matched.structuralEvidence["dataStop"]["matched"] is True

    missing_stop = tmp_path / "missing-stop.xlsx"
    workbook = load_workbook(with_stop)
    workbook["Sheet1"]["A100"] = None
    workbook.save(missing_stop)
    workbook.close()
    drifted = build_structural_fingerprint(
        inspect_workbook(missing_stop),
        FingerprintDefinition.model_validate(definition),
    )
    assert drifted.matched is False
    assert any(
        reason.code == "FINGERPRINT_DATA_STOP_MISMATCH" for reason in drifted.reasons
    )
    assert drifted.hash != matched.hash


def test_fingerprint_reports_declared_data_start_without_a_data_row(
    tmp_path: Path,
) -> None:
    missing_start = tmp_path / "missing-start.xlsx"
    workbook = load_workbook(LAYOUT_A)
    sheet = workbook["Sheet1"]
    for cell in sheet[7]:
        cell.value = None
    workbook.save(missing_start)
    workbook.close()

    result = build_structural_fingerprint(
        inspect_workbook(missing_start),
        FingerprintDefinition.model_validate(_definition()),
    )
    assert result.matched is False
    assert any(
        reason.code == "FINGERPRINT_DATA_START_MISMATCH" for reason in result.reasons
    )
