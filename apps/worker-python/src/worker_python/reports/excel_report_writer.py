from __future__ import annotations

import json
import hashlib
import math
import os
import re
import shutil
import tempfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter

from worker_python.reports.cell_map import (
    COMPANY_VALUE_CELL,
    CONTAINER_VALUE_CELL,
    DATE_VALUE_CELL,
    DESTINATION_ROWS,
    SHEET_NAME,
    TIME_VALUE_CELL,
    TOTAL_CARTONS_CELL,
)
from worker_python.reports.row_layout import (
    MAX_ROW_HEIGHT_POINTS,
    CellLayoutInput,
    TextRun,
    calculate_cell_layout,
    calculate_row_layout,
    excel_column_width_to_points,
)
from worker_python.time_utils import operational_now


REPO_ROOT = Path(__file__).resolve().parents[5]
DEFAULT_TEMPLATE_PATH = REPO_ROOT / "samples" / "templates" / "卸柜报告-En.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "storage" / "reports"
REPORT_MANIFEST_FILENAME = "report_manifest.json"
DEFAULT_REPORT_ROW_HEIGHT = 16.5
CELL_HORIZONTAL_PADDING_POINTS = 8.0
REPORT_PRINT_AREA = "A1:P25"
STANDARDS_CELL = "C21"
MISSING_DESTINATION_PLACEHOLDER = "NEED_MANUAL_DESTINATION"


@dataclass(frozen=True)
class ExcelReportIssue:
    code: str
    message: str
    destinationCode: str | None = None
    sheet: str | None = None
    row: int | None = None
    stage: str | None = None
    ordinal: int | None = None
    expectedCount: int | None = None
    actualCount: int | None = None
    requiredHeightPoints: float | None = None
    availableHeightPoints: float | None = None


@dataclass(frozen=True)
class CanonicalPlanIdentity:
    ordinal: int
    destination: str
    finalPallets: int
    totalCartons: int


@dataclass(frozen=True)
class ExcelReportResult:
    outputPath: Path
    manifestPath: Path
    warnings: tuple[ExcelReportIssue, ...]
    errors: tuple[ExcelReportIssue, ...]
    writtenDestinationCount: int
    totalDestinationCount: int
    totalCartons: int
    orderedDestinationDigest: str


def write_excel_report(
    *,
    parsed_result: Any,
    pallet_result: Any,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    template_path: Path = DEFAULT_TEMPLATE_PATH,
    report_datetime: datetime | None = None,
    company: str = "Bestar",
) -> ExcelReportResult:
    warnings: list[ExcelReportIssue] = []
    errors: list[ExcelReportIssue] = []
    report_datetime = report_datetime or operational_now()

    if not template_path.is_file():
        errors.append(
            ExcelReportIssue(
                code="MISSING_TEMPLATE",
                message=f"Excel report template does not exist: {template_path}",
            )
        )
        return _error_result(output_dir, warnings, errors)

    container_no = getattr(parsed_result, "containerNo", None)
    if not container_no:
        warnings.append(
            ExcelReportIssue(
                code="MISSING_CONTAINER_NO",
                message="Container number is missing; report filename uses UNKNOWN-CONTAINER.",
            )
        )
        container_no = "UNKNOWN-CONTAINER"

    plans = tuple(getattr(pallet_result, "plans", ()))
    expected_plans = _canonical_plan_identities(plans)
    expected_digest = _ordered_plan_digest(expected_plans)
    if not plans:
        warnings.append(
            ExcelReportIssue(
                code="NO_DESTINATION_PLANS",
                message="No pallet plans were provided for report rows.",
            )
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{_safe_filename(container_no)}卸柜报告-En.xlsx"
    manifest_path = output_dir / REPORT_MANIFEST_FILENAME
    temporary_output_path: Path | None = None

    # Preserve every untouched rich-text template cell when saving the report.
    workbook = load_workbook(template_path, rich_text=True)
    try:
        first_sheet = workbook[SHEET_NAME]
        _prepare_report_sheet(
            first_sheet,
            report_datetime=report_datetime,
            container_no=container_no,
            company=company,
        )
        page_plans, layout_error = _plan_report_pages(first_sheet, plans)
        if layout_error is not None:
            errors.append(layout_error)
            return _error_result(
                output_dir,
                warnings,
                errors,
                output_path=output_path,
                total_destination_count=len(expected_plans),
                ordered_destination_digest=expected_digest,
            )

        worksheets = _report_worksheets(workbook, len(page_plans))
        for worksheet, plans_for_sheet in zip(worksheets, page_plans):
            _configure_page_contract(worksheet)
            _write_destination_rows(worksheet, plans_for_sheet, warnings)
            worksheet[TOTAL_CARTONS_CELL] = sum(
                int(getattr(plan, "totalCartons", 0) or 0) for plan in plans_for_sheet
            )
            _apply_written_row_layout(
                worksheet,
                row=worksheet[TOTAL_CARTONS_CELL].row,
                coordinates=(TOTAL_CARTONS_CELL,),
            )
        total_cartons = sum(
            int(getattr(plan, "totalCartons", 0) or 0) for plan in plans
        )
        worksheets[0][TOTAL_CARTONS_CELL] = total_cartons
        temporary_output_path = _temporary_path(
            output_dir,
            prefix=f".{_safe_filename(container_no)}-",
            suffix=".xlsx",
        )
        try:
            workbook.save(temporary_output_path)
        except Exception:
            temporary_output_path.unlink(missing_ok=True)
            raise
    finally:
        workbook.close()

    assert temporary_output_path is not None
    try:
        written_count, validation_error = _validate_saved_report(
            temporary_output_path,
            expected_plans=expected_plans,
            page_plans=page_plans,
        )
        if validation_error is not None:
            errors.append(validation_error)
            return _error_result(
                output_dir,
                warnings,
                errors,
                output_path=output_path,
                total_destination_count=len(expected_plans),
                ordered_destination_digest=expected_digest,
            )

        _replace_report_and_manifest(
            temporary_output_path=temporary_output_path,
            output_path=output_path,
            manifest_path=manifest_path,
            template_path=template_path,
            container_no=container_no,
            report_datetime=report_datetime,
            company=company,
            warnings=warnings,
            expected_count=len(expected_plans),
            written_count=written_count,
            ordered_digest=expected_digest,
        )
    finally:
        temporary_output_path.unlink(missing_ok=True)

    return ExcelReportResult(
        outputPath=output_path,
        manifestPath=manifest_path,
        warnings=tuple(warnings),
        errors=tuple(errors),
        writtenDestinationCount=written_count,
        totalDestinationCount=len(expected_plans),
        totalCartons=total_cartons,
        orderedDestinationDigest=expected_digest,
    )


def _write_header(
    worksheet: Any,
    *,
    report_datetime: datetime,
    container_no: str,
    company: str,
) -> None:
    worksheet[DATE_VALUE_CELL] = report_datetime.date().isoformat()
    worksheet[TIME_VALUE_CELL] = report_datetime.strftime("%H:%M")
    worksheet[CONTAINER_VALUE_CELL] = container_no
    worksheet[COMPANY_VALUE_CELL] = company


def _prepare_report_sheet(
    worksheet: Any,
    *,
    report_datetime: datetime,
    container_no: str,
    company: str,
) -> None:
    _configure_page_contract(worksheet)
    _write_header(
        worksheet,
        report_datetime=report_datetime,
        container_no=container_no,
        company=company,
    )
    _apply_written_row_layout(
        worksheet,
        row=1,
        coordinates=(DATE_VALUE_CELL, TIME_VALUE_CELL, CONTAINER_VALUE_CELL),
    )
    _apply_written_row_layout(
        worksheet,
        row=2,
        coordinates=(COMPANY_VALUE_CELL,),
    )
    _ensure_merged_cell_height(worksheet, STANDARDS_CELL)


def _report_worksheets(workbook: Any, page_count: int) -> list[Any]:
    page_count = max(1, page_count)
    first_sheet = workbook[SHEET_NAME]
    if page_count == 1:
        return [first_sheet]

    for worksheet in list(workbook.worksheets):
        if worksheet is not first_sheet:
            workbook.remove(worksheet)

    worksheets = [first_sheet]
    for page_number in range(2, page_count + 1):
        copied = workbook.copy_worksheet(first_sheet)
        copied.title = f"Sheet{page_number}"
        worksheets.append(copied)
    return worksheets


def _plan_report_pages(
    worksheet: Any,
    plans: tuple[Any, ...],
) -> tuple[tuple[tuple[Any, ...], ...], ExcelReportIssue | None]:
    if not plans:
        return ((),), None

    for ordinal, plan in enumerate(plans, start=1):
        destination = str(
            getattr(plan, "destinationCode", None)
            or MISSING_DESTINATION_PLACEHOLDER
        )
        row_cells = DESTINATION_ROWS[(ordinal - 1) % len(DESTINATION_ROWS)]
        required_height = _destination_row_height(
            worksheet,
            row_cells,
            destination=destination,
            final_pallets=int(getattr(plan, "finalPallets", 0) or 0),
            total_cartons=int(getattr(plan, "totalCartons", 0) or 0),
        )
        if required_height > MAX_ROW_HEIGHT_POINTS:
            return (), ExcelReportIssue(
                code="REPORT_LAYOUT_REVIEW_REQUIRED",
                message="REPORT_LAYOUT_REVIEW_REQUIRED",
                sheet=worksheet.title,
                row=row_cells.row,
                stage="planning.layout-review",
                ordinal=ordinal,
                requiredHeightPoints=required_height,
                availableHeightPoints=MAX_ROW_HEIGHT_POINTS,
            )

    capacity = len(DESTINATION_ROWS)
    pages = tuple(
        tuple(plans[index : index + capacity])
        for index in range(0, len(plans), capacity)
    )
    expected = _canonical_plan_identities(plans)
    actual = _canonical_plan_identities(
        tuple(plan for page in pages for plan in page)
    )
    if actual != expected:
        return (), _conservation_issue(
            stage="planning",
            expected_count=len(expected),
            actual_count=len(actual),
        )
    if any(len(page) != capacity for page in pages[:-1]):
        return (), _conservation_issue(
            stage="planning.capacity",
            expected_count=capacity,
            actual_count=min((len(page) for page in pages[:-1]), default=0),
        )
    return pages, None


def _required_sheet_height(worksheet: Any, plans: tuple[Any, ...]) -> float:
    heights = {
        row: _row_height(worksheet, row) for row in range(1, worksheet.max_row + 1)
    }
    for row_cells, plan in zip(DESTINATION_ROWS, plans):
        destination = str(
            getattr(plan, "destinationCode", None) or "NEED_MANUAL_DESTINATION"
        )
        row_height = _destination_row_height(
            worksheet,
            row_cells,
            destination=destination,
            final_pallets=int(getattr(plan, "finalPallets", 0) or 0),
            total_cartons=int(getattr(plan, "totalCartons", 0) or 0),
        )
        heights[row_cells.row] = max(heights[row_cells.row], row_height)
    return math.ceil(sum(heights.values()) * 4.0) / 4.0


def _write_destination_rows(
    worksheet: Any,
    plans: tuple[Any, ...],
    warnings: list[ExcelReportIssue],
) -> None:
    for row_cells, plan in zip(DESTINATION_ROWS, plans):
        destination = getattr(plan, "destinationCode", None)
        if not destination:
            destination = MISSING_DESTINATION_PLACEHOLDER
            warnings.append(
                ExcelReportIssue(
                    code="MISSING_DESTINATION",
                    message="Destination is missing; report row requires manual destination.",
                )
            )

        final_pallets = int(getattr(plan, "finalPallets", 0) or 0)
        total_cartons = int(getattr(plan, "totalCartons", 0) or 0)

        worksheet[row_cells.pallet_label_cell] = destination
        worksheet[row_cells.destination_cell] = destination
        worksheet[row_cells.pallet_count_cell] = final_pallets
        worksheet[row_cells.carton_count_cell] = total_cartons
        _apply_destination_row_layout(
            worksheet,
            row_cells,
            destination,
            final_pallets=final_pallets,
            total_cartons=total_cartons,
        )


def _apply_destination_row_layout(
    worksheet: Any,
    row_cells: Any,
    destination: str,
    *,
    final_pallets: int,
    total_cartons: int,
) -> None:
    wrapped_cells = (
        worksheet[row_cells.pallet_label_cell],
        worksheet[row_cells.destination_cell],
    )
    for cell in wrapped_cells:
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        cell.alignment = alignment

    worksheet.row_dimensions[row_cells.row].height = _destination_row_height(
        worksheet,
        row_cells,
        destination=destination,
        final_pallets=final_pallets,
        total_cartons=total_cartons,
    )


def _destination_row_height(
    worksheet: Any,
    row_cells: Any,
    *,
    destination: str,
    final_pallets: int,
    total_cartons: int,
) -> float:
    coordinates_and_values = (
        (row_cells.pallet_label_cell, destination),
        (row_cells.destination_cell, destination),
        (row_cells.pallet_count_cell, str(final_pallets)),
        (row_cells.carton_count_cell, str(total_cartons)),
    )
    cells = tuple(
        _cell_layout_input(
            worksheet,
            coordinate,
            visible_value=value,
            force_wrap=coordinate
            in {row_cells.pallet_label_cell, row_cells.destination_cell},
        )
        for coordinate, value in coordinates_and_values
    )
    result = calculate_row_layout(
        cells,
        template_height_points=_row_height(worksheet, row_cells.row),
        maximum_height_points=MAX_ROW_HEIGHT_POINTS,
    )
    return max(
        result.required_height_points,
        *(cell.required_height_points for cell in result.cell_results),
    )


def _apply_written_row_layout(
    worksheet: Any,
    *,
    row: int,
    coordinates: tuple[str, ...],
) -> None:
    cells = tuple(
        _cell_layout_input(worksheet, coordinate) for coordinate in coordinates
    )
    worksheet.row_dimensions[row].height = calculate_row_layout(
        cells,
        template_height_points=_row_height(worksheet, row),
        maximum_height_points=MAX_ROW_HEIGHT_POINTS,
    ).required_height_points


def _ensure_merged_cell_height(worksheet: Any, coordinate: str) -> None:
    cell = worksheet[coordinate]
    alignment = copy(cell.alignment)
    alignment.vertical = "top"
    cell.alignment = alignment

    merged_range = _merged_range_for_coordinate(worksheet, coordinate)
    if merged_range is None:
        _apply_written_row_layout(
            worksheet,
            row=cell.row,
            coordinates=(coordinate,),
        )
        return

    rows = tuple(range(merged_range.min_row, merged_range.max_row + 1))
    existing_height = sum(_row_height(worksheet, row) for row in rows)
    required_height = calculate_cell_layout(
        _cell_layout_input(worksheet, coordinate)
    ).required_height_points
    if required_height <= existing_height:
        return

    # Excel and LibreOffice both calculate the printable text box for a
    # vertically merged cell from the participating rows. Concentrating the
    # complete growth in the final row can leave the text box at its old
    # height in Print Preview even though the worksheet reports a larger
    # aggregate height. Grow every row in the merge so both renderers observe
    # the same physical note region.
    growth_per_row = (required_height - existing_height) / len(rows)
    for row in rows:
        worksheet.row_dimensions[row].height = (
            math.ceil((_row_height(worksheet, row) + growth_per_row) * 4.0) / 4.0
        )


def _cell_layout_input(
    worksheet: Any,
    coordinate: str,
    *,
    visible_value: str | None = None,
    force_wrap: bool = False,
) -> CellLayoutInput:
    cell = worksheet[coordinate]
    if force_wrap:
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        cell.alignment = alignment
    value = cell.value if visible_value is None else visible_value
    font_size = float(cell.font.sz or 11.0)
    return CellLayoutInput(
        visible_value="" if value is None else str(value),
        printable_width_points=_cell_printable_width_points(worksheet, coordinate),
        font_name=cell.font.name,
        font_size=font_size,
        bold=bool(cell.font.bold),
        wrap_text=bool(cell.alignment.wrap_text or force_wrap),
        indent=float(cell.alignment.indent or 0.0),
        rotation=int(cell.alignment.textRotation or 0),
        runs=_rich_text_runs(cell.value, cell),
    )


def _rich_text_runs(value: Any, cell: Any) -> tuple[TextRun, ...]:
    if not isinstance(value, CellRichText):
        return ()
    runs: list[TextRun] = []
    for item in value:
        if isinstance(item, TextBlock):
            font = item.font
            runs.append(
                TextRun(
                    text=item.text,
                    font_name=font.rFont or cell.font.name,
                    font_size=float(font.sz or cell.font.sz or 11.0),
                    bold=bool(font.b),
                )
            )
        else:
            runs.append(
                TextRun(
                    text=str(item),
                    font_name=cell.font.name,
                    font_size=float(cell.font.sz or 11.0),
                    bold=bool(cell.font.bold),
                )
            )
    return tuple(runs)


def _cell_printable_width_points(worksheet: Any, coordinate: str) -> float:
    cell = worksheet[coordinate]
    merged_range = _merged_range_for_coordinate(worksheet, coordinate)
    if merged_range is None:
        columns = (cell.column,)
    else:
        columns = range(merged_range.min_col, merged_range.max_col + 1)
    width = sum(
        excel_column_width_to_points(
            float(
                worksheet.column_dimensions[get_column_letter(column)].width
                or worksheet.sheet_format.defaultColWidth
                or 8.43
            )
        )
        for column in columns
    )
    return max(width - CELL_HORIZONTAL_PADDING_POINTS, 1.0)


def _merged_range_for_coordinate(worksheet: Any, coordinate: str) -> Any | None:
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range
    return None


def _row_height(worksheet: Any, row: int) -> float:
    return float(
        worksheet.row_dimensions[row].height
        or worksheet.sheet_format.defaultRowHeight
        or DEFAULT_REPORT_ROW_HEIGHT
    )


def _configure_page_contract(worksheet: Any) -> None:
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    # Use one sizing model. Leaving the template's fixed scale beside fit
    # attributes makes Excel and LibreOffice select different effective widths.
    worksheet.page_setup.scale = None
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    worksheet.print_area = REPORT_PRINT_AREA
    worksheet.row_breaks = worksheet.row_breaks.__class__()
    worksheet.col_breaks = worksheet.col_breaks.__class__()


def _replace_report_and_manifest(
    *,
    temporary_output_path: Path,
    output_path: Path,
    manifest_path: Path,
    template_path: Path,
    container_no: str,
    report_datetime: datetime,
    company: str,
    warnings: list[ExcelReportIssue],
    expected_count: int,
    written_count: int,
    ordered_digest: str,
) -> None:
    manifest = _load_manifest(manifest_path)
    record = {
        "generated_at": report_datetime.isoformat(),
        "container_no": container_no,
        "company": company,
        "output_path": str(output_path),
        "template_path": str(template_path),
        "warnings": [warning.code for warning in warnings],
        "expected_destination_count": expected_count,
        "written_destination_count": written_count,
        "ordered_destination_digest": ordered_digest,
    }
    manifest["records"] = [
        existing
        for existing in manifest["records"]
        if existing.get("output_path") != str(output_path)
    ]
    manifest["records"].append(record)
    temporary_manifest_path = _temporary_path(
        manifest_path.parent,
        prefix=".report-manifest-",
        suffix=".json",
    )
    temporary_manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    backup_path: Path | None = None
    try:
        if output_path.exists():
            backup_path = _temporary_path(
                output_path.parent,
                prefix=f".{output_path.stem}-backup-",
                suffix=output_path.suffix,
            )
            shutil.copy2(output_path, backup_path)
        os.replace(temporary_output_path, output_path)
        try:
            os.replace(temporary_manifest_path, manifest_path)
        except Exception:
            if backup_path is None:
                output_path.unlink(missing_ok=True)
            else:
                os.replace(backup_path, output_path)
                backup_path = None
            raise
    finally:
        temporary_manifest_path.unlink(missing_ok=True)
        if backup_path is not None:
            backup_path.unlink(missing_ok=True)


def _load_manifest(manifest_path: Path) -> dict[str, Any]:
    if not manifest_path.exists():
        return {"schema_version": 1, "records": []}

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("schema_version") != 1:
        raise ValueError(f"Unsupported report manifest schema: {manifest_path}")
    if not isinstance(manifest.get("records"), list):
        raise ValueError(f"Report manifest records must be a list: {manifest_path}")
    return manifest


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", value).strip("-") or "UNKNOWN-CONTAINER"


def _canonical_plan_identities(
    plans: tuple[Any, ...],
) -> tuple[CanonicalPlanIdentity, ...]:
    return tuple(
        CanonicalPlanIdentity(
            ordinal=ordinal,
            destination=str(
                getattr(plan, "destinationCode", None)
                or MISSING_DESTINATION_PLACEHOLDER
            ),
            finalPallets=int(getattr(plan, "finalPallets", 0) or 0),
            totalCartons=int(getattr(plan, "totalCartons", 0) or 0),
        )
        for ordinal, plan in enumerate(plans, start=1)
    )


def _ordered_plan_digest(plans: tuple[CanonicalPlanIdentity, ...]) -> str:
    payload = [
        {
            "ordinal": plan.ordinal,
            "destination": plan.destination,
            "finalPallets": plan.finalPallets,
            "totalCartons": plan.totalCartons,
        }
        for plan in plans
    ]
    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()


def _validate_saved_report(
    path: Path,
    *,
    expected_plans: tuple[CanonicalPlanIdentity, ...],
    page_plans: tuple[tuple[Any, ...], ...],
) -> tuple[int, ExcelReportIssue | None]:
    workbook = load_workbook(path, rich_text=True, data_only=False)
    try:
        actual: list[CanonicalPlanIdentity] = []
        expected_page_count = max(1, len(page_plans))
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            sheet_rows: list[tuple[int, str, int, int]] = []
            for row_cells in DESTINATION_ROWS:
                destination = worksheet[row_cells.destination_cell].value
                pallets = worksheet[row_cells.pallet_count_cell].value
                cartons = worksheet[row_cells.carton_count_cell].value
                mirror = worksheet[row_cells.pallet_label_cell].value
                if (
                    destination is None
                    and pallets is None
                    and cartons is None
                    and mirror is None
                ):
                    continue
                if str(mirror or "") != str(destination or ""):
                    return len(actual), _conservation_issue(
                        stage="reopen.mirror",
                        expected_count=len(expected_plans),
                        actual_count=len(actual),
                        sheet=worksheet.title,
                        row=row_cells.row,
                        ordinal=len(actual) + 1,
                    )
                sheet_rows.append(
                    (
                        row_cells.row,
                        str(destination or ""),
                        int(pallets or 0),
                        int(cartons or 0),
                    )
                )

            if sheet_index >= expected_page_count and sheet_rows:
                return len(actual), _conservation_issue(
                    stage="reopen.extra-sheet",
                    expected_count=len(expected_plans),
                    actual_count=len(actual) + len(sheet_rows),
                    sheet=worksheet.title,
                    row=sheet_rows[0][0],
                    ordinal=len(actual) + 1,
                )

            for row, destination, pallets, cartons in sheet_rows:
                actual.append(
                    CanonicalPlanIdentity(
                        ordinal=len(actual) + 1,
                        destination=destination,
                        finalPallets=pallets,
                        totalCartons=cartons,
                    )
                )
                if (
                    len(actual) > len(expected_plans)
                    or actual[-1] != expected_plans[len(actual) - 1]
                ):
                    return len(actual), _conservation_issue(
                        stage="reopen.row",
                        expected_count=len(expected_plans),
                        actual_count=len(actual),
                        sheet=worksheet.title,
                        row=row,
                        ordinal=len(actual),
                    )

            if sheet_index < expected_page_count:
                expected_total = (
                    sum(plan.totalCartons for plan in expected_plans)
                    if sheet_index == 0
                    else sum(
                        int(getattr(plan, "totalCartons", 0) or 0)
                        for plan in page_plans[sheet_index]
                    )
                )
                if int(worksheet[TOTAL_CARTONS_CELL].value or 0) != expected_total:
                    return len(actual), _conservation_issue(
                        stage="reopen.total",
                        expected_count=expected_total,
                        actual_count=int(worksheet[TOTAL_CARTONS_CELL].value or 0),
                        sheet=worksheet.title,
                        row=worksheet[TOTAL_CARTONS_CELL].row,
                    )

        actual_tuple = tuple(actual)
        if actual_tuple != expected_plans:
            return len(actual_tuple), _conservation_issue(
                stage="reopen.count",
                expected_count=len(expected_plans),
                actual_count=len(actual_tuple),
            )
        if _ordered_plan_digest(actual_tuple) != _ordered_plan_digest(expected_plans):
            return len(actual_tuple), _conservation_issue(
                stage="reopen.digest",
                expected_count=len(expected_plans),
                actual_count=len(actual_tuple),
            )
        return len(actual_tuple), None
    finally:
        workbook.close()


def _conservation_issue(
    *,
    stage: str,
    expected_count: int,
    actual_count: int,
    sheet: str | None = None,
    row: int | None = None,
    ordinal: int | None = None,
) -> ExcelReportIssue:
    return ExcelReportIssue(
        code="REPORT_DESTINATION_CONSERVATION_FAILED",
        message="REPORT_DESTINATION_CONSERVATION_FAILED",
        stage=stage,
        expectedCount=expected_count,
        actualCount=actual_count,
        sheet=sheet,
        row=row,
        ordinal=ordinal,
    )


def _temporary_path(directory: Path, *, prefix: str, suffix: str) -> Path:
    descriptor, raw_path = tempfile.mkstemp(
        dir=directory,
        prefix=prefix,
        suffix=suffix,
    )
    os.close(descriptor)
    return Path(raw_path)


def _error_result(
    output_dir: Path,
    warnings: list[ExcelReportIssue],
    errors: list[ExcelReportIssue],
    *,
    output_path: Path | None = None,
    total_destination_count: int = 0,
    ordered_destination_digest: str = "",
) -> ExcelReportResult:
    return ExcelReportResult(
        outputPath=output_path
        or output_dir / "UNKNOWN-CONTAINER卸柜报告-En.xlsx",
        manifestPath=output_dir / REPORT_MANIFEST_FILENAME,
        warnings=tuple(warnings),
        errors=tuple(errors),
        writtenDestinationCount=0,
        totalDestinationCount=total_destination_count,
        totalCartons=0,
        orderedDestinationDigest=ordered_destination_digest,
    )
