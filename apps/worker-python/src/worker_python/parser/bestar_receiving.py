from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from worker_python.parser.detector import FormatType, detect_excel_format
from worker_python.parser.workbook_warnings import ignore_openpyxl_conditional_formatting_warning


PARSER_VERSION = "bestar-receiving-v1"
CONTAINER_PATTERN = re.compile(r"\b[A-Z]{4}\d{7}[A-Z]?\b")
SUMMARY_MARKERS = {"TOTAL", "SUMMARY"}
NEED_MANUAL_DESTINATION = "NEED_MANUAL_DESTINATION"


@dataclass(frozen=True)
class BestarParseIssue:
    code: str
    message: str
    row_number: int | None = None
    field: str | None = None


@dataclass(frozen=True)
class BestarReceivingLine:
    rowNumber: int
    itemNo: str | None
    description: str | None
    totalCartons: int | None
    totalSkidCount: int | None
    raw_json: dict[str, Any]


@dataclass(frozen=True)
class BestarDestinationSummary:
    destinationCode: str | None
    status: str
    totalCartons: int
    totalSkidCount: int | None
    lineCount: int


@dataclass(frozen=True)
class BestarReceivingParseResult:
    containerNo: str | None
    poNumber: str | None
    customer: str | None
    clearOrderNo: str | None
    formatType: FormatType
    confidence: float
    parserVersion: str
    lines: tuple[BestarReceivingLine, ...]
    destinationSummaries: tuple[BestarDestinationSummary, ...]
    warnings: tuple[BestarParseIssue, ...]
    errors: tuple[BestarParseIssue, ...]
    rawMetadata: dict[str, Any]


FIELD_ALIASES = {
    "itemNo": ("ITEM#", "ITEM #"),
    "description": ("DESCRIPTION",),
    "totalCartons": ("TOTAL # OF CARTONS",),
    "totalSkidCount": ("TOTAL SKID COUNT",),
}


def parse_bestar_receiving(path: Path) -> BestarReceivingParseResult:
    detection = detect_excel_format(path)
    metadata: dict[str, Any] = {
        "sourceFile": str(path),
        "detectorReason": detection.reason,
        "matchedSheet": detection.matched_sheet,
        "matchedRow": detection.matched_row,
    }

    if detection.format_type != FormatType.BESTAR_RECEIVING:
        return BestarReceivingParseResult(
            containerNo=None,
            poNumber=None,
            customer=None,
            clearOrderNo=None,
            formatType=detection.format_type,
            confidence=detection.confidence,
            parserVersion=PARSER_VERSION,
            lines=(),
            destinationSummaries=(),
            warnings=tuple(
                BestarParseIssue(code="DETECTOR_WARNING", message=warning)
                for warning in detection.warnings
            ),
            errors=(
                BestarParseIssue(
                    code="UNSUPPORTED_FORMAT",
                    message=(
                        "BESTAR_RECEIVING parser only accepts files detected as "
                        f"BESTAR_RECEIVING; got {detection.format_type}."
                    ),
                ),
            )
            + tuple(
                BestarParseIssue(code="DETECTOR_ERROR", message=error)
                for error in detection.errors
            ),
            rawMetadata=metadata,
        )

    if detection.matched_sheet is None or detection.matched_row is None:
        return _empty_error_result(
            path=path,
            detection_format=detection.format_type,
            confidence=detection.confidence,
            metadata=metadata,
            issue=BestarParseIssue(
                code="HEADER_NOT_FOUND",
                message="Detector did not return a sheet and header row for BESTAR_RECEIVING.",
            ),
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return _empty_error_result(
            path=path,
            detection_format=detection.format_type,
            confidence=detection.confidence,
            metadata=metadata,
            issue=BestarParseIssue(code="WORKBOOK_READ_FAILED", message=str(exc)),
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            worksheet = workbook[detection.matched_sheet]
            header_values = next(
                worksheet.iter_rows(
                    min_row=detection.matched_row,
                    max_row=detection.matched_row,
                    values_only=True,
                )
            )
            headers = _unique_headers(header_values)
            field_columns = _field_columns(headers)
            header_metadata = _metadata_values(worksheet, path)
            lines, warnings = _parse_lines(
                worksheet=worksheet,
                header_row=detection.matched_row,
                headers=headers,
                field_columns=field_columns,
            )
    finally:
        workbook.close()

    errors: list[BestarParseIssue] = []
    if header_metadata["containerNo"] is None:
        errors.append(
            BestarParseIssue(
                code="MISSING_CONTAINER_NO",
                message="Container number was not found in workbook content or filename.",
                field="containerNo",
            )
        )

    warnings.append(
        BestarParseIssue(
            code=NEED_MANUAL_DESTINATION,
            message="Bestar receiving report does not contain destination; manual destination is required.",
            field="destinationCode",
        )
    )

    metadata["containerSource"] = header_metadata["containerSource"]
    metadata["fieldColumns"] = field_columns

    return BestarReceivingParseResult(
        containerNo=header_metadata["containerNo"],
        poNumber=header_metadata["poNumber"],
        customer=header_metadata["customer"],
        clearOrderNo=header_metadata["clearOrderNo"],
        formatType=FormatType.BESTAR_RECEIVING,
        confidence=detection.confidence,
        parserVersion=PARSER_VERSION,
        lines=tuple(lines),
        destinationSummaries=tuple(_destination_summaries(lines)),
        warnings=tuple(warnings),
        errors=tuple(errors),
        rawMetadata=metadata,
    )


def _parse_lines(
    *,
    worksheet: Any,
    header_row: int,
    headers: tuple[str, ...],
    field_columns: dict[str, int],
) -> tuple[list[BestarReceivingLine], list[BestarParseIssue]]:
    lines: list[BestarReceivingLine] = []
    warnings: list[BestarParseIssue] = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if _is_empty_row(row):
            continue

        raw_json = _raw_json(headers, row)
        if _is_summary_row(raw_json):
            warnings.append(
                BestarParseIssue(
                    code="SUMMARY_ROW_SKIPPED",
                    message="Skipped Bestar total row so item cartons are not double counted.",
                    row_number=row_number,
                )
            )
            continue

        line, line_warnings = _parse_line(row_number, raw_json, field_columns)
        warnings.extend(line_warnings)
        if not _has_item_signal(line):
            warnings.append(
                BestarParseIssue(
                    code="UNPARSEABLE_ROW_SKIPPED",
                    message="Skipped a non-empty row without recognizable Bestar item fields.",
                    row_number=row_number,
                )
            )
            continue

        lines.append(line)

    return lines, warnings


def _parse_line(
    row_number: int,
    raw_json: dict[str, Any],
    field_columns: dict[str, int],
) -> tuple[BestarReceivingLine, list[BestarParseIssue]]:
    warnings: list[BestarParseIssue] = []
    total_cartons, cartons_warning = _number(raw_json, field_columns, "totalCartons", row_number)
    total_skids, skids_warning = _number(raw_json, field_columns, "totalSkidCount", row_number)
    warnings.extend(issue for issue in (cartons_warning, skids_warning) if issue)

    return (
        BestarReceivingLine(
            rowNumber=row_number,
            itemNo=_text(raw_json, field_columns, "itemNo"),
            description=_text(raw_json, field_columns, "description"),
            totalCartons=int(total_cartons) if total_cartons is not None else None,
            totalSkidCount=int(total_skids) if total_skids is not None else None,
            raw_json=raw_json,
        ),
        warnings,
    )


def _destination_summaries(lines: list[BestarReceivingLine]) -> list[BestarDestinationSummary]:
    total_cartons = sum(line.totalCartons or 0 for line in lines)
    skid_values = [line.totalSkidCount for line in lines if line.totalSkidCount is not None]

    return [
        BestarDestinationSummary(
            destinationCode=None,
            status=NEED_MANUAL_DESTINATION,
            totalCartons=total_cartons,
            totalSkidCount=sum(skid_values) if skid_values else None,
            lineCount=len(lines),
        )
    ]


def _metadata_values(worksheet: Any, path: Path) -> dict[str, str | None]:
    metadata = {
        "containerNo": None,
        "containerSource": None,
        "poNumber": None,
        "customer": None,
        "clearOrderNo": None,
    }
    label_map = {
        "CONTAINER#": "containerNo",
        "PO#": "poNumber",
        "CUSTOMER": "customer",
        "CLEARORDER#": "clearOrderNo",
    }

    for row in worksheet.iter_rows(max_row=20, values_only=True):
        cells = [_cell_text(value) for value in row]
        for index, cell in enumerate(cells):
            key = label_map.get(_normalize_label(cell))
            if key is None:
                continue

            value = _next_value(cells, index)
            if key == "containerNo":
                match = CONTAINER_PATTERN.search((value or cell).upper())
                if match:
                    metadata["containerNo"] = match.group(0)
                    metadata["containerSource"] = "content"
            else:
                metadata[key] = value

    if metadata["containerNo"] is None:
        match = CONTAINER_PATTERN.search(path.name.upper())
        if match:
            metadata["containerNo"] = match.group(0)
            metadata["containerSource"] = "filename"

    return metadata


def _field_columns(headers: tuple[str, ...]) -> dict[str, int]:
    normalized_headers = {_normalize_label(header): index for index, header in enumerate(headers)}
    result: dict[str, int] = {}

    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            index = normalized_headers.get(_normalize_label(alias))
            if index is not None:
                result[field_name] = index
                break

    return result


def _raw_json(headers: tuple[str, ...], row: tuple[Any, ...]) -> dict[str, Any]:
    width = max(len(headers), len(row))
    result: dict[str, Any] = {}

    for index in range(width):
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        value = row[index] if index < len(row) else None
        result[header] = _json_value(value)

    return result


def _unique_headers(header_values: tuple[Any, ...]) -> tuple[str, ...]:
    counts: dict[str, int] = {}
    headers: list[str] = []

    for index, value in enumerate(header_values, start=1):
        header = _cell_text(value) or f"column_{index}"
        count = counts.get(header, 0) + 1
        counts[header] = count
        headers.append(header if count == 1 else f"{header}__{count}")

    return tuple(headers)


def _number(
    raw_json: dict[str, Any],
    field_columns: dict[str, int],
    field_name: str,
    row_number: int,
) -> tuple[Decimal | None, BestarParseIssue | None]:
    value = _raw_field(raw_json, field_columns, field_name)
    if value in (None, ""):
        return None, None

    try:
        return Decimal(str(value).replace(",", "").strip()), None
    except (InvalidOperation, ValueError):
        return (
            None,
            BestarParseIssue(
                code="INVALID_NUMBER",
                message=f"Could not parse numeric field {field_name}: {value}",
                row_number=row_number,
                field=field_name,
            ),
        )


def _text(raw_json: dict[str, Any], field_columns: dict[str, int], field_name: str) -> str | None:
    value = _raw_field(raw_json, field_columns, field_name)
    text = _cell_text(value)
    if text in {"", "/"}:
        return None
    return text


def _raw_field(raw_json: dict[str, Any], field_columns: dict[str, int], field_name: str) -> Any:
    index = field_columns.get(field_name)
    if index is None:
        return None

    key = tuple(raw_json.keys())[index]
    return raw_json[key]


def _next_value(cells: list[str], label_index: int) -> str | None:
    for value in cells[label_index + 1 :]:
        if value:
            return value
    return None


def _is_summary_row(raw_json: dict[str, Any]) -> bool:
    return any(_normalize_label(str(value)) in SUMMARY_MARKERS for value in raw_json.values())


def _has_item_signal(line: BestarReceivingLine) -> bool:
    return any(
        value is not None
        for value in (
            line.itemNo,
            line.description,
            line.totalCartons,
            line.totalSkidCount,
        )
    )


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_cell_text(value) == "" for value in row)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _normalize_label(value: str) -> str:
    text = value.replace("＃", "#").replace("：", ":")
    return re.sub(r"\s+", "", text).upper()


def _empty_error_result(
    *,
    path: Path,
    detection_format: FormatType,
    confidence: float,
    metadata: dict[str, Any],
    issue: BestarParseIssue,
) -> BestarReceivingParseResult:
    metadata["sourceFile"] = str(path)
    return BestarReceivingParseResult(
        containerNo=None,
        poNumber=None,
        customer=None,
        clearOrderNo=None,
        formatType=detection_format,
        confidence=confidence,
        parserVersion=PARSER_VERSION,
        lines=(),
        destinationSummaries=(),
        warnings=(),
        errors=(issue,),
        rawMetadata=metadata,
    )
