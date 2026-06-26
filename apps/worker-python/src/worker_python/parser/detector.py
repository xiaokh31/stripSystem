from __future__ import annotations

import re
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path

from openpyxl import load_workbook


class FormatType(StrEnum):
    UNLOADING_PLAN_CN = "UNLOADING_PLAN_CN"
    BESTAR_RECEIVING = "BESTAR_RECEIVING"
    UNKNOWN = "UNKNOWN"


@dataclass(frozen=True)
class DetectionResult:
    format_type: FormatType
    confidence: float
    reason: str
    warnings: tuple[str, ...] = ()
    errors: tuple[str, ...] = ()
    matched_sheet: str | None = None
    matched_row: int | None = None
    matched_headers: tuple[str, ...] = ()


@dataclass(frozen=True)
class _ScannedRow:
    sheet_name: str
    row_number: int
    cells: tuple[str, ...]


@dataclass(frozen=True)
class _HeaderPattern:
    name: str
    terms: tuple[str, ...]
    required_terms: tuple[str, ...]
    minimum_matches: int


MAX_ROWS_PER_SHEET = 40

BESTAR_HEADER_TERMS = ("RECEIVING REPORT", "CONTAINER #", "PO#", "CUSTOMER", "CLEAR ORDER #")
BESTAR_DETAIL_TERMS = (
    "ITEM#",
    "DESCRIPTION",
    "TOTAL # OF CARTONS",
    "TOTAL SKID COUNT",
)

UNLOADING_PLAN_CN_PATTERNS = (
    _HeaderPattern(
        name="standard_cn_waybill",
        terms=(
            "运单号",
            "客户单号",
            "扩展单号",
            "转单号",
            "服务名称",
            "仓库代码",
            "PO Number",
            "收件人国家",
            "件数",
            "实际重量",
            "材积重",
            "体积",
        ),
        required_terms=("运单号", "件数", "体积"),
        minimum_matches=6,
    ),
    _HeaderPattern(
        name="delivery_plan_cn",
        terms=(
            "运单号",
            "FBA NO.",
            "PO#",
            "箱数/件数",
            "重量",
            "体积",
            "派送目的地",
            "派送方式",
        ),
        required_terms=("运单号", "箱数/件数", "体积"),
        minimum_matches=5,
    ),
    _HeaderPattern(
        name="receiving_dispatch_cn",
        terms=(
            "SO",
            "FBA",
            "Reference ID",
            "件数",
            "重量",
            "体积",
            "地址类型",
            "仓库代码",
            "仓库地址",
            "派送方式",
        ),
        required_terms=("SO", "件数", "体积"),
        minimum_matches=6,
    ),
)


def detect_excel_format(path: Path) -> DetectionResult:
    if not path.is_file():
        return DetectionResult(
            format_type=FormatType.UNKNOWN,
            confidence=0.0,
            reason=f"Excel file does not exist: {path}",
            errors=(f"Excel file does not exist: {path}",),
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return DetectionResult(
            format_type=FormatType.UNKNOWN,
            confidence=0.0,
            reason=f"Unable to read Excel workbook: {exc}",
            errors=(f"Unable to read Excel workbook: {exc}",),
        )

    try:
        scanned_rows = tuple(_scan_rows(workbook))
    finally:
        workbook.close()

    bestar_result = _detect_bestar_receiving(scanned_rows)
    if bestar_result is not None:
        return bestar_result

    cn_result = _detect_unloading_plan_cn(scanned_rows)
    if cn_result is not None:
        return cn_result

    return DetectionResult(
        format_type=FormatType.UNKNOWN,
        confidence=0.0,
        reason=f"No supported header pattern found in the first {MAX_ROWS_PER_SHEET} rows.",
        warnings=("Unsupported Excel format for Phase 0 parser detector.",),
    )


def _scan_rows(workbook) -> list[_ScannedRow]:  # noqa: ANN001
    rows: list[_ScannedRow] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index > MAX_ROWS_PER_SHEET:
                break

            cells = tuple(_normalize_cell(value) for value in row if _normalize_cell(value))
            if cells:
                rows.append(
                    _ScannedRow(
                        sheet_name=sheet_name,
                        row_number=row_index,
                        cells=cells,
                    )
                )

    return rows


def _detect_bestar_receiving(scanned_rows: tuple[_ScannedRow, ...]) -> DetectionResult | None:
    all_cells = tuple(cell for row in scanned_rows for cell in row.cells)
    header_matches = _matched_terms(BESTAR_HEADER_TERMS, all_cells)
    detail_matches = _matched_terms(BESTAR_DETAIL_TERMS, all_cells)

    if len(header_matches) < 4 or len(detail_matches) < 3:
        return None

    matched_row = _first_matching_row(scanned_rows, BESTAR_DETAIL_TERMS)
    matched_headers = header_matches + detail_matches
    confidence = min(0.99, len(matched_headers) / (len(BESTAR_HEADER_TERMS) + len(BESTAR_DETAIL_TERMS)))

    return DetectionResult(
        format_type=FormatType.BESTAR_RECEIVING,
        confidence=confidence,
        reason="Matched Bestar receiving report header and detail header area.",
        matched_sheet=matched_row.sheet_name if matched_row else None,
        matched_row=matched_row.row_number if matched_row else None,
        matched_headers=matched_headers,
    )


def _detect_unloading_plan_cn(scanned_rows: tuple[_ScannedRow, ...]) -> DetectionResult | None:
    best_match: tuple[_HeaderPattern, _ScannedRow, tuple[str, ...]] | None = None

    for pattern in UNLOADING_PLAN_CN_PATTERNS:
        for row in scanned_rows:
            matched_terms = _matched_terms(pattern.terms, row.cells)
            required_matches = _matched_terms(pattern.required_terms, row.cells)
            if (
                len(required_matches) == len(pattern.required_terms)
                and len(matched_terms) >= pattern.minimum_matches
            ):
                if best_match is None or len(matched_terms) > len(best_match[2]):
                    best_match = (pattern, row, matched_terms)

    if best_match is None:
        return None

    pattern, row, matched_terms = best_match
    confidence = min(0.99, len(matched_terms) / len(pattern.terms))

    return DetectionResult(
        format_type=FormatType.UNLOADING_PLAN_CN,
        confidence=confidence,
        reason=f"Matched Chinese unloading plan header pattern: {pattern.name}.",
        matched_sheet=row.sheet_name,
        matched_row=row.row_number,
        matched_headers=matched_terms,
    )


def _first_matching_row(
    scanned_rows: tuple[_ScannedRow, ...],
    terms: tuple[str, ...],
) -> _ScannedRow | None:
    for row in scanned_rows:
        if _matched_terms(terms, row.cells):
            return row
    return None


def _matched_terms(terms: tuple[str, ...], cells: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(term for term in terms if _has_term(cells, term))


def _has_term(cells: tuple[str, ...], term: str) -> bool:
    expected = _compact(term)
    return any(expected in _compact(cell) for cell in cells)


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    text = text.replace("＃", "#").replace("：", ":")
    return re.sub(r"\s+", " ", text).upper()


def _compact(value: str) -> str:
    return re.sub(r"\s+", "", _normalize_cell(value))
