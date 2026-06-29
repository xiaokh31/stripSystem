from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from worker_python.parser.detector import FormatType, detect_excel_format
from worker_python.parser.workbook_warnings import ignore_openpyxl_conditional_formatting_warning


PARSER_VERSION = "unloading-plan-cn-v1"
CONTAINER_PATTERN = re.compile(r"\b[A-Z]{4}\d{7}[A-Z]?\b")
MIN_VOLUME_CBM = 0.01
COURIER_DELIVERY_TERMS = (
    "快递",
    "快遞",
    "快递派送",
    "快遞派送",
    "COURIER",
    "EXPRESS",
    "PARCEL",
)
COURIER_CARRIER_TERMS = (
    "UPS",
    "PUROLATOR",
    "FEDEX",
    "CANPAR",
    "DHL",
    "CANADA POST",
    "CANADAPOST",
    "USPS",
    "INTELCOM",
    "UNIUNI",
)


@dataclass(frozen=True)
class ParseIssue:
    code: str
    message: str
    row_number: int | None = None
    field: str | None = None


@dataclass(frozen=True)
class ParsedLine:
    rowNumber: int
    waybillNo: str | None
    fbaNo: str | None
    poNumber: str | None
    cartons: int | None
    weight: float | None
    volumeCbm: float | None
    destinationCode: str | None
    deliveryMethod: str | None
    note: str | None
    raw_json: dict[str, Any]


@dataclass(frozen=True)
class DestinationSummary:
    destinationCode: str
    totalCartons: int
    totalVolumeCbm: float
    lineCount: int


@dataclass(frozen=True)
class UnloadingPlanParseResult:
    containerNo: str | None
    formatType: FormatType
    confidence: float
    parserVersion: str
    lines: tuple[ParsedLine, ...]
    destinationSummaries: tuple[DestinationSummary, ...]
    warnings: tuple[ParseIssue, ...]
    errors: tuple[ParseIssue, ...]
    rawMetadata: dict[str, Any]


FIELD_ALIASES = {
    "waybillNo": ("运单号", "SO"),
    "fbaNo": ("FBA NO.", "FBA", "客户单号"),
    "poNumber": ("PO#", "PO NUMBER", "REFERENCE ID"),
    "cartons": ("箱数/件数", "件数"),
    "weight": ("重量", "实际重量(KG)"),
    "volumeCbm": ("体积", "体积(M³)", "体积(M3)"),
    "destinationCode": ("派送目的地", "仓库代码"),
    "deliveryMethod": ("派送方式", "服务名称"),
    "note": ("备注", "特殊指令/备注", "内部备注"),
}

SUMMARY_MARKERS = ("汇总", "合计", "总计", "TOTAL")


def parse_unloading_plan_cn(path: Path) -> UnloadingPlanParseResult:
    detection = detect_excel_format(path)
    metadata: dict[str, Any] = {
        "sourceFile": str(path),
        "detectorReason": detection.reason,
        "matchedSheet": detection.matched_sheet,
        "matchedRow": detection.matched_row,
    }

    if detection.format_type != FormatType.UNLOADING_PLAN_CN:
        return UnloadingPlanParseResult(
            containerNo=None,
            formatType=detection.format_type,
            confidence=detection.confidence,
            parserVersion=PARSER_VERSION,
            lines=(),
            destinationSummaries=(),
            warnings=tuple(
                ParseIssue(code="DETECTOR_WARNING", message=warning)
                for warning in detection.warnings
            ),
            errors=(
                ParseIssue(
                    code="UNSUPPORTED_FORMAT",
                    message=(
                        "UNLOADING_PLAN_CN parser only accepts files detected as "
                        f"UNLOADING_PLAN_CN; got {detection.format_type}."
                    ),
                ),
            )
            + tuple(ParseIssue(code="DETECTOR_ERROR", message=error) for error in detection.errors),
            rawMetadata=metadata,
        )

    if detection.matched_sheet is None or detection.matched_row is None:
        return _empty_error_result(
            path=path,
            detection_format=detection.format_type,
            confidence=detection.confidence,
            metadata=metadata,
            issue=ParseIssue(
                code="HEADER_NOT_FOUND",
                message="Detector did not return a sheet and header row for UNLOADING_PLAN_CN.",
            ),
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return _empty_error_result(
            path=path,
            detection_format=detection.format_type,
            confidence=detection.confidence,
            metadata=metadata,
            issue=ParseIssue(code="WORKBOOK_READ_FAILED", message=str(exc)),
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            worksheet = workbook[detection.matched_sheet]
            header_values = next(
                worksheet.iter_rows(
                    min_row=detection.matched_row,
                    max_row=detection.matched_row,
                    values_only=True,
                )
            )
            raw_headers = _unique_headers(header_values)
            field_columns = _field_columns(raw_headers)
            container_no, container_source = _extract_container_no(workbook, path)
            lines, warnings = _parse_lines(
                worksheet=worksheet,
                header_row=detection.matched_row,
                raw_headers=raw_headers,
                field_columns=field_columns,
            )
    finally:
        workbook.close()

    errors: list[ParseIssue] = []
    if container_no is None:
        errors.append(
            ParseIssue(
                code="MISSING_CONTAINER_NO",
                message="Container number was not found in workbook content or filename.",
                field="containerNo",
            )
        )

    metadata["containerSource"] = container_source
    metadata["fieldColumns"] = field_columns

    return UnloadingPlanParseResult(
        containerNo=container_no,
        formatType=FormatType.UNLOADING_PLAN_CN,
        confidence=detection.confidence,
        parserVersion=PARSER_VERSION,
        lines=tuple(lines),
        destinationSummaries=tuple(_destination_summaries(lines)),
        warnings=tuple(warnings),
        errors=tuple(errors),
        rawMetadata=metadata,
    )


def _parse_lines(
    *,
    worksheet: Any,
    header_row: int,
    raw_headers: tuple[str, ...],
    field_columns: dict[str, int],
) -> tuple[list[ParsedLine], list[ParseIssue]]:
    lines: list[ParsedLine] = []
    warnings: list[ParseIssue] = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if _is_empty_row(row):
            continue

        raw_json = _raw_json(raw_headers, row)
        if _is_summary_row(raw_json):
            warnings.append(
                ParseIssue(
                    code="NON_DETAIL_ROW_SKIPPED",
                    message="Skipped a summary or total row after the header.",
                    row_number=row_number,
                )
            )
            continue

        line, line_warnings = _parse_line(row_number, raw_json, field_columns)
        warnings.extend(line_warnings)

        if not _has_detail_signal(line):
            warnings.append(
                ParseIssue(
                    code="UNPARSEABLE_ROW_SKIPPED",
                    message="Skipped a non-empty row without recognizable detail fields.",
                    row_number=row_number,
                )
            )
            continue

        lines.append(line)

    return lines, warnings


def _parse_line(
    row_number: int,
    raw_json: dict[str, Any],
    field_columns: dict[str, int],
) -> tuple[ParsedLine, list[ParseIssue]]:
    warnings: list[ParseIssue] = []

    cartons, cartons_warning = _number(raw_json, field_columns, "cartons", row_number)
    weight, weight_warning = _number(raw_json, field_columns, "weight", row_number)
    volume, volume_warning = _number(raw_json, field_columns, "volumeCbm", row_number)
    warnings.extend(issue for issue in (cartons_warning, weight_warning, volume_warning) if issue)

    cartons_int = int(cartons) if cartons is not None else None
    volume_float = float(volume) if volume is not None else None
    weight_float = float(weight) if weight is not None else None
    waybill_no = _text(raw_json, field_columns, "waybillNo")
    destination_code = _text(raw_json, field_columns, "destinationCode")
    delivery_method = _text(raw_json, field_columns, "deliveryMethod")
    note = _text(raw_json, field_columns, "note")

    if destination_code is None:
        warnings.append(
            ParseIssue(
                code="MISSING_DESTINATION",
                message="Destination code is missing.",
                row_number=row_number,
                field="destinationCode",
            )
        )
    if cartons_int is None or cartons_int == 0:
        warnings.append(
            ParseIssue(
                code="MISSING_CARTONS",
                message="Cartons are missing or zero.",
                row_number=row_number,
                field="cartons",
            )
        )
    if volume_float is None:
        warnings.append(
            ParseIssue(
                code="MISSING_VOLUME",
                message="Volume is missing.",
                row_number=row_number,
                field="volumeCbm",
            )
        )
    elif cartons_int is not None and cartons_int > 0 and volume_float == 0:
        warnings.append(
            ParseIssue(
                code="ZERO_VOLUME_WITH_CARTONS",
                message=f"第{row_number}行体积为0，共{cartons_int}箱，已按0.01 CBM参与托盘计算。",
                row_number=row_number,
                field="volumeCbm",
            )
        )
        volume_float = MIN_VOLUME_CBM

    destination_code, destination_warning = _destination_with_waybill(
        destination_code=destination_code,
        waybill_no=waybill_no,
        row_number=row_number,
    )
    if destination_warning is not None:
        warnings.append(destination_warning)
    courier_warning = _courier_delivery_warning(
        delivery_method=delivery_method,
        note=note,
        row_number=row_number,
    )
    if courier_warning is not None:
        warnings.append(courier_warning)

    return (
        ParsedLine(
            rowNumber=row_number,
            waybillNo=waybill_no,
            fbaNo=_text(raw_json, field_columns, "fbaNo"),
            poNumber=_text(raw_json, field_columns, "poNumber"),
            cartons=cartons_int,
            weight=weight_float,
            volumeCbm=volume_float,
            destinationCode=destination_code,
            deliveryMethod=delivery_method,
            note=note,
            raw_json=raw_json,
        ),
        warnings,
    )


def _destination_summaries(lines: list[ParsedLine]) -> list[DestinationSummary]:
    grouped: dict[str, dict[str, Decimal | int]] = {}

    for line in lines:
        destination = line.destinationCode or ""
        if destination not in grouped:
            grouped[destination] = {
                "totalCartons": 0,
                "totalVolumeCbm": Decimal("0"),
                "lineCount": 0,
            }

        grouped[destination]["totalCartons"] += line.cartons or 0
        grouped[destination]["totalVolumeCbm"] += Decimal(str(line.volumeCbm or 0))
        grouped[destination]["lineCount"] += 1

    return [
        DestinationSummary(
            destinationCode=destination,
            totalCartons=int(values["totalCartons"]),
            totalVolumeCbm=float(values["totalVolumeCbm"]),
            lineCount=int(values["lineCount"]),
        )
        for destination, values in sorted(grouped.items())
    ]


def _extract_container_no(workbook: Any, path: Path) -> tuple[str | None, str | None]:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(max_row=20, values_only=True):
            cells = [_cell_text(value) for value in row]
            for index, cell in enumerate(cells):
                if "柜号" not in cell and "CONTAINER" not in cell.upper():
                    continue

                match = CONTAINER_PATTERN.search(cell.upper())
                if match:
                    return match.group(0), "content"

                for neighbor in cells[index + 1 : index + 3]:
                    match = CONTAINER_PATTERN.search(neighbor.upper())
                    if match:
                        return match.group(0), "content"

    match = CONTAINER_PATTERN.search(path.name.upper())
    if match:
        return match.group(0), "filename"

    return None, None


def _field_columns(headers: tuple[str, ...]) -> dict[str, int]:
    normalized_headers = {_normalize_header(header): index for index, header in enumerate(headers)}
    result: dict[str, int] = {}

    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            index = normalized_headers.get(_normalize_header(alias))
            if index is not None:
                result[field_name] = index
                break

    return result


def _raw_json(headers: tuple[str, ...], row: tuple[Any, ...]) -> dict[str, Any]:
    width = max(len(headers), len(row))
    result: dict[str, Any] = {}

    for index in range(width):
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        value = row[index] if index < len(row) else None
        result[header] = _json_value(value)

    return result


def _unique_headers(header_values: tuple[Any, ...]) -> tuple[str, ...]:
    counts: dict[str, int] = {}
    headers: list[str] = []

    for index, value in enumerate(header_values, start=1):
        header = _cell_text(value) or f"column_{index}"
        count = counts.get(header, 0) + 1
        counts[header] = count
        headers.append(header if count == 1 else f"{header}__{count}")

    return tuple(headers)


def _number(
    raw_json: dict[str, Any],
    field_columns: dict[str, int],
    field_name: str,
    row_number: int,
) -> tuple[Decimal | None, ParseIssue | None]:
    value = _raw_field(raw_json, field_columns, field_name)
    if value in (None, ""):
        return None, None

    try:
        return Decimal(str(value).replace(",", "").strip()), None
    except (InvalidOperation, ValueError):
        return (
            None,
            ParseIssue(
                code="INVALID_NUMBER",
                message=f"Could not parse numeric field {field_name}: {value}",
                row_number=row_number,
                field=field_name,
            ),
        )


def _destination_with_waybill(
    *,
    destination_code: str | None,
    waybill_no: str | None,
    row_number: int,
) -> tuple[str | None, ParseIssue | None]:
    if destination_code is None or not _is_address_destination(destination_code):
        return destination_code, None

    if waybill_no is None:
        return (
            destination_code,
            ParseIssue(
                code="MISSING_WAYBILL_FOR_ADDRESS_DESTINATION",
                message="Commercial or private address destination requires a waybill number.",
                row_number=row_number,
                field="waybillNo",
            ),
        )

    if waybill_no in destination_code:
        return destination_code, None

    return f"{destination_code} / {waybill_no}", None


def _is_address_destination(destination_code: str) -> bool:
    normalized = _normalize_address_destination(destination_code)
    return any(
        term in normalized
        for term in (
            "PRIVATE",
            "PRIVATEADDRESS",
            "COMMERCIAL",
            "COMMERCIALADDRESS",
            "BUSINESSADDRESS",
            "私人",
            "私人地址",
            "商业",
            "商业地址",
            "商業",
            "商業地址",
        )
    )


def _courier_delivery_warning(
    *,
    delivery_method: str | None,
    note: str | None,
    row_number: int,
) -> ParseIssue | None:
    text = " ".join(value for value in (delivery_method, note) if value).upper()
    normalized = re.sub(r"[\s/_-]+", "", text)
    if not text:
        return None

    is_courier_delivery = any(
        term in text or term in normalized for term in COURIER_DELIVERY_TERMS
    )
    has_carrier = any(
        term in text or term in normalized for term in COURIER_CARRIER_TERMS
    )
    if not is_courier_delivery or has_carrier:
        return None

    return ParseIssue(
        code="COURIER_DELIVERY_METHOD_MISSING_CARRIER",
        message=(
            "Courier delivery is requested, but the delivery method or note does "
            "not specify a carrier such as UPS, Purolator, FedEx, Canpar, DHL, "
            "or Canada Post."
        ),
        row_number=row_number,
        field="deliveryMethod",
    )


def _text(raw_json: dict[str, Any], field_columns: dict[str, int], field_name: str) -> str | None:
    value = _raw_field(raw_json, field_columns, field_name)
    text = _cell_text(value)
    if text in {"", "/"}:
        return None
    return text


def _raw_field(raw_json: dict[str, Any], field_columns: dict[str, int], field_name: str) -> Any:
    index = field_columns.get(field_name)
    if index is None:
        return None

    key = tuple(raw_json.keys())[index]
    return raw_json[key]


def _has_detail_signal(line: ParsedLine) -> bool:
    return any(
        value is not None
        for value in (
            line.waybillNo,
            line.fbaNo,
            line.poNumber,
            line.destinationCode,
            line.deliveryMethod,
            line.note,
            line.cartons,
            line.volumeCbm,
        )
    )


def _is_summary_row(raw_json: dict[str, Any]) -> bool:
    values = [_cell_text(value).upper() for value in raw_json.values()]
    if any(value in SUMMARY_MARKERS for value in values):
        return True

    text_values = [value for value in values if value and not _looks_numeric(value)]
    numeric_values = [value for value in values if _looks_numeric(value)]
    return not text_values and len(numeric_values) >= 2


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_cell_text(value) == "" for value in row)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _normalize_header(value: str) -> str:
    text = value.replace("＃", "#").replace("：", ":")
    text = text.replace("³", "3")
    return re.sub(r"\s+", "", text).upper()


def _normalize_address_destination(value: str) -> str:
    return re.sub(r"[\s/_-]+", "", value.upper())


def _looks_numeric(value: str) -> bool:
    try:
        Decimal(value.replace(",", ""))
    except InvalidOperation:
        return False
    return True


def _empty_error_result(
    *,
    path: Path,
    detection_format: FormatType,
    confidence: float,
    metadata: dict[str, Any],
    issue: ParseIssue,
) -> UnloadingPlanParseResult:
    metadata["sourceFile"] = str(path)
    return UnloadingPlanParseResult(
        containerNo=None,
        formatType=detection_format,
        confidence=confidence,
        parserVersion=PARSER_VERSION,
        lines=(),
        destinationSummaries=(),
        warnings=(),
        errors=(issue,),
        rawMetadata=metadata,
    )
