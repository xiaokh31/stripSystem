from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

from worker_python.time_utils import operational_now


SUMMARY_COLUMN_WIDTHS = {
    "A": 35.082,
    "B": 22.6914,
    "C": 27.8398,
    "D": 19.7109,
    "E": 26.6094,
    "F": 31.0312,
    "G": 28.7031,
    "H": 30.4102,
}
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class UnloadingSummaryIssue:
    code: str
    message: str
    containerId: str | None = None
    containerNo: str | None = None
    field: str | None = None


@dataclass(frozen=True)
class UnloadingSummaryExportResult:
    outputPath: Path
    warnings: tuple[UnloadingSummaryIssue, ...]
    errors: tuple[UnloadingSummaryIssue, ...]
    rowCount: int
    sourceContainerCount: int


def write_unloading_summary_workbook(
    *,
    payload: dict[str, Any],
    output_dir: Path,
    generated_at: datetime | None = None,
) -> UnloadingSummaryExportResult:
    generated_at = generated_at or operational_now()
    output_dir.mkdir(parents=True, exist_ok=True)

    month = _required_month(payload.get("month"))
    rows = _rows(payload.get("rows"))
    review_items = _review_items(payload.get("reviewItems"))
    warnings = _warnings(rows, review_items)
    output_path = output_dir / f"monthly-unloading-summary-{month}-{generated_at.strftime('%Y%m%d%H%M%S')}.xlsx"

    workbook = Workbook()
    try:
        summary_sheet = workbook.active
        summary_sheet.title = _sheet_title(month)
        _configure_summary_sheet(summary_sheet)
        _write_summary_rows(summary_sheet, rows)
        _write_review_sheet(workbook, review_items, warnings, month)
        workbook.save(output_path)
    finally:
        workbook.close()

    return UnloadingSummaryExportResult(
        outputPath=output_path,
        warnings=tuple(warnings),
        errors=(),
        rowCount=len(rows),
        sourceContainerCount=len(
            {
                str(row.get("containerId") or row.get("containerNo") or "")
                for row in rows
                if row.get("containerId") or row.get("containerNo")
            }
        ),
    )


def result_payload(result: UnloadingSummaryExportResult) -> dict[str, Any]:
    return {
        "task_status": "GENERATED" if not result.errors else "ERROR",
        "summary_result": {
            "outputPath": str(result.outputPath),
            "mimeType": EXCEL_MIME_TYPE,
            "warnings": [asdict(issue) for issue in result.warnings],
            "errors": [asdict(issue) for issue in result.errors],
            "rowCount": result.rowCount,
            "sourceContainerCount": result.sourceContainerCount,
        },
    }


def _configure_summary_sheet(sheet: Any) -> None:
    for column, width in SUMMARY_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A1"


def _write_summary_rows(sheet: Any, rows: list[dict[str, Any]]) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    first_container_row = True
    excel_row = 1
    last_container_id: str | None = None

    for row in rows:
        container_id = _text(row.get("containerId")) or _text(row.get("containerNo"))
        if (
            last_container_id is not None
            and container_id
            and container_id != last_container_id
        ):
            excel_row += 1
            first_container_row = True

        container_label = ""
        date_label = ""
        if first_container_row:
            sequence = _text(row.get("sequence"))
            container_no = _text(row.get("containerNo"))
            container_label = f"{sequence}、{container_no}" if sequence else container_no
            date_label = _text(row.get("dateBusinessTag"))

        values = [
            container_label,
            date_label,
            _text(row.get("destinationText")),
            _text(row.get("quantityText")),
            _text(row.get("referenceText")),
            _text(row.get("appointmentText")),
            _text(row.get("splitOrVarianceText")),
            _text(row.get("operationNote")),
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(excel_row, index, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        if first_container_row:
            sheet.cell(excel_row, 1).font = Font(bold=True)
            sheet.cell(excel_row, 2).font = Font(bold=True)

        sheet.row_dimensions[excel_row].height = _row_height(values)
        last_container_id = container_id or last_container_id
        first_container_row = False
        excel_row += 1


def _write_review_sheet(
    workbook: Workbook,
    review_items: list[dict[str, Any]],
    warnings: list[UnloadingSummaryIssue],
    month: str,
) -> None:
    sheet = workbook.create_sheet("Review")
    headers = ["Month", "Code", "Container", "Field", "Message"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 90

    records: list[UnloadingSummaryIssue] = warnings[:]
    for item in review_items:
        records.append(
            UnloadingSummaryIssue(
                code=_text(item.get("code")) or "REVIEW_ITEM",
                message=_text(item.get("message")),
                containerId=_text(item.get("containerId")),
                containerNo=_text(item.get("containerNo")),
                field=_text(item.get("field")),
            )
        )

    for index, issue in enumerate(records, start=2):
        values = [
            month,
            issue.code,
            issue.containerNo or issue.containerId or "",
            issue.field or "",
            issue.message,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _warnings(
    rows: list[dict[str, Any]],
    review_items: list[dict[str, Any]],
) -> list[UnloadingSummaryIssue]:
    warnings: list[UnloadingSummaryIssue] = []
    for row in rows:
        container_id = _text(row.get("containerId"))
        container_no = _text(row.get("containerNo"))
        if not _text(row.get("referenceText")):
            warnings.append(
                UnloadingSummaryIssue(
                    code="MISSING_REFERENCE_TEXT",
                    message="Reference, appointment number, shipment, or raw note is missing for this summary row.",
                    containerId=container_id,
                    containerNo=container_no,
                    field="referenceText",
                )
            )
        if not _text(row.get("appointmentText")):
            warnings.append(
                UnloadingSummaryIssue(
                    code="MISSING_APPOINTMENT_TEXT",
                    message="Appointment or unloading time is missing for this summary row.",
                    containerId=container_id,
                    containerNo=container_no,
                    field="appointmentText",
                )
            )
    if review_items:
        warnings.append(
            UnloadingSummaryIssue(
                code="REVIEW_ITEMS_PRESENT",
                message="One or more completed-status containers require office review before monthly assignment.",
            )
        )
    return warnings


def _rows(value: object) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [row for row in value if isinstance(row, dict)]


def _review_items(value: object) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _required_month(value: object) -> str:
    month = _text(value)
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("month must use YYYY-MM format.")
    return month


def _sheet_title(month: str) -> str:
    return f"{int(month[-2:])}月拆柜数据"


def _row_height(values: list[str]) -> float:
    line_count = max(1, *(value.count("\n") + 1 for value in values if value))
    return max(18.0, min(90.0, line_count * 18.0))


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
