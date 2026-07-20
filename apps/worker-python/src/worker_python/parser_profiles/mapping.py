from __future__ import annotations

import re
import time
from collections import OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import regex as bounded_regex  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from worker_python.imports import compute_sha256
from worker_python.pallets.rules import detect_package_type_from_values
from worker_python.parser.workbook_warnings import (
    ignore_openpyxl_conditional_formatting_warning,
)
from worker_python.parser_profiles.contracts import (
    FINGERPRINT_ALGORITHM_VERSION,
    MAPPING_SCHEMA_VERSION,
    BlankOperation,
    CaseOperation,
    CoalesceOperation,
    ColumnSource,
    ConcatenateOperation,
    ConstantSource,
    ContractIssue,
    DivideOperation,
    FieldMapping,
    FieldProvenance,
    IncludePredicate,
    InspectionLimits,
    LookupOperation,
    MappingDefinition,
    MultiplyOperation,
    ParseDecimalOperation,
    ParseIntegerOperation,
    ProfileDestinationSummary,
    ProfileParseResult,
    ProfileParsedLine,
    RegexExtractOperation,
    SkipBlankPredicate,
    SourceReference,
    StopPredicate,
    TrimOperation,
    UnitConversionOperation,
    SkipSummaryPredicate,
    WorkbookInspectionError,
    json_value,
)
from worker_python.parser_profiles.inspection import inspect_workbook
from worker_python.parser_profiles.normalization import normalize_header as _normalize


MIN_VOLUME_CBM = 0.01
MAX_REGEX_INPUT_LENGTH = 10_000
MAX_REGEX_OPERATIONS = 1_000
MAX_REGEX_TOTAL_SECONDS = 2.0
CONTAINER_NUMBER_PATTERN = re.compile(r"\b[A-Z]{4}\d{7}[A-Z]?\b")


class RegexBudgetExceeded(RuntimeError):
    pass


class RegexInputLimitExceeded(RuntimeError):
    pass


@dataclass
class RegexExecutionBudget:
    remaining_operations: int
    deadline: float

    def next_timeout(self) -> float:
        remaining_seconds = self.deadline - time.monotonic()
        if self.remaining_operations <= 0 or remaining_seconds <= 0:
            raise RegexBudgetExceeded
        self.remaining_operations -= 1
        return min(0.05, remaining_seconds)


def execute_mapping(
    path: Path,
    definition: MappingDefinition | dict[str, Any],
    *,
    replay_input_hash: str,
    limits: InspectionLimits | None = None,
) -> ProfileParseResult:
    parsed_definition = (
        definition
        if isinstance(definition, MappingDefinition)
        else MappingDefinition.validate_definition(definition)
    )
    path = Path(path)
    limits = limits or InspectionLimits()
    errors: list[ContractIssue] = []
    warnings: list[ContractIssue] = []
    regex_budget = RegexExecutionBudget(
        remaining_operations=MAX_REGEX_OPERATIONS,
        deadline=time.monotonic() + MAX_REGEX_TOTAL_SECONDS,
    )

    try:
        inspection = inspect_workbook(path, limits=limits)
    except WorkbookInspectionError as exc:
        return _empty_result(
            parsed_definition,
            replay_input_hash,
            path,
            exc.issues[0],
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            formula_workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
    except Exception:
        return _empty_result(
            parsed_definition,
            replay_input_hash,
            path,
            ContractIssue(code="MAPPING_WORKBOOK_READ_FAILED", path="workbook"),
        )

    try:
        sheet_name = _sheet_name(workbook.sheetnames, parsed_definition)
        if sheet_name is None:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(code="MAPPING_SHEET_NOT_FOUND", path="sheet"),
            )
        worksheet = workbook[sheet_name]
        formula_worksheet = formula_workbook[sheet_name]
        worksheet_max_row = int(worksheet.max_row or 0)
        worksheet_max_column = int(worksheet.max_column or 0)
        inspected_sheet = next(
            (sheet for sheet in inspection.sheets if sheet.name == sheet_name),
            None,
        )
        if inspected_sheet is None:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(code="MAPPING_SHEET_LIMIT_EXCEEDED", path="sheet"),
            )
        if worksheet_max_column > limits.maxColumnsPerSheet:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(
                    code="MAPPING_COLUMN_LIMIT_EXCEEDED",
                    path="limits.maxColumnsPerSheet",
                    params={
                        "observed": worksheet_max_column,
                        "limit": limits.maxColumnsPerSheet,
                    },
                ),
            )
        if worksheet_max_row and parsed_definition.header.row > worksheet_max_row:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(code="MAPPING_HEADER_NOT_FOUND", path="header.row"),
            )

        headers, columns = _headers(
            worksheet,
            start_row=parsed_definition.header.row,
            row_count=parsed_definition.header.rowCount,
        )
        required_headers = _required_column_headers(parsed_definition)
        missing_headers = sorted(
            header for header in required_headers if _normalize(header) not in columns
        )
        if missing_headers:
            errors.extend(
                ContractIssue(
                    code="MAPPING_SOURCE_COLUMN_NOT_FOUND",
                    path="fields",
                    rawValue=header,
                )
                for header in missing_headers
            )

        container_no: str | None = None
        metadata_values: dict[str, str | None] = {
            "company": None,
            "poNumber": None,
            "customer": None,
            "clearOrderNo": None,
        }
        workbook_provenance: dict[str, FieldProvenance] = {}
        if parsed_definition.container is not None:
            container_no, provenance, field_issues = _map_field(
                "containerNo",
                parsed_definition.container,
                worksheet=worksheet,
                formula_worksheet=formula_worksheet,
                row_number=None,
                columns=columns,
                sheet_name=sheet_name,
                row_values=None,
                formula_cells=None,
                regex_budget=regex_budget,
            )
            errors.extend(field_issues)
            workbook_provenance["containerNo"] = provenance
            container_no = _optional_text(container_no)

        for field_name, field_mapping in parsed_definition.metadataFields.items():
            value, provenance, field_issues = _map_field(
                field_name,
                field_mapping,
                worksheet=worksheet,
                formula_worksheet=formula_worksheet,
                row_number=None,
                columns=columns,
                sheet_name=sheet_name,
                row_values=None,
                formula_cells=None,
                regex_budget=regex_budget,
            )
            errors.extend(field_issues)
            workbook_provenance[field_name] = provenance
            metadata_values[field_name] = _optional_text(value)

        if container_no is None:
            filename_match = CONTAINER_NUMBER_PATTERN.search(path.name.upper())
            if filename_match:
                container_no = filename_match.group(0)
                workbook_provenance["containerNo"] = FieldProvenance(
                    field="containerNo",
                    sourceRefs=(
                        SourceReference(
                            sheet="<filename>",
                            row=None,
                            column=None,
                            cell=None,
                            rawValue=path.name,
                        ),
                    ),
                    transformChain=("filename_container_fallback",),
                )

        if container_no is None:
            errors.append(
                ContractIssue(
                    code="MISSING_CONTAINER_NO",
                    path="container",
                    field="containerNo",
                )
            )

        lines: list[ProfileParsedLine] = []
        max_end = (
            parsed_definition.dataRange.startRow
            + parsed_definition.dataRange.maxRows
            - 1
        )
        if parsed_definition.dataRange.endRow is not None:
            max_end = min(max_end, parsed_definition.dataRange.endRow)
        observed_max_row = worksheet_max_row
        if observed_max_row > max_end and parsed_definition.dataRange.endRow is None:
            warnings.append(
                ContractIssue(
                    code="MAPPING_ROW_LIMIT_EXCEEDED",
                    path="dataRange.maxRows",
                    params={"observed": observed_max_row, "limitEndRow": max_end},
                )
            )

        iteration_end = min(max_end, observed_max_row) if observed_max_row else max_end
        iteration_count = max(
            iteration_end - parsed_definition.dataRange.startRow + 1, 0
        )
        if iteration_count > limits.maxRowsPerSheet:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(
                    code="MAPPING_ROW_BUDGET_EXCEEDED",
                    path="limits.maxRowsPerSheet",
                    params={
                        "observed": iteration_count,
                        "limit": limits.maxRowsPerSheet,
                    },
                ),
            )
        required_cells = iteration_count * worksheet_max_column
        if required_cells > limits.maxCells:
            return _empty_result(
                parsed_definition,
                replay_input_hash,
                path,
                ContractIssue(
                    code="MAPPING_CELL_LIMIT_EXCEEDED",
                    path="limits.maxCells",
                    params={"observed": required_cells, "limit": limits.maxCells},
                ),
            )
        cached_rows = worksheet.iter_rows(
            min_row=parsed_definition.dataRange.startRow,
            max_row=iteration_end,
            values_only=True,
        )
        formula_rows = formula_worksheet.iter_rows(
            min_row=parsed_definition.dataRange.startRow,
            max_row=iteration_end,
            values_only=False,
        )
        for row_number, (row, formula_row) in enumerate(
            zip(cached_rows, formula_rows, strict=True),
            start=parsed_definition.dataRange.startRow,
        ):
            raw_json = _raw_json(headers, row)
            errors.extend(
                _row_formula_issues(
                    required_headers,
                    row_number=row_number,
                    columns=columns,
                    row_values=row,
                    formula_cells=formula_row,
                    sheet_name=sheet_name,
                )
            )
            if all(value in (None, "") for value in raw_json.values()):
                continue
            action, predicate_issues = _predicate_action(
                parsed_definition,
                raw_json,
                row_number=row_number,
                columns=columns,
                worksheet=worksheet,
                formula_worksheet=formula_worksheet,
                sheet_name=sheet_name,
                row_values=row,
                formula_cells=formula_row,
                regex_budget=regex_budget,
            )
            errors.extend(predicate_issues)
            if action == "stop":
                break
            if action == "skip":
                continue

            values: dict[str, Any] = {}
            provenance_by_field: dict[str, FieldProvenance] = {}
            row_errors: list[ContractIssue] = []
            for field_name, field_mapping in parsed_definition.fields.items():
                value, provenance, field_issues = _map_field(
                    field_name,
                    field_mapping,
                    worksheet=worksheet,
                    formula_worksheet=formula_worksheet,
                    row_number=row_number,
                    columns=columns,
                    sheet_name=sheet_name,
                    row_values=row,
                    formula_cells=formula_row,
                    regex_budget=regex_budget,
                )
                values[field_name] = value
                provenance_by_field[field_name] = provenance
                row_errors.extend(field_issues)
            errors.extend(row_errors)

            values = _canonicalize_values(values, raw_json)
            row_warnings = _canonical_warnings(
                values, row_number, parsed_definition.formatType
            )
            warnings.extend(row_warnings)
            if values.get("volumeCbm") == 0 and _as_int(values.get("cartons")) not in (
                None,
                0,
            ):
                values["volumeCbm"] = MIN_VOLUME_CBM

            if not _has_detail_signal(values):
                continue
            lines.append(
                ProfileParsedLine(
                    rowNumber=row_number,
                    waybillNo=_optional_text(values.get("waybillNo")),
                    fbaNo=_optional_text(values.get("fbaNo")),
                    poNumber=_optional_text(values.get("poNumber")),
                    itemNo=_optional_text(values.get("itemNo")),
                    description=_optional_text(values.get("description")),
                    cartons=_as_int(values.get("cartons")),
                    weight=_as_float(values.get("weight")),
                    volumeCbm=_as_float(values.get("volumeCbm")),
                    destinationCode=_optional_text(values.get("destinationCode")),
                    packageType=_optional_text(values.get("packageType")),
                    deliveryMethod=_optional_text(values.get("deliveryMethod")),
                    note=_optional_text(values.get("note")),
                    totalSkidCount=_as_int(values.get("totalSkidCount")),
                    raw_json=raw_json,
                    provenance=provenance_by_field,
                )
            )
    finally:
        workbook.close()
        formula_workbook.close()

    if parsed_definition.formatType == "BESTAR_RECEIVING":
        warnings.append(
            ContractIssue(
                code="NEED_MANUAL_DESTINATION",
                path="destinationSummaries",
                field="destinationCode",
            )
        )

    return ProfileParseResult(
        containerNo=container_no,
        company=metadata_values["company"],
        poNumber=metadata_values["poNumber"],
        customer=metadata_values["customer"],
        clearOrderNo=metadata_values["clearOrderNo"],
        formatType=parsed_definition.formatType,
        confidence=1.0 if not errors else 0.0,
        lines=tuple(lines),
        destinationSummaries=_destination_summaries(
            lines, parsed_definition.groupBy, parsed_definition.formatType
        ),
        warnings=tuple(_deduplicate_issues(warnings)),
        errors=tuple(_deduplicate_issues(errors)),
        rawMetadata=_metadata(
            parsed_definition, path, replay_input_hash, sheet_name, limits
        ),
        provenance=workbook_provenance,
    )


def _sheet_name(sheet_names: list[str], definition: MappingDefinition) -> str | None:
    if definition.sheet.name is not None:
        return definition.sheet.name if definition.sheet.name in sheet_names else None
    assert definition.sheet.index is not None
    return (
        sheet_names[definition.sheet.index]
        if definition.sheet.index < len(sheet_names)
        else None
    )


def _headers(
    worksheet: Any, *, start_row: int, row_count: int
) -> tuple[tuple[str, ...], dict[str, int]]:
    rows = list(
        worksheet.iter_rows(
            min_row=start_row,
            max_row=start_row + row_count - 1,
            values_only=True,
        )
    )
    width = max((len(row) for row in rows), default=0)
    counts: dict[str, int] = {}
    headers: list[str] = []
    aliases: dict[str, int] = {}
    for column in range(width):
        parts = [
            str(row[column]).strip()
            for row in rows
            if column < len(row) and row[column] not in (None, "")
        ]
        header = " / ".join(parts) if parts else f"column_{column + 1}"
        count = counts.get(header, 0) + 1
        counts[header] = count
        unique = header if count == 1 else f"{header}__{count}"
        headers.append(unique)
        aliases.setdefault(_normalize(header), column)
        aliases.setdefault(_normalize(unique), column)
        for part in parts:
            aliases.setdefault(_normalize(part), column)
    return tuple(headers), aliases


def _required_column_headers(definition: MappingDefinition) -> set[str]:
    headers: set[str] = set()
    mappings = list(definition.fields.values()) + list(
        definition.metadataFields.values()
    )
    if definition.container is not None:
        mappings.append(definition.container)
    for mapping in mappings:
        for source in mapping.sources:
            if isinstance(source, ColumnSource):
                headers.add(source.header)
    for predicate in definition.rowPredicates:
        if isinstance(predicate, SkipBlankPredicate):
            headers.update(predicate.headers)
        elif isinstance(predicate, SkipSummaryPredicate):
            headers.update(predicate.whenBlank)
            headers.update(predicate.whenPresent)
        elif isinstance(predicate.source, ColumnSource):
            headers.add(predicate.source.header)
    return headers


def _map_field(
    field_name: str,
    mapping: FieldMapping,
    *,
    worksheet: Any,
    formula_worksheet: Any,
    row_number: int | None,
    columns: dict[str, int],
    sheet_name: str,
    row_values: tuple[Any, ...] | None,
    formula_cells: tuple[Any, ...] | None,
    regex_budget: RegexExecutionBudget,
) -> tuple[Any, FieldProvenance, list[ContractIssue]]:
    values: list[Any] = []
    refs: list[SourceReference] = []
    issues: list[ContractIssue] = []
    for source in mapping.sources:
        value, ref, source_issue = _source_value(
            source,
            worksheet=worksheet,
            formula_worksheet=formula_worksheet,
            row_number=row_number,
            columns=columns,
            sheet_name=sheet_name,
            row_values=row_values,
            formula_cells=formula_cells,
        )
        values.append(value)
        refs.append(ref)
        if source_issue is not None:
            issues.append(source_issue)

    current: Any = values[0] if values else None
    for operation in mapping.transforms:
        try:
            current = _apply_operation(operation, current, values, regex_budget)
        except RegexInputLimitExceeded:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_INPUT_LIMIT_EXCEEDED",
                    path=f"fields.{field_name}",
                    row=row_number,
                    field=field_name,
                    params={"limit": MAX_REGEX_INPUT_LENGTH},
                )
            )
            current = None
            break
        except RegexBudgetExceeded:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_BUDGET_EXCEEDED",
                    path=f"fields.{field_name}",
                    row=row_number,
                    field=field_name,
                    params={"operation": operation.op},
                )
            )
            current = None
            break
        except TimeoutError:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_TIMEOUT",
                    path=f"fields.{field_name}",
                    row=row_number,
                    field=field_name,
                    params={"operation": operation.op},
                )
            )
            current = None
            break
        except (InvalidOperation, ValueError, TypeError, re.error, IndexError):
            issues.append(
                ContractIssue(
                    code="MAPPING_TRANSFORM_FAILED",
                    path=f"fields.{field_name}",
                    row=row_number,
                    field=field_name,
                    rawValue=json_value(current),
                    params={"operation": operation.op},
                )
            )
            current = None
            break

    return (
        current,
        FieldProvenance(
            field=field_name,
            sourceRefs=tuple(refs),
            transformChain=tuple(operation.op for operation in mapping.transforms),
        ),
        issues,
    )


def _source_value(
    source: Any,
    *,
    worksheet: Any,
    formula_worksheet: Any,
    row_number: int | None,
    columns: dict[str, int],
    sheet_name: str,
    row_values: tuple[Any, ...] | None,
    formula_cells: tuple[Any, ...] | None,
) -> tuple[Any, SourceReference, ContractIssue | None]:
    if isinstance(source, ConstantSource):
        return (
            source.value,
            SourceReference(
                sheet=sheet_name,
                row=None,
                column=None,
                cell=None,
                rawValue=source.value,
            ),
            None,
        )

    if isinstance(source, ColumnSource):
        column_index = columns.get(_normalize(source.header))
        if (
            column_index is None
            or row_number is None
            or row_values is None
            or formula_cells is None
        ):
            return (
                None,
                SourceReference(
                    sheet=sheet_name,
                    row=row_number,
                    column=None,
                    cell=None,
                    rawValue=None,
                ),
                ContractIssue(
                    code="MAPPING_SOURCE_COLUMN_NOT_FOUND",
                    path=source.header,
                    row=row_number,
                ),
            )
        column = column_index + 1
        value = row_values[column_index] if column_index < len(row_values) else None
        formula_cell = (
            formula_cells[column_index]
            if column_index < len(formula_cells)
            else formula_worksheet.cell(row=row_number, column=column)
        )
        issue = _formula_issue(formula_cell, value, sheet_name)
        return (
            value,
            SourceReference(
                sheet=sheet_name,
                row=row_number,
                column=column,
                cell=f"{get_column_letter(column)}{row_number}",
                rawValue=json_value(value),
            ),
            issue,
        )

    cell = source.cell
    value = worksheet[cell].value
    formula_cell = formula_worksheet[cell]
    issue = _formula_issue(formula_cell, value, sheet_name)
    return (
        value,
        SourceReference(
            sheet=sheet_name,
            row=formula_cell.row,
            column=formula_cell.column,
            cell=cell,
            rawValue=json_value(value),
        ),
        issue,
    )


def _formula_issue(
    formula_cell: Any, cached_value: Any, sheet_name: str
) -> ContractIssue | None:
    if formula_cell.data_type != "f" or cached_value is not None:
        return None
    return ContractIssue(
        code="MAPPING_FORMULA_CACHE_MISSING",
        path=formula_cell.coordinate,
        row=formula_cell.row,
        rawValue=formula_cell.value,
        params={"sheet": sheet_name, "cell": formula_cell.coordinate},
    )


def _apply_operation(
    operation: Any,
    current: Any,
    source_values: list[Any],
    regex_budget: RegexExecutionBudget,
) -> Any:
    if isinstance(operation, TrimOperation):
        return current.strip() if isinstance(current, str) else current
    if isinstance(operation, CaseOperation):
        text = str(current) if current is not None else None
        if text is None:
            return None
        return text.upper() if operation.mode == "upper" else text.lower()
    if isinstance(operation, BlankOperation):
        return None if current in operation.values else current
    if isinstance(operation, CoalesceOperation):
        return next((value for value in source_values if value not in (None, "")), None)
    if isinstance(operation, ConcatenateOperation):
        return operation.separator.join(
            str(value) for value in source_values if value not in (None, "")
        )
    if isinstance(operation, ParseDecimalOperation):
        return (
            float(
                _decimal(current, operation.groupSeparator, operation.decimalSeparator)
            )
            if current not in (None, "")
            else None
        )
    if isinstance(operation, ParseIntegerOperation):
        if current in (None, ""):
            return None
        number = _decimal(current, operation.groupSeparator, operation.decimalSeparator)
        if number != number.to_integral_value():
            raise ValueError("not integer")
        return int(number)
    if isinstance(operation, LookupOperation):
        if current is None:
            return None
        key = str(current)
        if operation.caseSensitive:
            return operation.dictionary.get(key, current)
        lookup = {
            candidate.casefold(): value
            for candidate, value in operation.dictionary.items()
        }
        return lookup.get(key.casefold(), current)
    if isinstance(operation, RegexExtractOperation):
        text = str(current or "")
        if len(text) > MAX_REGEX_INPUT_LENGTH:
            raise RegexInputLimitExceeded
        match = bounded_regex.search(
            operation.pattern,
            text,
            timeout=regex_budget.next_timeout(),
        )
        return match.group(operation.group) if match else None
    if isinstance(operation, MultiplyOperation):
        return (
            float(Decimal(str(current)) * operation.factor)
            if current not in (None, "")
            else None
        )
    if isinstance(operation, DivideOperation):
        return (
            float(Decimal(str(current)) / operation.divisor)
            if current not in (None, "")
            else None
        )
    if isinstance(operation, UnitConversionOperation):
        if current in (None, ""):
            return None
        factor = {
            "CBM": Decimal("1"),
            "CUBIC_METRES": Decimal("1"),
            "CUBIC_FEET": Decimal("0.0283168"),
            "CUBIC_INCHES": Decimal("0.000016387064"),
        }[operation.fromUnit]
        return float(Decimal(str(current)) * factor)
    raise ValueError("unknown operation")


def _decimal(value: Any, group_separator: str, decimal_separator: str) -> Decimal:
    text = str(value).strip()
    if group_separator:
        text = text.replace(group_separator, "")
    if decimal_separator != ".":
        text = text.replace(decimal_separator, ".")
    return Decimal(text)


def _predicate_action(
    definition: MappingDefinition,
    raw_json: dict[str, Any],
    *,
    row_number: int,
    columns: dict[str, int],
    worksheet: Any,
    formula_worksheet: Any,
    sheet_name: str,
    row_values: tuple[Any, ...],
    formula_cells: tuple[Any, ...],
    regex_budget: RegexExecutionBudget,
) -> tuple[str, list[ContractIssue]]:
    issues: list[ContractIssue] = []
    for predicate_index, predicate in enumerate(definition.rowPredicates):
        if isinstance(predicate, SkipBlankPredicate):
            values = [
                raw_json.get(_raw_key(raw_json, columns, header))
                for header in predicate.headers
            ]
            if all(value in (None, "") for value in values):
                return "skip", issues
            continue
        if isinstance(predicate, SkipSummaryPredicate):
            blank_values = [
                raw_json.get(_raw_key(raw_json, columns, header))
                for header in predicate.whenBlank
            ]
            present_values = [
                raw_json.get(_raw_key(raw_json, columns, header))
                for header in predicate.whenPresent
            ]
            if all(value in (None, "") for value in blank_values) and any(
                value not in (None, "") for value in present_values
            ):
                return "skip", issues
            continue
        value, _, source_issue = _source_value(
            predicate.source,
            worksheet=worksheet,
            formula_worksheet=formula_worksheet,
            row_number=row_number,
            columns=columns,
            sheet_name=sheet_name,
            row_values=row_values,
            formula_cells=formula_cells,
        )
        if source_issue is not None:
            issues.append(source_issue)
        try:
            matches = _predicate_matches(predicate, value, regex_budget)
        except RegexInputLimitExceeded:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_INPUT_LIMIT_EXCEEDED",
                    path=f"rowPredicates.{predicate_index}",
                    row=row_number,
                    params={"limit": MAX_REGEX_INPUT_LENGTH},
                )
            )
            return "skip", issues
        except RegexBudgetExceeded:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_BUDGET_EXCEEDED",
                    path=f"rowPredicates.{predicate_index}",
                    row=row_number,
                )
            )
            return "skip", issues
        except TimeoutError:
            issues.append(
                ContractIssue(
                    code="MAPPING_REGEX_TIMEOUT",
                    path=f"rowPredicates.{predicate_index}",
                    row=row_number,
                )
            )
            return "skip", issues
        if isinstance(predicate, StopPredicate) and matches:
            return "stop", issues
        if predicate.op == "exclude" and matches:
            return "skip", issues
        if isinstance(predicate, IncludePredicate) and not matches:
            return "skip", issues
    return "include", issues


def _row_formula_issues(
    required_headers: set[str],
    *,
    row_number: int,
    columns: dict[str, int],
    row_values: tuple[Any, ...],
    formula_cells: tuple[Any, ...],
    sheet_name: str,
) -> list[ContractIssue]:
    issues: list[ContractIssue] = []
    checked_columns: set[int] = set()
    for header in sorted(required_headers, key=_normalize):
        column_index = columns.get(_normalize(header))
        if column_index is None or column_index in checked_columns:
            continue
        checked_columns.add(column_index)
        cached_value = (
            row_values[column_index] if column_index < len(row_values) else None
        )
        if column_index >= len(formula_cells):
            continue
        issue = _formula_issue(formula_cells[column_index], cached_value, sheet_name)
        if issue is not None:
            issues.append(issue)
    return issues


def _predicate_matches(
    predicate: Any,
    value: Any,
    regex_budget: RegexExecutionBudget,
) -> bool:
    if predicate.operator == "is_blank":
        return value in (None, "")
    if predicate.operator == "equals":
        return _normalize_value(value) == _normalize_value(predicate.value)
    if predicate.operator == "not_equals":
        return _normalize_value(value) != _normalize_value(predicate.value)
    if predicate.operator == "contains":
        return _normalize_value(predicate.value) in _normalize_value(value)
    if predicate.operator == "in":
        return _normalize_value(value) in {
            _normalize_value(item) for item in predicate.values
        }
    if predicate.operator == "regex":
        text = str(value or "")
        if len(text) > MAX_REGEX_INPUT_LENGTH:
            raise RegexInputLimitExceeded
        return (
            bounded_regex.search(
                predicate.pattern or "",
                text,
                timeout=regex_budget.next_timeout(),
            )
            is not None
        )
    return False


def _raw_key(raw_json: dict[str, Any], columns: dict[str, int], header: str) -> str:
    column = columns.get(_normalize(header))
    return (
        tuple(raw_json)[column] if column is not None and column < len(raw_json) else ""
    )


def _raw_json(headers: tuple[str, ...], row: tuple[Any, ...]) -> dict[str, Any]:
    width = max(len(headers), len(row))
    return {
        headers[index] if index < len(headers) else f"column_{index + 1}": json_value(
            row[index] if index < len(row) else None
        )
        for index in range(width)
    }


def _canonicalize_values(
    values: dict[str, Any], raw_json: dict[str, Any]
) -> dict[str, Any]:
    result = dict(values)
    destination = _optional_text(result.get("destinationCode"))
    waybill = _optional_text(result.get("waybillNo"))
    if (
        destination
        and _normalize(destination) in {"PRIVATEADDRESS", "PRIVATE", "PRIVATEADDRESS/"}
        and waybill
    ):
        result["destinationCode"] = f"{destination.rstrip(' /')} / {waybill}"
    if not _optional_text(result.get("packageType")):
        result["packageType"] = detect_package_type_from_values(
            (*result.values(), *raw_json.values())
        )
    return result


def _canonical_warnings(
    values: dict[str, Any],
    row_number: int,
    format_type: str,
) -> list[ContractIssue]:
    warnings: list[ContractIssue] = []
    destination = _optional_text(values.get("destinationCode"))
    cartons = _as_int(values.get("cartons"))
    volume = _as_float(values.get("volumeCbm"))
    if destination is None and format_type != "BESTAR_RECEIVING":
        warnings.append(
            ContractIssue(
                code="MISSING_DESTINATION", row=row_number, field="destinationCode"
            )
        )
    if cartons in (None, 0):
        warnings.append(
            ContractIssue(code="MISSING_CARTONS", row=row_number, field="cartons")
        )
    if volume is None:
        warnings.append(
            ContractIssue(code="MISSING_VOLUME", row=row_number, field="volumeCbm")
        )
    elif volume == 0 and cartons not in (None, 0):
        warnings.append(
            ContractIssue(
                code="ZERO_VOLUME_WITH_CARTONS", row=row_number, field="volumeCbm"
            )
        )
    return warnings


def _has_detail_signal(values: dict[str, Any]) -> bool:
    return any(value not in (None, "") for value in values.values())


def _destination_summaries(
    lines: list[ProfileParsedLine],
    group_fields: tuple[str, ...],
    format_type: str,
) -> tuple[ProfileDestinationSummary, ...]:
    grouped: OrderedDict[tuple[Any, ...], list[ProfileParsedLine]] = OrderedDict()
    for line in lines:
        key = tuple(getattr(line, field) for field in group_fields)
        grouped.setdefault(key, []).append(line)
    summaries: list[ProfileDestinationSummary] = []
    for group_lines in grouped.values():
        skid_values = [
            line.totalSkidCount
            for line in group_lines
            if line.totalSkidCount is not None
        ]
        summaries.append(
            ProfileDestinationSummary(
                destinationCode=group_lines[0].destinationCode,
                packageType=group_lines[0].packageType,
                totalCartons=sum(line.cartons or 0 for line in group_lines),
                totalVolumeCbm=round(
                    sum(line.volumeCbm or 0 for line in group_lines), 6
                ),
                totalSkidCount=sum(skid_values) if skid_values else None,
                lineCount=len(group_lines),
                status=(
                    "NEED_MANUAL_DESTINATION"
                    if format_type == "BESTAR_RECEIVING"
                    and group_lines[0].destinationCode is None
                    else None
                ),
            )
        )
    return tuple(summaries)


def _metadata(
    definition: MappingDefinition,
    path: Path,
    replay_input_hash: str,
    sheet_name: str,
    limits: InspectionLimits,
) -> dict[str, Any]:
    return {
        "profileVersion": definition.profileVersion,
        "mappingSchemaVersion": MAPPING_SCHEMA_VERSION,
        "fingerprintVersion": FINGERPRINT_ALGORITHM_VERSION,
        "replayInputHash": replay_input_hash,
        "inputSha256": compute_sha256(path),
        "sourceSheet": sheet_name,
        "header": definition.header.model_dump(mode="json"),
        "dataRange": definition.dataRange.model_dump(mode="json"),
        "executionLimits": limits.model_dump(mode="json"),
    }


def _empty_result(
    definition: MappingDefinition,
    replay_input_hash: str,
    path: Path,
    issue: ContractIssue,
) -> ProfileParseResult:
    return ProfileParseResult(
        containerNo=None,
        company=None,
        poNumber=None,
        customer=None,
        clearOrderNo=None,
        formatType=definition.formatType,
        confidence=0.0,
        lines=(),
        destinationSummaries=(),
        warnings=(),
        errors=(issue,),
        rawMetadata={
            "profileVersion": definition.profileVersion,
            "mappingSchemaVersion": MAPPING_SCHEMA_VERSION,
            "fingerprintVersion": FINGERPRINT_ALGORITHM_VERSION,
            "replayInputHash": replay_input_hash,
            "inputSha256": compute_sha256(path) if path.is_file() else None,
        },
        provenance={},
    )


def _normalize_value(value: Any) -> str:
    return str(value or "").strip().casefold()


def _optional_text(value: Any) -> str | None:
    if value in (None, "", "/"):
        return None
    return str(value).strip()


def _as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _deduplicate_issues(issues: list[ContractIssue]) -> list[ContractIssue]:
    seen: set[tuple[str, int | None, str | None, str | None]] = set()
    result: list[ContractIssue] = []
    for issue in issues:
        key = (issue.code, issue.row, issue.field, issue.path)
        if key in seen:
            continue
        seen.add(key)
        result.append(issue)
    return result
