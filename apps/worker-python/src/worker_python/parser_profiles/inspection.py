from __future__ import annotations

import hashlib
import re
import zipfile
from collections.abc import Iterable
from pathlib import Path, PurePosixPath
from typing import Any, Literal
from xml.etree import ElementTree

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from worker_python.parser.workbook_warnings import (
    ignore_openpyxl_conditional_formatting_warning,
)
from worker_python.parser_profiles.contracts import (
    BoundedDimensions,
    ContractIssue,
    DataRangeCandidate,
    HeaderCandidate,
    InspectedCell,
    InspectionLimits,
    SheetInspection,
    WorkbookInspectionError,
    WorkbookInspection,
    json_value,
)


def inspect_workbook(
    path: Path,
    *,
    limits: InspectionLimits | None = None,
) -> WorkbookInspection:
    limits = limits or InspectionLimits()
    path = Path(path)
    workbook_type = _workbook_type(path)
    early_issues: list[ContractIssue] = []

    if not path.is_file():
        raise WorkbookInspectionError(
            (ContractIssue(code="WORKBOOK_NOT_FOUND", path="workbook"),)
        )
    if workbook_type is None or not zipfile.is_zipfile(path):
        raise WorkbookInspectionError(
            (ContractIssue(code="WORKBOOK_TYPE_UNSUPPORTED", path="workbook"),)
        )

    try:
        archive_entry_count, archive_total_bytes, oversize_entries = _archive_budget(
            path, limits.maxArchiveEntryBytes
        )
    except Exception as exc:
        raise WorkbookInspectionError(
            (ContractIssue(code="WORKBOOK_READ_FAILED", path="workbook"),)
        ) from exc
    if archive_entry_count > limits.maxArchiveEntries:
        raise WorkbookInspectionError(
            (
                ContractIssue(
                    code="WORKBOOK_ARCHIVE_ENTRY_COUNT_LIMIT_EXCEEDED",
                    path="workbook.archive",
                    params={
                        "observed": archive_entry_count,
                        "limit": limits.maxArchiveEntries,
                    },
                ),
            )
        )
    if archive_total_bytes > limits.maxArchiveTotalBytes:
        raise WorkbookInspectionError(
            (
                ContractIssue(
                    code="WORKBOOK_ARCHIVE_TOTAL_SIZE_LIMIT_EXCEEDED",
                    path="workbook.archive",
                    params={
                        "observed": archive_total_bytes,
                        "limit": limits.maxArchiveTotalBytes,
                    },
                ),
            )
        )
    if oversize_entries:
        raise WorkbookInspectionError(
            (
                ContractIssue(
                    code="WORKBOOK_ARCHIVE_ENTRY_LIMIT_EXCEEDED",
                    path="workbook.archive",
                    params={"entryCount": len(oversize_entries)},
                ),
            )
        )

    try:
        merged_ranges, merged_range_overflows = _merged_ranges_by_sheet(
            path,
            limits.maxMergedRangesPerSheet,
            limits.maxSheets,
        )
        digest = _sha256(path)
    except Exception as exc:
        raise WorkbookInspectionError(
            (ContractIssue(code="WORKBOOK_READ_FAILED", path="workbook"),)
        ) from exc
    for sheet_name, observed in sorted(merged_range_overflows.items()):
        early_issues.append(
            ContractIssue(
                code="INSPECTION_MERGED_RANGE_LIMIT_EXCEEDED",
                path="sheets.mergedRanges",
                params={
                    "sheet": sheet_name,
                    "observed": observed,
                    "limit": limits.maxMergedRangesPerSheet,
                },
            )
        )

    try:
        with ignore_openpyxl_conditional_formatting_warning():
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            cached_workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
    except Exception as exc:
        raise WorkbookInspectionError(
            (ContractIssue(code="WORKBOOK_READ_FAILED", path="workbook"),)
        ) from exc

    sheet_names = workbook.sheetnames
    if len(sheet_names) > limits.maxSheets:
        early_issues.append(
            ContractIssue(
                code="INSPECTION_SHEET_LIMIT_EXCEEDED",
                path="sheets",
                params={"observed": len(sheet_names), "limit": limits.maxSheets},
            )
        )

    sheets: list[SheetInspection] = []
    total_scanned_cells = 0
    try:
        for index, name in enumerate(sheet_names[: limits.maxSheets]):
            worksheet = workbook[name]
            cached_worksheet = cached_workbook[name]
            declared_max_row = max(int(worksheet.max_row or 0), 0)
            declared_max_column = max(int(worksheet.max_column or 0), 0)
            scanned_rows = min(
                declared_max_row or limits.maxRowsPerSheet, limits.maxRowsPerSheet
            )
            scanned_columns = min(
                declared_max_column or limits.maxColumnsPerSheet,
                limits.maxColumnsPerSheet,
            )

            if declared_max_row > limits.maxRowsPerSheet:
                early_issues.append(
                    ContractIssue(
                        code="INSPECTION_ROW_LIMIT_EXCEEDED",
                        path=f"sheets.{index}.boundedDimensions.maxRow",
                        params={
                            "observed": declared_max_row,
                            "limit": limits.maxRowsPerSheet,
                        },
                    )
                )
            if declared_max_column > limits.maxColumnsPerSheet:
                early_issues.append(
                    ContractIssue(
                        code="INSPECTION_COLUMN_LIMIT_EXCEEDED",
                        path=f"sheets.{index}.boundedDimensions.maxColumn",
                        params={
                            "observed": declared_max_column,
                            "limit": limits.maxColumnsPerSheet,
                        },
                    )
                )

            row_cells: list[tuple[InspectedCell, ...]] = []
            sample_cells: list[InspectedCell] = []
            cell_budget_exhausted = False
            inferred_max_row = 0
            inferred_max_column = 0
            trailing_empty_rows = 0

            if scanned_rows and scanned_columns:
                source_rows = worksheet.iter_rows(
                    min_row=1,
                    max_row=scanned_rows,
                    min_col=1,
                    max_col=scanned_columns,
                )
                cached_rows = cached_worksheet.iter_rows(
                    min_row=1,
                    max_row=scanned_rows,
                    min_col=1,
                    max_col=scanned_columns,
                )
                for source_row, cached_row in zip(
                    source_rows, cached_rows, strict=True
                ):
                    inspected_row: list[InspectedCell] = []
                    source_row_number = len(row_cells) + 1
                    for column_number, (source_cell, cached_cell) in enumerate(
                        zip(source_row, cached_row, strict=True),
                        start=1,
                    ):
                        if total_scanned_cells >= limits.maxCells:
                            cell_budget_exhausted = True
                            break
                        total_scanned_cells += 1
                        inspected = _inspect_cell(
                            source_cell,
                            cached_cell,
                            row=source_row_number,
                            column=column_number,
                        )
                        if inspected.value is not None or inspected.isFormula:
                            inspected_row.append(inspected)
                            inferred_max_row = max(inferred_max_row, inspected.row)
                            inferred_max_column = max(
                                inferred_max_column, inspected.column
                            )
                            if len(sample_cells) < limits.maxSampleCellsPerSheet:
                                sample_cells.append(inspected)
                            if inspected.isFormula and not inspected.hasCachedValue:
                                early_issues.append(
                                    ContractIssue(
                                        code="FORMULA_CACHED_VALUE_MISSING",
                                        path=f"sheets.{index}.sampleCells",
                                        row=inspected.row,
                                        rawValue=inspected.value,
                                        params={"sheet": name, "cell": inspected.cell},
                                    )
                                )
                    row_cells.append(tuple(inspected_row))
                    trailing_empty_rows = (
                        0 if inspected_row else trailing_empty_rows + 1
                    )
                    if cell_budget_exhausted:
                        break
                    if (
                        declared_max_row == 0
                        and inferred_max_row
                        and trailing_empty_rows >= 50
                    ):
                        break

            if cell_budget_exhausted:
                early_issues.append(
                    ContractIssue(
                        code="INSPECTION_CELL_LIMIT_EXCEEDED",
                        path="limits.maxCells",
                        params={"limit": limits.maxCells},
                    )
                )

            header_candidates = _header_candidates(
                row_cells,
                limit=limits.maxHeaderCandidatesPerSheet,
            )
            sheets.append(
                SheetInspection(
                    index=index,
                    name=name,
                    visibility=str(worksheet.sheet_state),
                    boundedDimensions=BoundedDimensions(
                        maxRow=declared_max_row or inferred_max_row,
                        maxColumn=declared_max_column or inferred_max_column,
                        scannedRows=len(row_cells),
                        scannedColumns=scanned_columns,
                    ),
                    mergedRanges=tuple(merged_ranges.get(name, ())),
                    sampleCells=tuple(sample_cells),
                    candidateHeaderAreas=header_candidates,
                    candidateDataRanges=_data_range_candidates(
                        row_cells, header_candidates
                    ),
                )
            )
            if total_scanned_cells >= limits.maxCells:
                break
    finally:
        workbook.close()
        cached_workbook.close()

    return WorkbookInspection(
        workbookType=workbook_type,
        inputSha256=digest,
        sheets=tuple(sheets),
        limits=limits,
        issues=tuple(_deduplicate_issues(early_issues)),
    )


def _inspect_cell(source: Any, cached: Any, *, row: int, column: int) -> InspectedCell:
    is_formula = source.data_type == "f"
    value = source.value
    cached_value = cached.value if is_formula else None
    return InspectedCell(
        row=row,
        column=column,
        cell=f"{get_column_letter(column)}{row}",
        value=json_value(value),
        valueType=_value_type(value, is_formula),
        isFormula=is_formula,
        hasCachedValue=(cached_value is not None) if is_formula else None,
        cachedValue=json_value(cached_value) if is_formula else None,
        cachedValueType=(
            _value_type(cached_value, False)
            if is_formula and cached_value is not None
            else None
        ),
    )


def _value_type(value: Any, is_formula: bool) -> str:
    if is_formula:
        return "formula"
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float):
        return "number"
    if hasattr(value, "isoformat"):
        return "date"
    return "string"


def _header_candidates(
    rows: list[tuple[InspectedCell, ...]],
    *,
    limit: int,
) -> tuple[HeaderCandidate, ...]:
    candidates: list[HeaderCandidate] = []
    for index, row in enumerate(rows):
        strings = [
            cell
            for cell in row
            if cell.valueType == "string" and str(cell.value).strip()
        ]
        if len(strings) < 2:
            continue
        candidates.append(
            HeaderCandidate(
                row=index + 1,
                rowCount=1,
                nonEmptyCells=len(row),
                cells=row,
            )
        )
        if len(candidates) >= limit:
            break

    multi_row: list[HeaderCandidate] = []
    for first, second in zip(candidates, candidates[1:]):
        if second.row == first.row + 1:
            combined = tuple(
                sorted(
                    first.cells + second.cells, key=lambda cell: (cell.row, cell.column)
                )
            )
            multi_row.append(
                HeaderCandidate(
                    row=first.row,
                    rowCount=2,
                    nonEmptyCells=len(combined),
                    cells=combined,
                )
            )
    return tuple((candidates + multi_row)[:limit])


def _data_range_candidates(
    rows: list[tuple[InspectedCell, ...]],
    headers: tuple[HeaderCandidate, ...],
) -> tuple[DataRangeCandidate, ...]:
    if not rows:
        return ()
    start = min(
        (candidate.row + candidate.rowCount for candidate in headers), default=1
    )
    ranges: list[DataRangeCandidate] = []
    range_start: int | None = None
    non_empty = 0
    last_row = start
    for row_number in range(start, len(rows) + 1):
        has_values = bool(rows[row_number - 1])
        if has_values:
            if range_start is None:
                range_start = row_number
            non_empty += 1
            last_row = row_number
        elif range_start is not None:
            ranges.append(
                DataRangeCandidate(
                    startRow=range_start,
                    endRow=last_row,
                    nonEmptyRows=non_empty,
                )
            )
            range_start = None
            non_empty = 0
    if range_start is not None:
        ranges.append(
            DataRangeCandidate(
                startRow=range_start, endRow=last_row, nonEmptyRows=non_empty
            )
        )
    return tuple(ranges[:20])


def _workbook_type(path: Path) -> Literal["OOXML_XLSX", "OOXML_XLSM"] | None:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return "OOXML_XLSX"
    if suffix == ".xlsm":
        return "OOXML_XLSM"
    return None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _archive_budget(
    path: Path, maximum_entry_bytes: int
) -> tuple[int, int, tuple[str, ...]]:
    with zipfile.ZipFile(path) as archive:
        entries = archive.infolist()
        return (
            len(entries),
            sum(info.file_size for info in entries),
            tuple(
                info.filename
                for info in entries
                if info.file_size > maximum_entry_bytes
            ),
        )


def _merged_ranges_by_sheet(
    path: Path, maximum_per_sheet: int, maximum_sheets: int
) -> tuple[dict[str, tuple[str, ...]], dict[str, int]]:
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = {
            relation.attrib["Id"]: relation.attrib["Target"]
            for relation in rel_root
            if relation.attrib.get("Id") and relation.attrib.get("Target")
        }
        result: dict[str, tuple[str, ...]] = {}
        overflows: dict[str, int] = {}
        relationship_key = (
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        for sheet in workbook_root.findall(".//{*}sheet")[:maximum_sheets]:
            sheet_name = sheet.attrib.get("name")
            relationship_id = sheet.attrib.get(relationship_key)
            if not sheet_name or not relationship_id:
                continue
            target = relationship_targets.get(relationship_id)
            if not target:
                continue
            entry = _relationship_entry(target)
            if entry not in names:
                continue
            xml = archive.read(entry)
            refs = re.findall(rb"<mergeCell\s+ref=\"([^\"]+)\"", xml)
            if len(refs) > maximum_per_sheet:
                overflows[sheet_name] = len(refs)
            result[sheet_name] = tuple(
                ref.decode("utf-8") for ref in refs[:maximum_per_sheet]
            )
        return result, overflows


def _relationship_entry(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return str(PurePosixPath("xl") / target)


def _deduplicate_issues(issues: Iterable[ContractIssue]) -> list[ContractIssue]:
    seen: set[tuple[str, str | None, int | None]] = set()
    result: list[ContractIssue] = []
    for issue in issues:
        key = (issue.code, issue.path, issue.row)
        if key in seen:
            continue
        seen.add(key)
        result.append(issue)
    return result
